"""业绩报表与业务洞察模块

提供以下能力：
- 销售预测（基于商机 amount × probability 加权）
- 阶段转化率分析
- 团队/个人业绩对比
- 同比环比趋势
- 自动业务洞察（瓶颈识别、异常检测、最佳表现者）
- Excel 报表导出

权限模型：主任/院长看全部，普通用户看自己负责的数据。
"""
from io import BytesIO
from datetime import datetime, timedelta
from flask import request, jsonify, send_file

from extensions import get_db, token_required

from . import reports_bp


# 阶段分段定义（与 Dashboard 漏斗、Business.vue 概率筛选保持一致）
STAGE_RANGES = [
    {'key': 'lead', 'name': '引导需求', 'min': 0, 'max': 29},
    {'key': 'demo', 'name': '能力展示', 'min': 30, 'max': 59},
    {'key': 'proposal', 'name': '方案确定', 'min': 60, 'max': 79},
    {'key': 'negotiation', 'name': '商务谈判', 'min': 80, 'max': 89},
    {'key': 'contract', 'name': '合同签订', 'min': 90, 'max': 99},
    {'key': 'won', 'name': '销售实现', 'min': 100, 'max': 100},
]


def _get_stage_index(probability):
    """根据 probability 返回阶段索引（0-5），未知返回 None。"""
    if probability is None:
        return None
    prob = int(probability)
    for idx, stage in enumerate(STAGE_RANGES):
        if stage['min'] <= prob <= stage['max']:
            return idx
    return None


def _compute_trend_comparison(db, cursor, role, username, time_range, year=None):
    """计算同比/环比：当前周期 vs 上一周期 5 个核心指标。

    被 dashboard.py 和 reports.py 共用，确保趋势数据一致。
    返回结构：{ total_customers: {current, previous, growth_rate}, ... }
    """
    now = datetime.now()

    if time_range == 'month':
        # 本月 vs 上月
        this_start = now.strftime('%Y-%m-01')
        next_month = now.replace(day=28) + timedelta(days=4)
        prev_start = (next_month.replace(day=1) - timedelta(days=1)).replace(day=1)
        prev_end = (next_month.replace(day=1) - timedelta(days=1)).strftime('%Y-%m-%d')
        cur_cond = "AND created_at >= ? AND created_at <= ?"
        cur_params_start = this_start
        # 上一周期
        prev_cond = "AND created_at >= ? AND created_at <= ?"
        prev_params_start = prev_start.strftime('%Y-%m-%d')
        prev_params_end = prev_end
    elif time_range == 'quarter':
        # 本季度 vs 上季度
        quarter = (now.month - 1) // 3 + 1
        start_month = (quarter - 1) * 3 + 1
        this_start = f"{now.year}-{start_month:02d}-01"
        if quarter == 1:
            prev_start = f"{now.year - 1}-10-01"
            prev_end = f"{now.year - 1}-12-31"
        else:
            prev_start_month = (quarter - 2) * 3 + 1
            prev_start = f"{now.year}-{prev_start_month:02d}-01"
            prev_end = (datetime(now.year, prev_start_month + 2, 28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            prev_end = prev_end.strftime('%Y-%m-%d')
        cur_cond = "AND created_at >= ?"
        cur_params_start = this_start
        prev_cond = "AND created_at >= ? AND created_at <= ?"
        prev_params_start = prev_start
        prev_params_end = prev_end
    elif time_range == 'year':
        # 本年 vs 去年
        chart_year = year if year else now.year
        cur_cond = "AND strftime('%Y', created_at) = ?"
        cur_params_start = str(chart_year)
        prev_cond = "AND strftime('%Y', created_at) = ?"
        prev_params_start = str(chart_year - 1)
        prev_params_end = None
    else:
        # all：本年 vs 去年（兜底）
        chart_year = now.year
        cur_cond = "AND strftime('%Y', created_at) = ?"
        cur_params_start = str(chart_year)
        prev_cond = "AND strftime('%Y', created_at) = ?"
        prev_params_start = str(chart_year - 1)
        prev_params_end = None

    is_admin = role in ('主任', '院长')
    owner_filter = "" if is_admin else "AND owner_id = ?"
    owner_params = [] if is_admin else [username]

    def _count(table, date_field, cond, params, extra_where="", extra_params=None):
        extra_params = extra_params or []
        # customers/business/contracts 用 owner_id；payment_records 需 JOIN contracts 取 owner_id
        if table == 'payment_records':
            sql = (f"SELECT COUNT(*) as total FROM payment_records pr "
                   f"JOIN contracts c ON pr.contract_id = c.id "
                   f"WHERE 1=1 {owner_filter.replace('owner_id', 'c.owner_id')} "
                   f"AND {date_field} IS NOT NULL AND {date_field} != '' {cond.replace(date_field, 'pr.' + date_field) if date_field in cond else cond}")
            # payment_date 用字符串比较
            sql = (f"SELECT COUNT(*) as total FROM payment_records pr "
                   f"JOIN contracts c ON pr.contract_id = c.id "
                   f"WHERE 1=1 {'AND c.owner_id = ?' if not is_admin else ''} "
                   f"AND pr.payment_date IS NOT NULL AND pr.payment_date != '' {cond}")
            cursor.execute(sql, owner_params + params)
        else:
            sql = f"SELECT COUNT(*) as total FROM {table} WHERE 1=1 {owner_filter} {cond}"
            cursor.execute(sql, owner_params + params)
        return cursor.fetchone()['total'] or 0

    def _sum(table, amount_field, date_field, cond, params):
        if table == 'payment_records':
            sql = (f"SELECT COALESCE(SUM(pr.{amount_field}), 0) as total FROM payment_records pr "
                   f"JOIN contracts c ON pr.contract_id = c.id "
                   f"WHERE 1=1 {'AND c.owner_id = ?' if not is_admin else ''} "
                   f"AND pr.{date_field} IS NOT NULL AND pr.{date_field} != '' {cond}")
            cursor.execute(sql, owner_params + params)
        else:
            sql = (f"SELECT COALESCE(SUM({amount_field}), 0) as total FROM {table} "
                   f"WHERE 1=1 {owner_filter} {cond}")
            cursor.execute(sql, owner_params + params)
        return cursor.fetchone()['total'] or 0

    # 当前周期参数
    cur_params = [cur_params_start] if prev_params_end is None and time_range != 'month' else (
        [cur_params_start, now.strftime('%Y-%m-%d')] if time_range == 'month' else [cur_params_start]
    )
    prev_params = [prev_params_start, prev_params_end] if prev_params_end else [prev_params_start]

    # 简化：直接用 strftime 比较 year，或 >= 比较 month/quarter
    if time_range == 'year' or time_range == 'all':
        cur_cond_customers = f"AND strftime('%Y', created_at) = ?"
        prev_cond_customers = f"AND strftime('%Y', created_at) = ?"
        cur_params_c = [cur_params_start]
        prev_params_c = [prev_params_start]

        cur_cond_contracts = f"AND strftime('%Y', sign_date) = ?"
        prev_cond_contracts = f"AND strftime('%Y', sign_date) = ?"

        cur_cond_payments = f"AND strftime('%Y', pr.payment_date) = ?"
        prev_cond_payments = f"AND strftime('%Y', pr.payment_date) = ?"
    else:
        # month/quarter：用 >= 比较
        if time_range == 'month':
            cur_end = now.strftime('%Y-%m-%d')
        else:
            quarter = (now.month - 1) // 3 + 1
            end_month = quarter * 3
            cur_end = (datetime(now.year, end_month, 28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            cur_end = cur_end.strftime('%Y-%m-%d')

        cur_cond_customers = f"AND created_at >= ? AND created_at <= ?"
        prev_cond_customers = f"AND created_at >= ? AND created_at <= ?"
        cur_params_c = [cur_params_start, cur_end]
        prev_params_c = [prev_params_start, prev_params_end]

        cur_cond_contracts = f"AND sign_date >= ? AND sign_date <= ?"
        prev_cond_contracts = f"AND sign_date >= ? AND sign_date <= ?"

        cur_cond_payments = f"AND pr.payment_date >= ? AND pr.payment_date <= ?"
        prev_cond_payments = f"AND pr.payment_date >= ? AND pr.payment_date <= ?"

    # 1. 客户数
    cur_customers = _count('customers', 'created_at', cur_cond_customers, cur_params_c)
    prev_customers = _count('customers', 'created_at', prev_cond_customers, prev_params_c)

    # 2. 商机数（仅 active）
    def _count_business(cond, params):
        sql = f"SELECT COUNT(*) as total FROM business WHERE status = 'active' {owner_filter} {cond}"
        cursor.execute(sql, owner_params + params)
        return cursor.fetchone()['total'] or 0

    cur_business = _count_business(cur_cond_customers.replace('created_at', 'created_at'), cur_params_c)
    prev_business = _count_business(prev_cond_customers.replace('created_at', 'created_at'), prev_params_c)

    # 3. 合同数
    cur_contracts = _count('contracts', 'sign_date', cur_cond_contracts, cur_params_c)
    prev_contracts = _count('contracts', 'sign_date', prev_cond_contracts, prev_params_c)

    # 4. 合同额
    def _sum_contracts(cond, params):
        sql = f"SELECT COALESCE(SUM(total_amt), 0) as total FROM contracts WHERE 1=1 {owner_filter} {cond.replace('created_at', 'sign_date') if 'created_at' in cond else cond}"
        cursor.execute(sql, owner_params + params)
        return cursor.fetchone()['total'] or 0

    cur_contract_amt = _sum_contracts(cur_cond_contracts, cur_params_c)
    prev_contract_amt = _sum_contracts(prev_cond_contracts, prev_params_c)

    # 5. 回款额
    def _sum_payments(cond, params):
        sql = (f"SELECT COALESCE(SUM(pr.amount), 0) as total FROM payment_records pr "
               f"JOIN contracts c ON pr.contract_id = c.id "
               f"WHERE 1=1 {'AND c.owner_id = ?' if not is_admin else ''} "
               f"AND pr.payment_date IS NOT NULL AND pr.payment_date != '' {cond}")
        cursor.execute(sql, owner_params + params)
        return cursor.fetchone()['total'] or 0

    cur_payment_amt = _sum_payments(cur_cond_payments, cur_params_c)
    prev_payment_amt = _sum_payments(prev_cond_payments, prev_params_c)

    def _growth(cur, prev):
        if prev == 0:
            return 100.0 if cur > 0 else 0.0
        return round((cur - prev) / prev * 100, 1)

    return {
        'total_customers': {'current': cur_customers, 'previous': prev_customers, 'growth_rate': _growth(cur_customers, prev_customers)},
        'total_business': {'current': cur_business, 'previous': prev_business, 'growth_rate': _growth(cur_business, prev_business)},
        'total_contracts': {'current': cur_contracts, 'previous': prev_contracts, 'growth_rate': _growth(cur_contracts, prev_contracts)},
        'contracts_amount': {'current': cur_contract_amt, 'previous': prev_contract_amt, 'growth_rate': _growth(cur_contract_amt, prev_contract_amt)},
        'total_payments': {'current': cur_payment_amt, 'previous': prev_payment_amt, 'growth_rate': _growth(cur_payment_amt, prev_payment_amt)},
    }


@reports_bp.route('/api/reports/forecast', methods=['GET'])
@token_required
def get_forecast():
    """销售预测：按 predict_date 月份分组，加权预测金额 = SUM(amount × probability / 100)。"""
    payload = request.current_user
    username = payload['username']
    role = payload['role']
    year = request.args.get('year', type=int) or datetime.now().year

    db = get_db()
    cursor = db.cursor()

    is_admin = role in ('主任', '院长')
    owner_filter = "" if is_admin else "AND b.owner_id = ?"
    owner_params = [] if is_admin else [username]

    # 按月份统计预测金额（predict_date 格式为 YYYY-MM 或 YYYY-MM-DD，取前 7 位归一化）
    cursor.execute(f"""
        SELECT substr(predict_date, 1, 7) as month,
               SUM(amount * COALESCE(probability, 0) / 100.0) as forecast,
               SUM(CASE WHEN probability >= 100 THEN amount ELSE 0 END) as signed
        FROM business b
        WHERE b.status = 'active'
          AND b.predict_date IS NOT NULL AND b.predict_date != ''
          AND substr(predict_date, 1, 4) = ?
          {owner_filter}
        GROUP BY substr(predict_date, 1, 7)
        ORDER BY month
    """, [str(year)] + owner_params)

    month_map = {}
    for row in cursor.fetchall():
        month_map[row['month']] = {'forecast': row['forecast'] or 0, 'signed': row['signed'] or 0}

    months = []
    forecast_data = []
    signed_data = []
    for m in range(1, 13):
        month_str = f"{year}-{m:02d}"
        key = month_str
        data = month_map.get(key, {'forecast': 0, 'signed': 0})
        months.append(f"{m}月")
        forecast_data.append(round(data['forecast'] / 10000, 1))  # 元→万元
        signed_data.append(round(data['signed'] / 10000, 1))

    total_forecast = round(sum(forecast_data), 1)
    total_signed = round(sum(signed_data), 1)
    accuracy = round(total_signed / total_forecast * 100, 1) if total_forecast > 0 else 0

    return jsonify({
        'code': 200,
        'message': 'success',
        'data': {
            'months': months,
            'forecast_data': forecast_data,
            'signed_data': signed_data,
            'total_forecast': total_forecast,
            'total_signed': total_signed,
            'accuracy': accuracy
        }
    })


@reports_bp.route('/api/reports/conversion', methods=['GET'])
@token_required
def get_conversion():
    """阶段转化率：6 个阶段的商机数量、金额、转化率、流失率。"""
    payload = request.current_user
    username = payload['username']
    role = payload['role']

    db = get_db()
    cursor = db.cursor()

    is_admin = role in ('主任', '院长')
    owner_filter = "" if is_admin else "AND owner_id = ?"
    owner_params = [] if is_admin else [username]

    cursor.execute(f"""
        SELECT probability, amount FROM business
        WHERE status = 'active' {owner_filter}
    """, owner_params)

    stage_counts = [0] * len(STAGE_RANGES)
    stage_amounts = [0.0] * len(STAGE_RANGES)

    for row in cursor.fetchall():
        idx = _get_stage_index(row['probability'])
        if idx is not None:
            stage_counts[idx] += 1
            stage_amounts[idx] += (row['amount'] or 0)

    stages = []
    for i, stage in enumerate(STAGE_RANGES):
        count = stage_counts[i]
        amount = stage_amounts[i]
        # 转化率：累计到此阶段的商机数 / 第一阶段商机数 × 100%
        # 漏斗转化：当前阶段数 / 上一阶段数
        if i == 0:
            conversion_rate = 100.0
            drop_rate = 0.0
        else:
            prev_count = stage_counts[i - 1]
            conversion_rate = round(count / prev_count * 100, 1) if prev_count > 0 else 0.0
            drop_rate = round(100 - conversion_rate, 1) if prev_count > 0 else 100.0

        stages.append({
            'key': stage['key'],
            'name': stage['name'],
            'count': count,
            'amount': round(amount / 10000, 1),  # 万元
            'conversion_rate': conversion_rate,
            'drop_rate': drop_rate
        })

    return jsonify({
        'code': 200,
        'message': 'success',
        'data': {'stages': stages}
    })


@reports_bp.route('/api/reports/team-performance', methods=['GET'])
@token_required
def get_team_performance():
    """团队业绩对比：按负责人统计商机/合同/回款/预测/胜率。仅主任/院长可调用。"""
    payload = request.current_user
    role = payload['role']

    if role not in ('主任', '院长'):
        return jsonify({'code': 403, 'message': '无权查看团队业绩对比', 'data': None})

    db = get_db()
    cursor = db.cursor()

    # 商机统计（按 owner_id）
    cursor.execute("""
        SELECT u.name as owner_name, u.role,
               COUNT(b.id) as business_count,
               COALESCE(SUM(b.amount), 0) as business_amount,
               SUM(CASE WHEN b.probability >= 100 THEN 1 ELSE 0 END) as won_count,
               SUM(CASE WHEN b.probability >= 100 THEN b.amount ELSE 0 END) as won_amount,
               SUM(CASE WHEN b.probability < 100 THEN b.amount * COALESCE(b.probability, 0) / 100.0 ELSE 0 END) as forecast_amount
        FROM users u
        LEFT JOIN business b ON u.username = b.owner_id AND b.status = 'active'
        WHERE u.status = '在职'
        GROUP BY u.username, u.name, u.role
        ORDER BY business_amount DESC
    """)
    business_stats = {row['owner_name']: dict(row) for row in cursor.fetchall()}

    # 合同统计
    cursor.execute("""
        SELECT u.name as owner_name,
               COUNT(c.id) as contract_count,
               COALESCE(SUM(c.total_amt), 0) as contract_amount
        FROM users u
        LEFT JOIN contracts c ON u.username = c.owner_id
        GROUP BY u.username, u.name
    """)
    contract_stats = {row['owner_name']: dict(row) for row in cursor.fetchall()}

    # 回款统计
    cursor.execute("""
        SELECT u.name as owner_name,
               COALESCE(SUM(pr.amount), 0) as payment_amount
        FROM users u
        LEFT JOIN contracts c ON u.username = c.owner_id
        LEFT JOIN payment_records pr ON c.id = pr.contract_id
        GROUP BY u.username, u.name
    """)
    payment_stats = {row['owner_name']: dict(row) for row in cursor.fetchall()}

    # 合并：取所有在职用户（有数据的优先）
    all_names = set(business_stats.keys()) | set(contract_stats.keys()) | set(payment_stats.keys())
    members = []
    for name in all_names:
        if not name:
            continue
        b = business_stats.get(name, {})
        c = contract_stats.get(name, {})
        p = payment_stats.get(name, {})
        b_count = b.get('business_count', 0) or 0
        won_count = b.get('won_count', 0) or 0
        win_rate = round(won_count / b_count * 100, 1) if b_count > 0 else 0.0
        members.append({
            'name': name,
            'role': b.get('role', ''),
            'business_count': b_count,
            'business_amount': round((b.get('business_amount', 0) or 0) / 10000, 1),
            'contract_count': c.get('contract_count', 0) or 0,
            'contract_amount': round((c.get('contract_amount', 0) or 0) / 10000, 1),
            'payment_amount': round((p.get('payment_amount', 0) or 0) / 10000, 1),
            'forecast_amount': round((b.get('forecast_amount', 0) or 0) / 10000, 1),
            'won_count': won_count,
            'win_rate': win_rate
        })

    # 按商机金额降序
    members.sort(key=lambda x: x['business_amount'], reverse=True)

    return jsonify({
        'code': 200,
        'message': 'success',
        'data': {'members': members}
    })


@reports_bp.route('/api/reports/trend-comparison', methods=['GET'])
@token_required
def get_trend_comparison():
    """同比环比：当前周期 vs 上一周期 5 个核心指标的增长率。"""
    payload = request.current_user
    username = payload['username']
    role = payload['role']
    time_range = request.args.get('time_range', 'month')
    year = request.args.get('year', type=int)

    db = get_db()
    cursor = db.cursor()

    trends = _compute_trend_comparison(db, cursor, role, username, time_range, year)

    return jsonify({
        'code': 200,
        'message': 'success',
        'data': {'metrics': trends}
    })


@reports_bp.route('/api/reports/insights', methods=['GET'])
@token_required
def get_insights():
    """自动业务洞察：瓶颈识别、异常检测、最佳表现者、风险预警、机会提示。"""
    payload = request.current_user
    username = payload['username']
    role = payload['role']
    year = request.args.get('year', type=int) or datetime.now().year

    db = get_db()
    cursor = db.cursor()
    insights = []

    is_admin = role in ('主任', '院长')
    owner_filter = "" if is_admin else "AND owner_id = ?"
    owner_params = [] if is_admin else [username]

    # 1. 瓶颈识别：转化率最低的阶段
    cursor.execute(f"""
        SELECT probability, amount FROM business
        WHERE status = 'active' {owner_filter}
    """, owner_params)
    stage_counts = [0] * len(STAGE_RANGES)
    for row in cursor.fetchall():
        idx = _get_stage_index(row['probability'])
        if idx is not None:
            stage_counts[idx] += 1

    # 找出流失率最大的阶段跳转（i → i+1 的转化率最低）
    bottleneck_stage = None
    min_conversion = 100.0
    for i in range(len(STAGE_RANGES) - 1):
        cur_count = stage_counts[i]
        next_count = stage_counts[i + 1]
        if cur_count > 0:
            conv = next_count / cur_count * 100
            if conv < min_conversion:
                min_conversion = conv
                bottleneck_stage = (STAGE_RANGES[i]['name'], STAGE_RANGES[i + 1]['name'], cur_count, next_count, conv)

    if bottleneck_stage and min_conversion < 50:
        insights.append({
            'type': 'bottleneck',
            'severity': 'high' if min_conversion < 30 else 'medium',
            'title': f"阶段瓶颈：{bottleneck_stage[0]} → {bottleneck_stage[1]}",
            'detail': f"从「{bottleneck_stage[0]}」到「{bottleneck_stage[1]}」转化率仅 {bottleneck_stage[4]:.1f}%（{bottleneck_stage[2]}个 → {bottleneck_stage[3]}个）",
            'suggestion': f"建议加强「{bottleneck_stage[1]}」阶段的跟进力度，复盘「{bottleneck_stage[0]}」阶段流失原因"
        })

    # 2. 异常检测：环比下滑超 20% 的指标
    trends = _compute_trend_comparison(db, cursor, role, username, 'month')
    metric_names = {
        'total_customers': '新增客户数',
        'total_business': '新增商机数',
        'total_contracts': '新增合同数',
        'contracts_amount': '合同总额',
        'total_payments': '回款总额'
    }
    for key, label in metric_names.items():
        growth = trends[key]['growth_rate']
        if growth < -20:
            insights.append({
                'type': 'anomaly',
                'severity': 'high' if growth < -40 else 'medium',
                'title': f"环比异常：{label}下滑 {abs(growth)}%",
                'detail': f"本月{label} {trends[key]['current']}，上月 {trends[key]['previous']}，环比 {growth}%",
                'suggestion': f"关注{label}下滑原因，是否需调整销售策略或加大客户开发力度"
            })

    # 3. 最佳表现者（仅主任/院长可见团队视角）
    if is_admin:
        cursor.execute("""
            SELECT u.name as owner_name,
                   COALESCE(SUM(c.total_amt), 0) as contract_amount,
                   COUNT(c.id) as contract_count
            FROM users u
            LEFT JOIN contracts c ON u.username = c.owner_id
            WHERE u.status = '在职' AND strftime('%Y', c.sign_date) = ?
            GROUP BY u.username, u.name
            ORDER BY contract_amount DESC
            LIMIT 1
        """, (str(year),))
        top_row = cursor.fetchone()
        if top_row and top_row['contract_amount'] > 0:
            insights.append({
                'type': 'top_performer',
                'severity': 'info',
                'title': f"业绩之星：{top_row['owner_name']}",
                'detail': f"{top_row['owner_name']} 本年合同总额 {top_row['contract_amount'] / 10000:.1f} 万元，签约 {top_row['contract_count']} 份",
                'suggestion': "可分享其销售经验，或在团队内推广其方法论"
            })

    # 4. 风险预警：预测金额高但转化率低的负责人
    if is_admin:
        cursor.execute("""
            SELECT u.name as owner_name,
                   COUNT(b.id) as biz_count,
                   SUM(b.amount * COALESCE(b.probability, 0) / 100.0) as forecast,
                   SUM(CASE WHEN b.probability >= 100 THEN 1 ELSE 0 END) as won_count
            FROM users u
            JOIN business b ON u.username = b.owner_id AND b.status = 'active'
            WHERE u.status = '在职'
            GROUP BY u.username, u.name
            HAVING forecast > 0
            ORDER BY forecast DESC
        """)
        for row in cursor.fetchall():
            win_rate = row['won_count'] / row['biz_count'] * 100 if row['biz_count'] > 0 else 0
            if row['forecast'] > 1000000 and win_rate < 30:  # 预测金额>100万且胜率<30%
                insights.append({
                    'type': 'risk_alert',
                    'severity': 'medium',
                    'title': f"风险预警：{row['owner_name']} 预测金额高但转化率低",
                    'detail': f"{row['owner_name']} 加权预测 {row['forecast'] / 10000:.1f} 万元，但当前胜率仅 {win_rate:.1f}%（{row['biz_count']}个商机，已签 {row['won_count']}个）",
                    'suggestion': "建议复核其商机概率评估的合理性，或提供销售支持"
                })

    # 5. 机会提示：高概率（≥80%）但未签的商机数量
    cursor.execute(f"""
        SELECT COUNT(*) as total, COALESCE(SUM(amount), 0) as amount
        FROM business
        WHERE status = 'active' AND probability >= 80 AND probability < 100
        {owner_filter}
    """, owner_params)
    opp_row = cursor.fetchone()
    if opp_row and opp_row['total'] > 0:
        insights.append({
            'type': 'opportunity',
            'severity': 'info',
            'title': f"成交机会：{opp_row['total']} 个高概率商机待签约",
            'detail': f"有 {opp_row['total']} 个商机概率 ≥80% 但尚未签约，潜在金额 {opp_row['amount'] / 10000:.1f} 万元",
            'suggestion': "建议优先推进这些商机，争取尽快签约落袋"
        })

    # 按严重程度排序
    severity_order = {'high': 0, 'medium': 1, 'info': 2}
    insights.sort(key=lambda x: severity_order.get(x['severity'], 3))

    return jsonify({
        'code': 200,
        'message': 'success',
        'data': {'insights': insights, 'total': len(insights)}
    })


@reports_bp.route('/api/reports/export', methods=['GET'])
@token_required
def export_report():
    """Excel 报表导出：多 Sheet 工作簿（销售预测、阶段转化、团队业绩、洞察摘要）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    payload = request.current_user
    username = payload['username']
    role = payload['role']
    year = request.args.get('year', type=int) or datetime.now().year

    is_admin = role in ('主任', '院长')
    owner_filter = "" if is_admin else "AND owner_id = ?"
    owner_filter_contract = "" if is_admin else "AND c.owner_id = ?"
    owner_params = [] if is_admin else [username]

    db = get_db()
    cursor = db.cursor()

    wb = Workbook()

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    def _style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    # Sheet 1: 销售预测
    ws1 = wb.active
    ws1.title = "销售预测"
    ws1.append(["月份", "加权预测金额(万元)", "已签约金额(万元)"])
    cursor.execute(f"""
        SELECT substr(predict_date, 1, 7) as month,
               SUM(amount * COALESCE(probability, 0) / 100.0) as forecast,
               SUM(CASE WHEN probability >= 100 THEN amount ELSE 0 END) as signed
        FROM business
        WHERE status = 'active' AND predict_date IS NOT NULL AND predict_date != ''
              AND substr(predict_date, 1, 4) = ?
              {owner_filter}
        GROUP BY substr(predict_date, 1, 7)
        ORDER BY month
    """, [str(year)] + owner_params)
    forecast_map = {row['month']: (row['forecast'] or 0, row['signed'] or 0) for row in cursor.fetchall()}
    total_forecast = 0
    total_signed = 0
    for m in range(1, 13):
        month_str = f"{year}-{m:02d}"
        f_val, s_val = forecast_map.get(month_str, (0, 0))
        ws1.append([f"{m}月", round(f_val / 10000, 2), round(s_val / 10000, 2)])
        total_forecast += f_val
        total_signed += s_val
    ws1.append(["合计", round(total_forecast / 10000, 2), round(total_signed / 10000, 2)])
    _style_header(ws1)

    # Sheet 2: 阶段转化
    ws2 = wb.create_sheet("阶段转化")
    ws2.append(["阶段", "商机数", "商机金额(万元)", "转化率(%)", "流失率(%)"])
    cursor.execute(f"""
        SELECT probability, amount FROM business
        WHERE status = 'active' {owner_filter}
    """, owner_params)
    stage_counts = [0] * len(STAGE_RANGES)
    stage_amounts = [0.0] * len(STAGE_RANGES)
    for row in cursor.fetchall():
        idx = _get_stage_index(row['probability'])
        if idx is not None:
            stage_counts[idx] += 1
            stage_amounts[idx] += (row['amount'] or 0)
    for i, stage in enumerate(STAGE_RANGES):
        if i == 0:
            conv, drop = 100.0, 0.0
        else:
            prev = stage_counts[i - 1]
            conv = round(stage_counts[i] / prev * 100, 1) if prev > 0 else 0.0
            drop = round(100 - conv, 1) if prev > 0 else 100.0
        ws2.append([stage['name'], stage_counts[i], round(stage_amounts[i] / 10000, 2), conv, drop])
    _style_header(ws2)

    # Sheet 3: 团队业绩（仅主任/院长）
    if is_admin:
        ws3 = wb.create_sheet("团队业绩")
        ws3.append(["负责人", "角色", "商机数", "商机金额(万元)", "合同数", "合同金额(万元)", "回款金额(万元)", "加权预测(万元)", "胜率(%)"])
        cursor.execute("""
            SELECT u.name as owner_name, u.role,
                   COUNT(b.id) as biz_count,
                   COALESCE(SUM(b.amount), 0) as biz_amount,
                   SUM(CASE WHEN b.probability >= 100 THEN 1 ELSE 0 END) as won_count,
                   SUM(CASE WHEN b.probability < 100 THEN b.amount * COALESCE(b.probability, 0) / 100.0 ELSE 0 END) as forecast
            FROM users u
            LEFT JOIN business b ON u.username = b.owner_id AND b.status = 'active'
            WHERE u.status = '在职'
            GROUP BY u.username, u.name, u.role
            ORDER BY biz_amount DESC
        """)
        biz_map = {row['owner_name']: dict(row) for row in cursor.fetchall()}

        cursor.execute("""
            SELECT u.name as owner_name,
                   COUNT(c.id) as contract_count,
                   COALESCE(SUM(c.total_amt), 0) as contract_amount
            FROM users u
            LEFT JOIN contracts c ON u.username = c.owner_id
            GROUP BY u.username, u.name
        """)
        contract_map = {row['owner_name']: dict(row) for row in cursor.fetchall()}

        cursor.execute("""
            SELECT u.name as owner_name,
                   COALESCE(SUM(pr.amount), 0) as payment_amount
            FROM users u
            LEFT JOIN contracts c ON u.username = c.owner_id
            LEFT JOIN payment_records pr ON c.id = pr.contract_id
            GROUP BY u.username, u.name
        """)
        payment_map = {row['owner_name']: dict(row) for row in cursor.fetchall()}

        for name in set(biz_map.keys()) | set(contract_map.keys()) | set(payment_map.keys()):
            if not name:
                continue
            b = biz_map.get(name, {})
            c = contract_map.get(name, {})
            p = payment_map.get(name, {})
            b_count = b.get('biz_count', 0) or 0
            won = b.get('won_count', 0) or 0
            win_rate = round(won / b_count * 100, 1) if b_count > 0 else 0
            ws3.append([
                name, b.get('role', ''),
                b_count, round((b.get('biz_amount', 0) or 0) / 10000, 2),
                c.get('contract_count', 0) or 0, round((c.get('contract_amount', 0) or 0) / 10000, 2),
                round((p.get('payment_amount', 0) or 0) / 10000, 2),
                round((b.get('forecast', 0) or 0) / 10000, 2),
                win_rate
            ])
        _style_header(ws3)

    # Sheet 4: 洞察摘要
    ws4 = wb.create_sheet("洞察摘要")
    ws4.append(["类型", "严重程度", "标题", "详情", "建议"])
    # 复用 insights 接口逻辑
    insights_data = _generate_insights(cursor, role, username, is_admin, owner_filter, owner_params, year)
    for ins in insights_data:
        type_labels = {
            'bottleneck': '阶段瓶颈',
            'anomaly': '环比异常',
            'top_performer': '业绩之星',
            'risk_alert': '风险预警',
            'opportunity': '成交机会'
        }
        ws4.append([
            type_labels.get(ins['type'], ins['type']),
            ins['severity'],
            ins['title'],
            ins['detail'],
            ins['suggestion']
        ])
    _style_header(ws4)

    # 调整列宽
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 4, 50)
            ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"销售业绩报表_{year}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def _generate_insights(cursor, role, username, is_admin, owner_filter, owner_params, year):
    """复用 insights 接口的核心逻辑，供 export 调用。"""
    insights = []

    # 1. 瓶颈识别
    cursor.execute(f"""
        SELECT probability FROM business
        WHERE status = 'active' {owner_filter}
    """, owner_params)
    stage_counts = [0] * len(STAGE_RANGES)
    for row in cursor.fetchall():
        idx = _get_stage_index(row['probability'])
        if idx is not None:
            stage_counts[idx] += 1

    bottleneck_stage = None
    min_conversion = 100.0
    for i in range(len(STAGE_RANGES) - 1):
        cur_count = stage_counts[i]
        next_count = stage_counts[i + 1]
        if cur_count > 0:
            conv = next_count / cur_count * 100
            if conv < min_conversion:
                min_conversion = conv
                bottleneck_stage = (STAGE_RANGES[i]['name'], STAGE_RANGES[i + 1]['name'], cur_count, next_count, conv)

    if bottleneck_stage and min_conversion < 50:
        insights.append({
            'type': 'bottleneck',
            'severity': 'high' if min_conversion < 30 else 'medium',
            'title': f"阶段瓶颈：{bottleneck_stage[0]} → {bottleneck_stage[1]}",
            'detail': f"从「{bottleneck_stage[0]}」到「{bottleneck_stage[1]}」转化率仅 {bottleneck_stage[4]:.1f}%（{bottleneck_stage[2]}个 → {bottleneck_stage[3]}个）",
            'suggestion': f"建议加强「{bottleneck_stage[1]}」阶段的跟进力度，复盘「{bottleneck_stage[0]}」阶段流失原因"
        })

    # 2. 异常检测
    trends = _compute_trend_comparison(get_db(), cursor, role, username, 'month')
    metric_names = {
        'total_customers': '新增客户数',
        'total_business': '新增商机数',
        'total_contracts': '新增合同数',
        'contracts_amount': '合同总额',
        'total_payments': '回款总额'
    }
    for key, label in metric_names.items():
        growth = trends[key]['growth_rate']
        if growth < -20:
            insights.append({
                'type': 'anomaly',
                'severity': 'high' if growth < -40 else 'medium',
                'title': f"环比异常：{label}下滑 {abs(growth)}%",
                'detail': f"本月{label} {trends[key]['current']}，上月 {trends[key]['previous']}，环比 {growth}%",
                'suggestion': f"关注{label}下滑原因，是否需调整销售策略或加大客户开发力度"
            })

    # 3. 最佳表现者
    if is_admin:
        cursor.execute("""
            SELECT u.name as owner_name,
                   COALESCE(SUM(c.total_amt), 0) as contract_amount,
                   COUNT(c.id) as contract_count
            FROM users u
            LEFT JOIN contracts c ON u.username = c.owner_id
            WHERE u.status = '在职' AND strftime('%Y', c.sign_date) = ?
            GROUP BY u.username, u.name
            ORDER BY contract_amount DESC
            LIMIT 1
        """, (str(year),))
        top_row = cursor.fetchone()
        if top_row and top_row['contract_amount'] > 0:
            insights.append({
                'type': 'top_performer',
                'severity': 'info',
                'title': f"业绩之星：{top_row['owner_name']}",
                'detail': f"{top_row['owner_name']} 本年合同总额 {top_row['contract_amount'] / 10000:.1f} 万元，签约 {top_row['contract_count']} 份",
                'suggestion': "可分享其销售经验，或在团队内推广其方法论"
            })

    # 4. 风险预警
    if is_admin:
        cursor.execute("""
            SELECT u.name as owner_name,
                   COUNT(b.id) as biz_count,
                   SUM(b.amount * COALESCE(b.probability, 0) / 100.0) as forecast,
                   SUM(CASE WHEN b.probability >= 100 THEN 1 ELSE 0 END) as won_count
            FROM users u
            JOIN business b ON u.username = b.owner_id AND b.status = 'active'
            WHERE u.status = '在职'
            GROUP BY u.username, u.name
            HAVING forecast > 0
            ORDER BY forecast DESC
        """)
        for row in cursor.fetchall():
            win_rate = row['won_count'] / row['biz_count'] * 100 if row['biz_count'] > 0 else 0
            if row['forecast'] > 1000000 and win_rate < 30:
                insights.append({
                    'type': 'risk_alert',
                    'severity': 'medium',
                    'title': f"风险预警：{row['owner_name']} 预测金额高但转化率低",
                    'detail': f"{row['owner_name']} 加权预测 {row['forecast'] / 10000:.1f} 万元，但当前胜率仅 {win_rate:.1f}%（{row['biz_count']}个商机，已签 {row['won_count']}个）",
                    'suggestion': "建议复核其商机概率评估的合理性，或提供销售支持"
                })

    # 5. 机会提示
    cursor.execute(f"""
        SELECT COUNT(*) as total, COALESCE(SUM(amount), 0) as amount
        FROM business
        WHERE status = 'active' AND probability >= 80 AND probability < 100
        {owner_filter}
    """, owner_params)
    opp_row = cursor.fetchone()
    if opp_row and opp_row['total'] > 0:
        insights.append({
            'type': 'opportunity',
            'severity': 'info',
            'title': f"成交机会：{opp_row['total']} 个高概率商机待签约",
            'detail': f"有 {opp_row['total']} 个商机概率 ≥80% 但尚未签约，潜在金额 {opp_row['amount'] / 10000:.1f} 万元",
            'suggestion': "建议优先推进这些商机，争取尽快签约落袋"
        })

    severity_order = {'high': 0, 'medium': 1, 'info': 2}
    insights.sort(key=lambda x: severity_order.get(x['severity'], 3))
    return insights


def register_routes(app):
    app.register_blueprint(reports_bp)
