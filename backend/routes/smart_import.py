"""
智能导入模块：上传 Excel/CSV 文件，自动识别模块、映射字段、校验数据，
支持人工调整后批量导入。

流程：
1. POST /api/smart-import/parse  — 上传文件，返回解析预览
2. POST /api/smart-import/execute — 确认后执行导入（按依赖顺序）

支持模块：客户、商机、合同、回款、拜访、线索
"""
import io
import json
import sqlite3
from datetime import datetime

import openpyxl

from flask import Blueprint, request, jsonify

from extensions import get_db, record_operation_log, DB_PATH, token_required

smart_import_bp = Blueprint('smart_import', __name__)

# ============================================================
# 字段映射字典：每个模块的 [字段名] → [中文关键词列表]
# 导入时按关键词模糊匹配 Excel 列头
# ============================================================
MODULE_FIELDS = {
    'customers': {
        'company':       ['客户名称', '公司名称', '公司', '企业名称', '单位名称', '单位'],
        'name':          ['联系人', '联系人姓名', '姓名', '客户姓名'],
        'phone':         ['电话', '手机', '联系方式', '电话号码', '手机号'],
        'email':         ['邮箱', '电子邮件', 'email', 'E-mail'],
        'level':         ['等级', '客户等级', '级别'],
        'source':        ['来源', '客户来源'],
        'industry':      ['行业', '所属行业'],
        'region':        ['地区', '区域', '所在地区', '省份'],
        'owner_id':      ['负责人', '归属人', '业务员', '销售', '跟进人'],
    },
    'business': {
        'title':         ['商机标题', '商机名称', '项目名称', '标题', '项目名'],
        'amount':        ['商机金额', '金额', '预算', '预计金额', '项目金额'],
        'stage':         ['阶段', '商机阶段', '当前阶段'],
        'predict_date':  ['预计成交日期', '预计成交', '预计日期', '成交日期'],
        'owner_id':      ['负责人', '归属人', '销售', '业务员'],
        'probability':   ['概率', '成交概率', '胜率'],
        'source':        ['来源', '商机来源'],
        'industry':      ['行业', '所属行业'],
        'region':        ['地区', '区域'],
        'note':          ['备注', '说明', '备注说明'],
    },
    'contracts': {
        'contract_no':       ['合同编号', '合同号', '编号'],
        'contract_name':     ['合同名称', '合同名'],
        'total_amt':         ['合同总额', '合同金额', '总金额', '金额', '合同总金额'],
        'sign_date':         ['签约日期', '签订日期', '合同日期', '签署日期'],
        'party_a':           ['甲方', '签约方', '客户名称', '客户'],
        'owner_id':           ['负责人', '归属人', '业务员'],
        'status':            ['状态', '合同状态'],
        'classification':    ['密级', '保密等级'],
        'business_type':     ['业态', '业务类型'],
        'acceptance_date':   ['验收日期', '验收时间'],
        'expected_income_date': ['预计收入日期', '预计回款日期', '预计到账日期'],
        'note':              ['备注', '说明'],
    },
    'payment_records': {
        'contract_no':   ['合同编号', '合同号'],
        'contract_name': ['合同名称', '合同名'],
        'payment_date':  ['回款日期', '日期', '到账日期', '付款日期'],
        'amount':        ['回款金额', '金额', '款额', '付款金额'],
        'note':          ['备注', '说明'],
    },
    'visits': {
        'plan_date':      ['拜访日期', '计划日期', '日期'],
        'plan_time':      ['拜访时间', '时间'],
        'purpose':        ['拜访目的', '目的'],
        'visitor_id':     ['拜访人', '负责人', '执行人'],
        'location':       ['地点', '拜访地点', '地址'],
        'contact_person': ['联系人'],
        'result':         ['结果', '拜访结果'],
        'notes':          ['备注', '说明'],
        'work_content':   ['工作内容', '内容'],
        'work_type':      ['工作类型', '类型'],
    },
    'scraped_leads': {
        'company':          ['招标单位', '公司', '公司名称', '单位', '客户名称', '企业名称',
                             '采购人', '招标人', '建设单位', '业主', '甲方', '委托方',
                             '采购方', '中标单位', '合作方', '供方', '供应商', '采购单位',
                             '投标单位', '竞标单位', '发包方', '使用单位', '用户单位',
                             '项目业主', '项目单位', '合作企业', '对方单位', '名称'],
        'contact_name':     ['招标联系人', '联系人', '姓名', '联系人姓名', '对接人', '项目联系人'],
        'phone':            ['招标联系电话', '招标电话', '电话', '手机', '联系电话', '联系方式',
                             '联系人电话', '手机'],
        'email':            ['邮箱', '电子邮件', 'email', 'E-mail', '联系邮箱'],
        'industry':         ['行业', '所属行业', '行业类别', '行业分类'],
        'region':           ['地区', '区域', '所在地区', '省份', '城市', '所在地'],
        'source':           ['来源', '线索来源', '信息来源', '渠道'],
        'opportunity_name': ['标题', '商机名称', '项目名称', '项目名', '招标标题',
                             '商机标题', '商机', '商机描述', '项目标题', '招标名称', '项目'],
        'tender_no':        ['招标编号', '标段编号', '项目编号', '采购编号', '招标号',
                             '项目代号', '招标编码'],
        'budget':           ['招标估价', '招标预算', '预算', '预算金额', '项目预算', '项目估算',
                             '招标控制价', '最高限价', '预算价', '概算金额'],
        'deadline':         ['投标截止时间', '投标截止日期', '截止日期', '截止时间', '报名截止',
                             '投标截止', '递交截止', '开标时间', '开标日期'],
        'publish_date':     ['发布时间', '发布日期', '公告日期', '公告时间', '发布',
                             '公示日期', '公示时间'],
        'agency':           ['招标代理机构', '代理机构', '招标代理', '代理', '招标机构'],
        'agency_phone':     ['招标代理机构联系电话', '代理机构电话', '代理机构联系电话', '代理联系电话',
                             '代理电话', '招标代理电话'],
        'link':             ['详情链接', '链接', '获取链接', '原文链接', '详情网址',
                             '原文网址', '跳转链接', '点击链接'],
        'remark':           ['备注', '说明', '备注说明', '补充说明', '其他'],
    },
    'enterprises': {
        'name':                    ['企业名称', '公司名称', '单位名称', '名称', '公司', '单位'],
        'established_date':        ['成立时间', '成立日期', '注册时间', '注册日期'],
        'location':                ['公司位置', '地址', '所在地', '公司地址', '位置'],
        'personnel_size':          ['人员规模', '员工人数', '人数', '规模'],
        'brief':                   ['单位简介', '公司简介', '简介', '企业简介'],
        'registered_capital':      ['注册资本', '注册资金', '资金'],
        'business_scope':          ['业务范围', '经营范围', '主营业务'],
        'main_qualifications':     ['主要资质', '资质', '资质等级', '资质信息'],
        'main_products':           ['主要产品和方案', '主要产品', '产品方案', '产品', '方案'],
        'relationship_status':     ['关系状态', '合作状态', '关系'],
        'cooperation_opportunities': ['合作机会点', '合作机会', '机会点', '机会'],
        'website':                 ['单位网址', '网址', '官网', '网站'],
        'contact_person':          ['联系人', '联系人姓名', '姓名'],
        'contact_info':            ['联系方式', '电话', '手机', '联系电话'],
    },
}

# 每个模块的必填字段
REQUIRED_FIELDS = {
    'customers': ['company'],
    'business': ['title'],
    'contracts': ['contract_no', 'contract_name', 'total_amt'],
    'payment_records': ['contract_no', 'payment_date', 'amount'],
    'visits': ['plan_date'],
    'scraped_leads': [],
    'enterprises': ['name'],
}

# 金额字段（单位为万元时需 ×10000 转元）
AMOUNT_FIELDS = {
    'business': ['amount'],
    'contracts': ['total_amt'],
    'payment_records': ['amount'],
}

# 日期字段（尝试多种格式解析）
DATE_FIELDS = {
    'customers': ['last_follow'],
    'business': ['predict_date'],
    'contracts': ['sign_date', 'acceptance_date', 'expected_income_date'],
    'payment_records': ['payment_date'],
    'visits': ['plan_date', 'actual_date'],
    'scraped_leads': ['deadline', 'publish_date'],
    'enterprises': ['established_date'],
}

# 链接类字段（优先用单元格超链接 URL 替换显示值，避免“点击查看原文”等无意义文本）
LINK_FIELDS = {'link', 'website'}

# 模块中文显示名
MODULE_NAMES = {
    'customers': '客户',
    'business': '商机',
    'contracts': '合同',
    'payment_records': '回款',
    'visits': '拜访',
    'scraped_leads': '线索',
    'enterprises': '企业信息库',
}

# 导入依赖顺序（先导入被依赖的表）
IMPORT_ORDER = ['customers', 'scraped_leads', 'enterprises', 'business', 'contracts', 'payment_records', 'visits']


def _match_keyword(header, keywords):
    """检查列头是否匹配某个关键词（包含匹配，大小写不敏感）。

    短关键词（≤2字符）使用精确匹配，避免 '公司' 匹配到 '公司地址' 等误匹配。
    """
    header_lower = str(header).strip().lower()
    for kw in keywords:
        kw_lower = kw.lower()
        if len(kw_lower) <= 2:
            # 短关键词：精确匹配
            if header_lower == kw_lower:
                return True
        else:
            # 长关键词：包含匹配
            if kw_lower in header_lower:
                return True
    return False


def detect_module(headers):
    """根据列头智能识别模块，返回 (模块名, 是否歧义, 各模块得分)。"""
    scores = {}
    for module, fields in MODULE_FIELDS.items():
        score = 0
        for field, keywords in fields.items():
            for header in headers:
                if _match_keyword(header, keywords):
                    score += 1
                    break
        scores[module] = score

    sorted_scores = sorted(scores.items(), key=lambda x: -x[1])
    best_module, best_score = sorted_scores[0]

    if best_score == 0:
        return None, False, scores

    # 第二名得分接近第一名（比值 > 0.7）则视为歧义
    second_score = sorted_scores[1][1] if len(sorted_scores) > 1 else 0
    is_ambiguous = second_score > 0 and (second_score / best_score) > 0.7

    return best_module, is_ambiguous, scores


def map_fields(headers, module):
    """将 Excel 列头映射到模块字段，返回 {列索引: 字段名} 和未映射的列。"""
    field_map = {}  # col_index → field_name
    unmapped = []   # 未匹配的列头

    if module not in MODULE_FIELDS:
        return field_map, unmapped

    fields = MODULE_FIELDS[module]
    for col_idx, header in enumerate(headers):
        matched = False
        for field, keywords in fields.items():
            if _match_keyword(header, keywords):
                # 避免一个字段被多列重复映射
                if field not in field_map.values():
                    field_map[col_idx] = field
                    matched = True
                    break
        if not matched:
            unmapped.append({'col_index': col_idx, 'header': header})

    return field_map, unmapped


def _parse_date(val):
    """尝试多种格式解析日期，返回 YYYY-MM-DD 字符串。"""
    if not val:
        return None
    val = str(val).strip()
    # 已经是标准格式
    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日', '%m/%d/%Y', '%d/%m/%Y']:
        try:
            return datetime.strptime(val, fmt).strftime('%Y-%m-%d')
        except (ValueError, TypeError):
            continue
    # Excel 数字日期
    try:
        import xlrd
        return datetime(*xlrd.xldate_as_tuple(float(val), 0)).strftime('%Y-%m-%d')
    except (ValueError, TypeError, ImportError):
        pass
    return val  # 无法解析，原样返回


def _parse_amount(val, is_wan=True):
    """解析金额，is_wan=True 时 ×10000 转元。"""
    if val is None or val == '':
        return 0
    try:
        num = float(str(val).replace(',', '').replace('，', '').replace('万', '').strip())
        return num * 10000 if is_wan else num
    except (ValueError, TypeError):
        return 0


def _extract_hyperlink_formula(val):
    """从 =HYPERLINK(url, text) 公式中提取 url，无法提取返回 None。
    支持 =HYPERLINK("https://...","点击查看原文") 等写法（参数分隔符兼容逗号/分号）。
    """
    if not isinstance(val, str):
        return None
    import re
    m = re.match(r'\s*=\s*HYPERLINK\s*\(\s*', val, re.IGNORECASE)
    if not m:
        return None
    rest = val[m.end():].strip()
    if not rest:
        return None
    # 双引号包裹的 URL
    if rest[0] == '"':
        end = rest.find('"', 1)
        if end > 0:
            return rest[1:end].strip()
    # 单引号包裹的 URL
    if rest[0] == "'":
        end = rest.find("'", 1)
        if end > 0:
            return rest[1:end].strip()
    # 无引号：取到分隔符（, ; ）为止
    m2 = re.match(r'([^,;)]+)', rest)
    if m2:
        return m2.group(1).strip()
    return None


def _resolve_owner_id(cursor, val):
    """将负责人姓名/用户名解析为 username。"""
    if not val:
        return None
    val = str(val).strip()
    cursor.execute("SELECT username FROM users WHERE username = ? OR name = ?", (val, val))
    row = cursor.fetchone()
    return row['username'] if row else val  # 找不到就用原值


def _resolve_customer(cursor, company_name, owner_id=None):
    """通过公司名匹配客户，返回 cust_id；找不到则返回 None。"""
    if not company_name:
        return None
    cursor.execute("SELECT id FROM customers WHERE company = ? ORDER BY id LIMIT 1", (str(company_name).strip(),))
    row = cursor.fetchone()
    return row['id'] if row else None


def _resolve_contract(cursor, contract_no, contract_name=None):
    """通过合同编号或名称匹配合同，返回 contract_id。"""
    if contract_no:
        cursor.execute("SELECT id FROM contracts WHERE contract_no = ?", (str(contract_no).strip(),))
        row = cursor.fetchone()
        if row:
            return row['id']
    if contract_name:
        cursor.execute("SELECT id FROM contracts WHERE contract_name = ? ORDER BY id LIMIT 1", (str(contract_name).strip(),))
        row = cursor.fetchone()
        if row:
            return row['id']
    return None


# ============================================================
# API: 解析文件
# ============================================================
@smart_import_bp.route('/api/smart-import/parse', methods=['POST'])
@token_required
def smart_import_parse():
    """上传 Excel/CSV 文件，自动识别模块和映射字段，返回预览数据。"""
    file = request.files.get('file')
    if not file:
        return jsonify({'code': 400, 'message': '请选择文件', 'data': None})

    filename = file.filename or ''
    try:
        if filename.endswith('.csv') or file.content_type == 'text/csv':
            # CSV 用 utf-8-sig 兼容 BOM
            content = file.read().decode('utf-8-sig', errors='ignore')
            import csv
            reader = csv.reader(io.StringIO(content))
            sheets = [{'name': 'Sheet1', 'headers': next(reader, []), 'rows': [r for r in reader if any(c.strip() for c in r)], 'hyperlinks': []}]
        else:
            # Excel：读两个版本，data_only=True 拿计算值，data_only=False 拿公式（用于提取 =HYPERLINK）
            buf = file.read()
            wb_vals = openpyxl.load_workbook(io.BytesIO(buf), data_only=True)
            wb_fml = openpyxl.load_workbook(io.BytesIO(buf), data_only=False)
            sheets = []
            for ws_val, ws_fml in zip(wb_vals.worksheets, wb_fml.worksheets):
                rows_val = list(ws_val.iter_rows())
                rows_fml = list(ws_fml.iter_rows())
                if not rows_fml:
                    continue
                headers = [str(c.value).strip() if c.value is not None else '' for c in rows_fml[0]]
                data_rows = []
                hyperlink_rows = []  # 与 data_rows 平行，存每行每列的超链接 URL（或 None）
                for row_fml, row_val in zip(rows_fml[1:], rows_val[1:]):
                    cells = []
                    hlinks = []
                    has_data = False
                    for c_fml, c_val in zip(row_fml, row_val):
                        # 显示值（用计算值版本，处理普通公式；标题等字段保持文本不被替换）
                        v = c_val.value
                        # 提取超链接 URL（单元格级 hyperlink 或 =HYPERLINK 公式），单独存不替换显示值
                        hl = None
                        if c_fml.hyperlink is not None:
                            hl = c_fml.hyperlink.target
                        else:
                            hl = _extract_hyperlink_formula(c_fml.value)
                        if v is not None and str(v).strip():
                            has_data = True
                        cells.append(str(v).strip() if v is not None else '')
                        hlinks.append(hl.strip() if isinstance(hl, str) and hl.strip() else None)
                    if has_data:
                        data_rows.append(cells)
                        hyperlink_rows.append(hlinks)
                if data_rows:
                    sheets.append({'name': ws_fml.title, 'headers': headers, 'rows': data_rows, 'hyperlinks': hyperlink_rows})
            wb_vals.close()
            wb_fml.close()
    except Exception as e:
        return jsonify({'code': 500, 'message': f'文件解析失败: {str(e)}', 'data': None})

    if not sheets:
        return jsonify({'code': 400, 'message': '文件没有有效数据', 'data': None})

    # 解析每个 sheet
    results = []
    for sheet in sheets:
        headers = sheet['headers']
        raw_rows = sheet['rows']

        # 智能识别模块
        module, is_ambiguous, scores = detect_module(headers)

        # 字段映射（最佳模块）
        field_map = {}
        unmapped = []
        if module:
            field_map, unmapped = map_fields(headers, module)

        # 为所有模块（不仅仅是得分>0）预计算字段映射，供前端切换时直接使用
        all_field_maps = {}
        for mod_name in MODULE_NAMES:
            fm, _ = map_fields(headers, mod_name)
            all_field_maps[mod_name] = {str(k): v for k, v in fm.items()}

        # 构建数据行
        parsed_rows = []
        raw_hyperlinks = sheet.get('hyperlinks', [])
        for row_idx, raw_row in enumerate(raw_rows, 1):
            row_hlinks = raw_hyperlinks[row_idx - 1] if row_idx - 1 < len(raw_hyperlinks) else []
            row_data = {}
            for col_idx, field in field_map.items():
                val = raw_row[col_idx] if col_idx < len(raw_row) else ''
                if field in LINK_FIELDS and col_idx < len(row_hlinks) and row_hlinks[col_idx]:
                    val = row_hlinks[col_idx]
                row_data[field] = val

            # 线索模块：用户已筛选数据，默认全部有效
            errors = []
            if module != 'scraped_leads':
                required = REQUIRED_FIELDS.get(module, [])
                for rf in required:
                    if not row_data.get(rf):
                        errors.append(f'缺少必填字段: {rf}')

            parsed_rows.append({
                'row_index': row_idx,
                'raw': raw_row,
                'data': row_data,
                'valid': len(errors) == 0,
                'errors': errors,
            })

        # 所有可选模块及其得分（供人工调整）—— 始终包含全部模块，即使得分为 0
        all_modules_list = sorted(scores.items(), key=lambda x: -x[1]) if scores else []
        # 补充没有出现在 scores 中的模块（score=0），确保用户始终能手动选择
        existing_modules = {m for m, _ in all_modules_list}
        for mod in MODULE_NAMES:
            if mod not in existing_modules:
                all_modules_list.append((mod, 0))

        results.append({
            'sheet_name': sheet['name'],
            'headers': headers,
            'detected_module': module,
            'is_ambiguous': is_ambiguous,
            'module_scores': [{'module': m, 'name': MODULE_NAMES.get(m, m), 'score': s}
                              for m, s in all_modules_list],
            'field_map': {str(k): v for k, v in field_map.items()},
            'all_field_maps': all_field_maps,
            'unmapped_columns': unmapped,
            'total_rows': len(raw_rows),
            'valid_count': sum(1 for r in parsed_rows if r['valid']),
            'invalid_count': sum(1 for r in parsed_rows if not r['valid']),
            'rows': parsed_rows,  # 返回全部解析行，避免确认导入时丢数据
        })

    return jsonify({
        'code': 200,
        'message': '解析成功',
        'data': {
            'filename': filename,
            'sheets': results,
            'module_names': MODULE_NAMES,
        }
    })


# ============================================================
# API: 执行导入
# ============================================================
@smart_import_bp.route('/api/smart-import/execute', methods=['POST'])
@token_required
def smart_import_execute():
    """确认后执行导入，按依赖顺序处理各 sheet。"""
    payload = request.get_json(silent=True) or {}
    sheets = payload.get('sheets', [])
    is_wan = payload.get('is_wan', True)  # 金额单位：True=万元, False=元
    mode = payload.get('mode', '')  # 'update_link' 表示仅更新已有线索链接（不新建）

    # 从 token 获取用户名（token_required 已校验有效性）
    username = request.current_user['username']
    db = get_db()
    cursor = db.cursor()

    # 按依赖顺序排列 sheets
    def sheet_order_key(sheet):
        module = sheet.get('module', '')
        return IMPORT_ORDER.index(module) if module in IMPORT_ORDER else 99

    sheets_sorted = sorted(sheets, key=sheet_order_key)

    total_success = 0
    total_fail = 0
    all_results = []

    for sheet in sheets_sorted:
        module = sheet.get('module')
        rows = sheet.get('rows', [])
        field_map = sheet.get('field_map', {})

        if not module or module not in MODULE_FIELDS:
            all_results.append({
                'sheet_name': sheet.get('sheet_name', ''),
                'module': module,
                'success_count': 0,
                'fail_count': len(rows),
                'results': [{'row_index': r.get('row_index', 0), 'success': False,
                             'message': '未识别模块，跳过'} for r in rows],
            })
            continue

        success_count = 0
        fail_count = 0
        row_results = []

        for row in rows:
            if not row.get('selected', True):  # 前端可取消勾选某些行
                continue
            row_data = row.get('data', {})
            row_index = row.get('row_index', 0)

            try:
                if mode == 'update_link' and module == 'scraped_leads':
                    matched, msg = _update_lead_link(cursor, row_data)
                    if matched > 0:
                        db.commit()
                        success_count += 1
                        row_results.append({'row_index': row_index, 'success': True, 'message': msg})
                    else:
                        db.rollback()
                        fail_count += 1
                        row_results.append({'row_index': row_index, 'success': False, 'message': msg})
                else:
                    _import_one_row(cursor, module, row_data, field_map, is_wan, username)
                    db.commit()
                    success_count += 1
                    row_results.append({'row_index': row_index, 'success': True, 'message': '导入成功'})
            except Exception as e:
                db.rollback()
                fail_count += 1
                row_results.append({'row_index': row_index, 'success': False, 'message': str(e)})

        total_success += success_count
        total_fail += fail_count
        all_results.append({
            'sheet_name': sheet.get('sheet_name', ''),
            'module': module,
            'module_name': MODULE_NAMES.get(module, module),
            'success_count': success_count,
            'fail_count': fail_count,
            'results': row_results,
        })

    try:
        record_operation_log(username, '导入', '智能导入',
                             f'成功 {total_success} 条，失败 {total_fail} 条')
    except Exception:
        pass

    return jsonify({
        'code': 200,
        'message': f'导入完成：成功 {total_success} 条，失败 {total_fail} 条',
        'data': {
            'total_success': total_success,
            'total_fail': total_fail,
            'sheets': all_results,
        }
    })


def _update_lead_link(cursor, row_data):
    """更新已有线索的 link 字段（按招标单位 + 商机名称匹配）。
    返回 (matched_count, message)。
    """
    company = (str(row_data.get('company') or '')).strip()
    opportunity_name = (str(row_data.get('opportunity_name') or '')).strip()
    link = (str(row_data.get('link') or '')).strip()
    if not link:
        return 0, '链接为空，跳过'
    if not company:
        return 0, '招标单位为空，无法匹配'
    # 优先 company + opportunity_name 精确匹配；无 opportunity_name 则单按 company 匹配
    if opportunity_name:
        cursor.execute(
            "SELECT id FROM scraped_leads WHERE company=? AND opportunity_name=?",
            (company, opportunity_name)
        )
    else:
        cursor.execute(
            "SELECT id FROM scraped_leads WHERE company=? "
            "AND (opportunity_name IS NULL OR opportunity_name='' OR opportunity_name=company)",
            (company,)
        )
    ids = [r['id'] for r in cursor.fetchall()]
    if not ids:
        return 0, f'未找到匹配线索（{company}）'
    for cid in ids:
        cursor.execute("UPDATE scraped_leads SET link=? WHERE id=?", (link, cid))
    return len(ids), f'已更新 {len(ids)} 条线索的链接'


def _import_one_row(cursor, module, row_data, field_map, is_wan, username):
    """导入单行数据到指定模块。"""
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 处理金额字段
    for amt_field in AMOUNT_FIELDS.get(module, []):
        if amt_field in row_data:
            row_data[amt_field] = _parse_amount(row_data[amt_field], is_wan)

    # 处理日期字段
    for date_field in DATE_FIELDS.get(module, []):
        if date_field in row_data:
            row_data[date_field] = _parse_date(row_data[date_field])

    # 解析负责人
    if 'owner_id' in row_data:
        row_data['owner_id'] = _resolve_owner_id(cursor, row_data['owner_id'])
    if 'visitor_id' in row_data:
        row_data['visitor_id'] = _resolve_owner_id(cursor, row_data['visitor_id'])

    if module == 'customers':
        cursor.execute("""
            INSERT INTO customers (company, name, phone, email, level, source, industry, region, owner_id, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (row_data.get('company', ''), row_data.get('name', ''), row_data.get('phone', ''),
              row_data.get('email', ''), row_data.get('level', ''), row_data.get('source', ''),
              row_data.get('industry', ''), row_data.get('region', ''),
              row_data.get('owner_id') or username, now))

    elif module == 'business':
        cust_id = _resolve_customer(cursor, row_data.get('company', ''))
        cursor.execute("""
            INSERT INTO business (cust_id, title, amount, stage, predict_date, owner_id,
                                  probability, source, industry, region, note, created_at, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active')
        """, (cust_id, row_data.get('title', ''), row_data.get('amount', 0),
              row_data.get('stage', ''), row_data.get('predict_date', ''),
              row_data.get('owner_id') or username, row_data.get('probability', 0),
              row_data.get('source', ''), row_data.get('industry', ''),
              row_data.get('region', ''), row_data.get('note', ''), now))

    elif module == 'contracts':
        # 尝试匹配客户和商机
        cust_id = _resolve_customer(cursor, row_data.get('party_a', ''))
        cursor.execute("""
            INSERT INTO contracts (contract_no, contract_name, total_amt, sign_date, party_a,
                                   owner_id, status, classification, business_type,
                                   acceptance_date, expected_income_date, note, cust_id, paid_amt)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0)
        """, (row_data.get('contract_no', ''), row_data.get('contract_name', ''),
              row_data.get('total_amt', 0), row_data.get('sign_date', ''),
              row_data.get('party_a', ''), row_data.get('owner_id') or username,
              row_data.get('status', '执行中'), row_data.get('classification', ''),
              row_data.get('business_type', ''), row_data.get('acceptance_date', ''),
              row_data.get('expected_income_date', ''), row_data.get('note', ''), cust_id))

    elif module == 'payment_records':
        contract_id = _resolve_contract(cursor, row_data.get('contract_no', ''),
                                         row_data.get('contract_name', ''))
        if not contract_id:
            raise ValueError(f"找不到合同: {row_data.get('contract_no') or row_data.get('contract_name')}")

        cursor.execute("""
            INSERT INTO payment_records (contract_id, payment_date, amount, note, created_at)
            VALUES (?, ?, ?, ?, ?)
        """, (contract_id, row_data.get('payment_date', ''), row_data.get('amount', 0),
              row_data.get('note', ''), now))

        # 回写合同已回款金额
        cursor.execute("UPDATE contracts SET paid_amt = paid_amt + ? WHERE id = ?",
                       (row_data.get('amount', 0), contract_id))

    elif module == 'visits':
        cust_id = _resolve_customer(cursor, row_data.get('company', ''))
        cursor.execute("""
            INSERT INTO visits (cust_id, visitor_id, plan_date, plan_time, purpose, status,
                                result, location, contact_person, notes, work_type, work_content, created_at)
            VALUES (?, ?, ?, ?, ?, 'planned', ?, ?, ?, ?, ?, ?, ?)
        """, (cust_id, row_data.get('visitor_id') or username,
              row_data.get('plan_date', ''), row_data.get('plan_time', ''),
              row_data.get('purpose', ''), row_data.get('result', ''),
              row_data.get('location', ''), row_data.get('contact_person', ''),
              row_data.get('notes', ''), row_data.get('work_type', 'visit'),
              row_data.get('work_content', ''), now))

    elif module == 'scraped_leads':
        # 人工导入统一写入原始情报库（raw_intelligence），
        # 后续经 AI 商机识别分析 → 转入CRM → 分配销售，与自动采集共用同一链路
        import hashlib
        from routes.leads import _ensure_manual_source

        company = (row_data.get('company') or '').strip()
        opp_name = (row_data.get('opportunity_name') or '').strip()
        if not company and not opp_name:
            raise ValueError('公司与商机名称均为空')
        title = opp_name or company
        link = (row_data.get('link') or '').strip()

        content_parts = [f'【人工导入】{title}']
        for label, key in [
            ('采购单位/公司', 'company'), ('联系人', 'contact_name'), ('电话', 'phone'),
            ('邮箱', 'email'), ('行业', 'industry'), ('地区', 'region'),
            ('预算', 'budget'), ('采购方式', 'procurement_method'),
            ('招标编号', 'tender_no'), ('代理机构', 'agency'), ('代理电话', 'agency_phone'),
            ('发布日期', 'publish_date'), ('截止日期', 'deadline'), ('备注', 'remark'),
        ]:
            val = (row_data.get(key) or '').strip()
            if val:
                content_parts.append(f'{label}：{val}')
        content = '；'.join(content_parts)

        # 去重：有链接按链接哈希，否则按 公司+项目+电话
        if link:
            url_hash = hashlib.sha256(link.encode()).hexdigest()
        else:
            url_hash = hashlib.sha256(
                f'{company}|{opp_name}|{row_data.get("phone", "")}'.encode()
            ).hexdigest()
        dup = cursor.execute(
            "SELECT id FROM raw_intelligence WHERE url_hash=?", (url_hash,)
        ).fetchone()
        if dup:
            raise ValueError('原始情报库已存在相同内容，跳过重复导入')

        source_id = _ensure_manual_source(cursor)
        cursor.execute("""
            INSERT INTO raw_intelligence (source_id, url, url_hash, title, content,
                                          snippet, publish_date, status, keywords_matched)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'pending', '人工导入')
        """, (source_id, link, url_hash, title, content, content[:300],
              (row_data.get('publish_date') or '').strip()))

    elif module == 'enterprises':
        cursor.execute("""
            INSERT INTO enterprises (name, established_date, location, personnel_size, brief,
                registered_capital, business_scope, main_qualifications, main_products,
                relationship_status, cooperation_opportunities, website, contact_person,
                contact_info, owner_id, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            row_data.get('name', ''),
            row_data.get('established_date', ''),
            row_data.get('location', ''),
            row_data.get('personnel_size', ''),
            row_data.get('brief', ''),
            row_data.get('registered_capital', ''),
            row_data.get('business_scope', ''),
            row_data.get('main_qualifications', ''),
            row_data.get('main_products', ''),
            row_data.get('relationship_status', '未接触'),
            row_data.get('cooperation_opportunities', ''),
            row_data.get('website', ''),
            row_data.get('contact_person', ''),
            row_data.get('contact_info', ''),
            row_data.get('owner_id') or username,
            now, now,
        ))
    else:
        raise ValueError(f"不支持的模块: {module}")
