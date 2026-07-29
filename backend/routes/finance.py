from io import BytesIO
from datetime import datetime
from flask import request, jsonify

from extensions import get_db, record_operation_log, token_required

from . import finance_bp


def detect_file_format(file_bytes):
    if len(file_bytes) < 8:
        return 'unknown'
    if file_bytes[:4] == b'\x50\x4B\x03\x04':
        return 'xlsx'
    if file_bytes[:8] == b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1':
        return 'xls'
    if file_bytes[:3] == b'\xEF\xBB\xBF':
        return 'csv'
    if b'<!DOCTYPE' in file_bytes[:200] or b'<html' in file_bytes[:200]:
        return 'html'
    try:
        import zipfile
        try:
            with zipfile.ZipFile(BytesIO(file_bytes)) as zf:
                if '[Content_Types].xml' in zf.namelist():
                    return 'xlsx'
        except:
            pass
    except:
        pass
    return 'unknown'


@finance_bp.route('/api/payment_records', methods=['GET'])
@token_required
def get_payment_records():
    payload = request.current_user
    contract_id = request.args.get('contract_id')
    username = payload['username']
    role = payload['role']

    db = get_db()
    cursor = db.cursor()

    if contract_id:
        cursor.execute("""
            SELECT pr.*, c.contract_name, c.contract_no, c.party_a, u.name as owner_name
            FROM payment_records pr
            LEFT JOIN contracts c ON pr.contract_id = c.id
            LEFT JOIN users u ON c.owner_id = u.username
            WHERE pr.contract_id = ?
            ORDER BY pr.payment_date DESC
        """, (contract_id,))
    else:
        if role == '主任' or role == '院长':
            cursor.execute("""
                SELECT pr.*, c.contract_name, c.contract_no, c.party_a, u.name as owner_name
                FROM payment_records pr
                LEFT JOIN contracts c ON pr.contract_id = c.id
                LEFT JOIN users u ON c.owner_id = u.username
                ORDER BY pr.payment_date DESC
            """)
        else:
            cursor.execute("""
                SELECT pr.*, c.contract_name, c.contract_no, c.party_a, u.name as owner_name
                FROM payment_records pr
                LEFT JOIN contracts c ON pr.contract_id = c.id
                LEFT JOIN users u ON c.owner_id = u.username
                WHERE c.owner_id = ?
                ORDER BY pr.payment_date DESC
            """, (username,))

    rows = cursor.fetchall()
    records = []
    for row in rows:
        records.append(dict(row))

    return jsonify({'code': 200, 'message': 'success', 'data': records})


@finance_bp.route('/api/payment_records', methods=['POST'])
@token_required
def create_payment_record():
    payload = request.current_user
    data = request.get_json(silent=True) or {}
    username = payload['username']

    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute("""
            INSERT INTO payment_records (contract_id, payment_date, amount, note)
            VALUES (?, ?, ?, ?)
        """, (data.get('contract_id'), data.get('payment_date'), data.get('amount'), data.get('note')))
        db.commit()

        cursor.execute("SELECT SUM(amount) as total FROM payment_records WHERE contract_id = ?", (data.get('contract_id'),))
        total_paid = cursor.fetchone()['total'] or 0
        cursor.execute("UPDATE contracts SET paid_amt = ? WHERE id = ?", (total_paid, data.get('contract_id')))
        db.commit()

        record_operation_log(username, '创建', '回款', f'创建回款记录，合同ID:{data.get("contract_id")}，金额:{data.get("amount")}')

        return jsonify({'code': 200, 'message': '回款记录创建成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@finance_bp.route('/api/payment_records/<int:record_id>', methods=['PUT'])
@token_required
def update_payment_record(record_id):
    payload = request.current_user
    data = request.get_json(silent=True) or {}
    username = payload['username']

    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute("SELECT contract_id FROM payment_records WHERE id = ?", (record_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'code': 404, 'message': '回款记录不存在', 'data': None})

        original_contract_id = row['contract_id']

        cursor.execute("""
            UPDATE payment_records
            SET payment_date = ?, amount = ?, note = ?
            WHERE id = ?
        """, (data.get('payment_date'), data.get('amount'), data.get('note'), record_id))
        db.commit()

        cursor.execute("SELECT SUM(amount) as total FROM payment_records WHERE contract_id = ?", (original_contract_id,))
        total_paid = cursor.fetchone()['total'] or 0
        cursor.execute("UPDATE contracts SET paid_amt = ? WHERE id = ?", (total_paid, original_contract_id))
        db.commit()

        record_operation_log(username, '编辑', '回款', f'编辑回款记录，ID:{record_id}，金额:{data.get("amount")}')

        return jsonify({'code': 200, 'message': '回款记录更新成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@finance_bp.route('/api/payment_records/<int:record_id>', methods=['DELETE'])
@token_required
def delete_payment_record(record_id):
    payload = request.current_user
    role = payload.get('role', '')
    if role != '主任' and role != '院长':
        return jsonify({'code': 403, 'message': '权限不足，仅主任和院长可删除回款记录', 'data': None})

    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute("SELECT contract_id FROM payment_records WHERE id = ?", (record_id,))
        row = cursor.fetchone()
        if row:
            contract_id = row['contract_id']

            cursor.execute("DELETE FROM payment_records WHERE id = ?", (record_id,))

            cursor.execute("SELECT SUM(amount) as total FROM payment_records WHERE contract_id = ?", (contract_id,))
            total_paid = cursor.fetchone()['total'] or 0
            cursor.execute("UPDATE contracts SET paid_amt = ? WHERE id = ?", (total_paid, contract_id))

        db.commit()

        record_operation_log(payload['username'], '删除', '回款', f'删除回款记录，ID:{record_id}')

        return jsonify({'code': 200, 'message': '回款记录删除成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@finance_bp.route('/api/payments/import-parse', methods=['POST'])
@token_required
def import_parse_payments():
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请选择文件', 'data': None})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'code': 400, 'message': '请选择文件', 'data': None})

    filename = file.filename.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
        return jsonify({'code': 400, 'message': '仅支持Excel文件（.xlsx/.xls）', 'data': None})

    try:
        file_bytes = file.read()
        file_format = detect_file_format(file_bytes)

        if file_format == 'xls':
            try:
                import xlrd
                wb = xlrd.open_workbook(file_contents=file_bytes)
                ws = wb.sheet_by_index(0)
                headers = [ws.cell_value(0, i) for i in range(ws.ncols)]
                use_xlrd = True
            except ImportError:
                return jsonify({'code': 400, 'message': '请安装xlrd库以支持.xls文件格式，或使用.xlsx格式', 'data': None})
            except Exception as e:
                return jsonify({'code': 400, 'message': f'无法读取.xls文件：{str(e)}', 'data': None})
        elif file_format == 'xlsx':
            try:
                from openpyxl import load_workbook
                wb = load_workbook(BytesIO(file_bytes), data_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                use_xlrd = False
            except Exception as e:
                return jsonify({'code': 400, 'message': f'无法读取.xlsx文件：{str(e)}。请确认文件未被加密或损坏，或尝试另存为.xlsx格式', 'data': None})
        else:
            return jsonify({'code': 400, 'message': f'文件格式不匹配：扩展名是{filename.split(".")[-1]}，但实际文件格式无法识别。请确认文件是有效的Excel文件', 'data': None})

        def normalize_header(header):
            if header is None:
                return ''
            s = str(header)
            s = s.replace('\u3000', '')
            s = s.replace('　', '')
            s = ''.join(s.split())
            return s

        col_map = {}
        actual_headers = []

        for idx, header in enumerate(headers):
            normalized = normalize_header(header)
            actual_headers.append(str(header) if header else '')

            if '合同编号' in normalized or '合同号' in normalized:
                col_map['contract_no'] = idx
            elif '合同名称' in normalized or '合同名' in normalized:
                col_map['contract_name'] = idx
            elif '回款日期' in normalized or '日期' in normalized:
                col_map['payment_date'] = idx
            elif '金额' in normalized or '款额' in normalized:
                col_map['amount'] = idx
            elif '备注' in normalized or '说明' in normalized:
                col_map['note'] = idx

        required_cols = ['contract_no', 'payment_date', 'amount']
        missing_cols = []
        for col in required_cols:
            if col not in col_map:
                missing_cols.append(col)

        if missing_cols:
            col_names = {'contract_no': '合同编号', 'payment_date': '回款日期', 'amount': '金额'}
            missing_names = ', '.join([col_names.get(c, c) for c in missing_cols])
            return jsonify({
                'code': 400,
                'message': f'缺少必要列：{missing_names}。文件中的表头为：{", ".join(actual_headers)}',
                'data': None
            })

        db = get_db()
        cursor = db.cursor()

        cursor.execute("SELECT id, contract_no, contract_name FROM contracts")
        contract_map = {}
        for row in cursor.fetchall():
            contract_map[row['contract_no'].strip()] = {'id': row['id'], 'name': row['contract_name']}

        cursor.execute("""
            SELECT pr.id, pr.contract_id, pr.payment_date, pr.amount, pr.note, pr.created_at, c.contract_no, c.contract_name
            FROM payment_records pr
            LEFT JOIN contracts c ON pr.contract_id = c.id
        """)
        existing_payments = {}
        for row in cursor.fetchall():
            key = f"{row['contract_no']}_{row['payment_date']}_{row['amount']}"
            existing_payments[key] = {
                'id': row['id'],
                'contract_id': row['contract_id'],
                'contract_name': row['contract_name'],
                'note': row['note'],
                'created_at': row['created_at']
            }

        rows = []
        valid_count = 0

        max_row = ws.nrows if use_xlrd else ws.max_row

        for row_idx in range(2, max_row + 1):
            row_data = {}
            errors = []
            is_duplicate = False

            for key, idx in col_map.items():
                if use_xlrd:
                    value = ws.cell_value(row_idx - 1, idx)
                else:
                    value = ws.cell(row=row_idx, column=idx + 1).value

                if key == 'amount':
                    if value is None or value == '':
                        errors.append('金额不能为空')
                    else:
                        try:
                            value = float(value) * 10000
                        except:
                            errors.append('金额格式错误')
                elif key == 'payment_date':
                    if value:
                        if use_xlrd:
                            if isinstance(value, float):
                                try:
                                    date_tuple = xlrd.xldate_as_tuple(value, wb.datemode)
                                    value = f"{date_tuple[0]}-{str(date_tuple[1]).zfill(2)}-{str(date_tuple[2]).zfill(2)}"
                                except:
                                    value = str(value)[:10]
                            else:
                                value = str(value)[:10]
                        elif isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d')
                        else:
                            value = str(value)[:10]
                    else:
                        errors.append('回款日期不能为空')
                elif key == 'contract_no':
                    if value:
                        value = str(value).strip()
                    else:
                        errors.append('合同编号不能为空')

                row_data[key] = value

            contract_no = row_data.get('contract_no', '').strip()
            if contract_no:
                if contract_no not in contract_map:
                    errors.append('合同编号不存在于系统中')
                else:
                    row_data['contract_id'] = contract_map[contract_no]['id']

            payment_date = row_data.get('payment_date', '')
            amount = row_data.get('amount', 0)

            if contract_no and payment_date and amount:
                check_key = f"{contract_no}_{payment_date}_{amount}"
                if check_key in existing_payments:
                    is_duplicate = True
                    existing_info = existing_payments[check_key]
                    row_data['duplicate_id'] = existing_info['id']
                    row_data['existing_data'] = {
                        'id': existing_info['id'],
                        'contract_name': existing_info['contract_name'],
                        'payment_date': payment_date,
                        'amount': amount,
                        'note': existing_info['note'],
                        'created_at': existing_info['created_at']
                    }

            valid = len(errors) == 0
            if valid:
                valid_count += 1

            rows.append({
                'row_index': row_idx,
                'data': row_data,
                'valid': valid,
                'errors': errors,
                'is_duplicate': is_duplicate
            })

        return jsonify({
            'code': 200,
            'message': '解析成功',
            'data': {
                'total': len(rows),
                'valid_count': valid_count,
                'invalid_count': len(rows) - valid_count,
                'duplicate_count': sum(1 for r in rows if r['is_duplicate']),
                'rows': rows
            }
        })

    except Exception as e:
        return jsonify({'code': 500, 'message': f'解析失败：{str(e)}', 'data': None})


@finance_bp.route('/api/payments/import-execute', methods=['POST'])
@token_required
def import_execute_payments():
    data = request.get_json(silent=True) or {}
    if not data or not isinstance(data, list):
        return jsonify({'code': 400, 'message': '数据格式错误', 'data': None})

    db = get_db()
    cursor = db.cursor()

    success_count = 0
    fail_count = 0
    results = []

    for item in data:
        row_data = item.get('data', {})
        row_index = item.get('row_index', 0)
        duplicate_action = item.get('duplicate_action', 'keep_import')

        contract_id = row_data.get('contract_id')
        payment_date = row_data.get('payment_date')
        amount = row_data.get('amount')
        note = row_data.get('note', '')
        duplicate_id = row_data.get('duplicate_id')

        if not contract_id or not payment_date or not amount:
            results.append({
                'row_index': row_index,
                'success': False,
                'message': '数据不完整'
            })
            fail_count += 1
            continue

        try:
            if duplicate_id and duplicate_action == 'keep_existing':
                results.append({
                    'row_index': row_index,
                    'success': True,
                    'message': '保留系统数据，跳过导入'
                })
                success_count += 1
                continue

            if duplicate_id and duplicate_action == 'replace':
                cursor.execute("DELETE FROM payment_records WHERE id = ?", (duplicate_id,))

            cursor.execute("""
                INSERT INTO payment_records (contract_id, payment_date, amount, note)
                VALUES (?, ?, ?, ?)
            """, (contract_id, payment_date, amount, note))

            cursor.execute("SELECT SUM(amount) as total FROM payment_records WHERE contract_id = ?", (contract_id,))
            total_paid = cursor.fetchone()['total'] or 0
            cursor.execute("UPDATE contracts SET paid_amt = ? WHERE id = ?", (total_paid, contract_id))

            success_count += 1
            results.append({
                'row_index': row_index,
                'success': True,
                'message': '导入成功' if not duplicate_id else '替换成功'
            })
        except Exception as e:
            fail_count += 1
            results.append({
                'row_index': row_index,
                'success': False,
                'message': str(e)
            })

    db.commit()

    return jsonify({
        'code': 200,
        'message': '导入完成',
        'data': {
            'total': len(data),
            'success_count': success_count,
            'fail_count': fail_count,
            'results': results
        }
    })


def register_routes(app):
    app.register_blueprint(finance_bp)
