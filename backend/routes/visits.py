from flask import request, jsonify
from extensions import get_db, record_operation_log, token_required
from datetime import datetime

from . import visits_bp


@visits_bp.route('/api/visits', methods=['GET'])
@token_required
def get_visits():
    payload = request.current_user
    username = payload['username']
    role = payload['role']
    
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    visitor_id = request.args.get('visitor_id', '')
    cust_id = request.args.get('cust_id', '')
    status = request.args.get('status', '')
    keyword = request.args.get('keyword', '')

    db = get_db()
    cursor = db.cursor()

    conditions = []
    params = []

    if role not in ('主任', '院长'):
        conditions.append("v.visitor_id = ?")
        params.append(username)

    if start_date:
        conditions.append("v.plan_date >= ?")
        params.append(start_date)
    if end_date:
        conditions.append("v.plan_date <= ?")
        params.append(end_date)
    if visitor_id:
        conditions.append("v.visitor_id = ?")
        params.append(visitor_id)
    if cust_id:
        conditions.append("v.cust_id = ?")
        params.append(cust_id)
    if status:
        conditions.append("v.status = ?")
        params.append(status)
    if keyword:
        conditions.append("(c.name LIKE ? OR c.company LIKE ? OR v.purpose LIKE ?)")
        params.extend([f'%{keyword}%', f'%{keyword}%', f'%{keyword}%'])

    where_clause = ' AND '.join(conditions) if conditions else '1=1'

    cursor.execute(f"""
        SELECT v.*, c.name as customer_name, c.company as customer_company,
               c.contact_name as contact_name, c.phone as customer_phone,
               u.name as visitor_name,
               (SELECT e.id FROM enterprises e
                JOIN enterprise_visits ev ON ev.enterprise_id = e.id
                WHERE ev.visit_id = v.id LIMIT 1) as enterprise_id,
               (SELECT e.name FROM enterprises e
                JOIN enterprise_visits ev ON ev.enterprise_id = e.id
                WHERE ev.visit_id = v.id LIMIT 1) as enterprise_name
        FROM visits v
        LEFT JOIN customers c ON v.cust_id = c.id
        LEFT JOIN users u ON v.visitor_id = u.username
        WHERE {where_clause}
        ORDER BY v.plan_date DESC, v.plan_time DESC
    """, params)

    rows = cursor.fetchall()
    visits = [dict(row) for row in rows]

    return jsonify({'code': 200, 'message': 'success', 'data': visits})


@visits_bp.route('/api/visits/<int:visit_id>', methods=['GET'])
@token_required
def get_visit(visit_id):
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute("""
        SELECT v.*, c.name as customer_name, c.company as customer_company,
               c.contact_name as contact_name, c.phone as customer_phone,
               u.name as visitor_name,
               (SELECT e.id FROM enterprises e
                JOIN enterprise_visits ev ON ev.enterprise_id = e.id
                WHERE ev.visit_id = v.id LIMIT 1) as enterprise_id,
               (SELECT e.name FROM enterprises e
                JOIN enterprise_visits ev ON ev.enterprise_id = e.id
                WHERE ev.visit_id = v.id LIMIT 1) as enterprise_name
        FROM visits v
        LEFT JOIN customers c ON v.cust_id = c.id
        LEFT JOIN users u ON v.visitor_id = u.username
        WHERE v.id = ?
    """, (visit_id,))
    
    row = cursor.fetchone()
    if not row:
        return jsonify({'code': 404, 'message': '拜访记录不存在', 'data': None})
    
    return jsonify({'code': 200, 'message': 'success', 'data': dict(row)})


@visits_bp.route('/api/visits', methods=['POST'])
@token_required
def create_visit():
    payload = request.current_user
    data = request.get_json(silent=True) or {}
    username = payload['username']

    db = get_db()
    cursor = db.cursor()

    work_type = data.get('work_type', 'visit')
    cust_id = data.get('cust_id') if work_type == 'visit' else None
    work_content = data.get('work_content') if work_type == 'other' else None

    try:
        cursor.execute("""
            INSERT INTO visits (cust_id, visitor_id, plan_date, plan_time, 
                               purpose, location, contact_person, notes, status,
                               work_type, work_content)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'planned', ?, ?)
        """, (
            cust_id,
            data.get('visitor_id', username),
            data.get('plan_date'),
            data.get('plan_time'),
            data.get('purpose'),
            data.get('location'),
            data.get('contact_person'),
            data.get('notes'),
            work_type,
            work_content
        ))
        db.commit()
        visit_id = cursor.lastrowid

        # 如果传入了 enterprise_id，自动写入关联表
        enterprise_id = data.get('enterprise_id')
        if enterprise_id:
            try:
                cursor.execute(
                    "INSERT OR IGNORE INTO enterprise_visits (enterprise_id, visit_id) VALUES (?, ?)",
                    (enterprise_id, visit_id)
                )
                db.commit()
            except Exception:
                pass

        work_desc = data.get("purpose", "") if work_type == "visit" else data.get("work_content", "")
        record_operation_log(username, '创建', '拜访排班', 
            f'创建{("拜访" if work_type == "visit" else "其它工作")}计划：{data.get("plan_date")} - {work_desc}')

        return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': visit_id}})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@visits_bp.route('/api/visits/<int:visit_id>', methods=['PUT'])
@token_required
def update_visit(visit_id):
    payload = request.current_user
    data = request.get_json(silent=True) or {}
    username = payload['username']
    role = payload['role']

    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT visitor_id, status FROM visits WHERE id=?", (visit_id,))
    visit = cursor.fetchone()
    if not visit:
        return jsonify({'code': 404, 'message': '拜访记录不存在', 'data': None})

    if role not in ('主任', '院长') and visit['visitor_id'] != username:
        return jsonify({'code': 403, 'message': '权限不足，只能编辑自己的拜访记录', 'data': None})

    try:
        updates = []
        params = []
        
        work_type = data.get('work_type')
        if work_type == 'other':
            data['cust_id'] = None
        
        allowed_fields = ['cust_id', 'visitor_id', 'plan_date', 'plan_time', 
                          'purpose', 'location', 'contact_person', 'notes', 'status',
                          'actual_date', 'actual_time', 'result', 'work_type', 'work_content']
        
        for field in allowed_fields:
            if field in data:
                updates.append(f"{field} = ?")
                params.append(data[field])
        
        if updates:
            updates.append("updated_at = ?")
            params.append(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            params.append(visit_id)
            
            cursor.execute(f"""
                UPDATE visits SET {', '.join(updates)} WHERE id = ?
            """, params)
            db.commit()

            record_operation_log(username, '编辑', '拜访排班', 
                f'编辑排班计划 ID:{visit_id}')

        # 如果传入了 enterprise_id，更新企业关联
        enterprise_id = data.get('enterprise_id')
        if enterprise_id is not None:
            # 先删除旧关联
            cursor.execute("DELETE FROM enterprise_visits WHERE visit_id = ?", (visit_id,))
            # 如果传入了新的企业ID，写入新关联
            if enterprise_id:
                try:
                    cursor.execute(
                        "INSERT OR IGNORE INTO enterprise_visits (enterprise_id, visit_id) VALUES (?, ?)",
                        (enterprise_id, visit_id)
                    )
                except Exception:
                    pass
            db.commit()

        return jsonify({'code': 200, 'message': '更新成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@visits_bp.route('/api/visits/<int:visit_id>', methods=['DELETE'])
@token_required
def delete_visit(visit_id):
    payload = request.current_user
    username = payload['username']
    role = payload['role']

    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT visitor_id FROM visits WHERE id=?", (visit_id,))
    visit = cursor.fetchone()
    if not visit:
        return jsonify({'code': 404, 'message': '拜访记录不存在', 'data': None})

    if role not in ('主任', '院长') and visit['visitor_id'] != username:
        return jsonify({'code': 403, 'message': '权限不足', 'data': None})

    try:
        cursor.execute("DELETE FROM enterprise_visits WHERE visit_id = ?", (visit_id,))
        cursor.execute("DELETE FROM visits WHERE id=?", (visit_id,))
        db.commit()

        record_operation_log(username, '删除', '拜访排班', f'删除拜访计划 ID:{visit_id}')

        return jsonify({'code': 200, 'message': '拜访记录删除成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@visits_bp.route('/api/visits/<int:visit_id>/complete', methods=['POST'])
@token_required
def complete_visit(visit_id):
    """完成拜访排班，同时创建跟进记录并同步到客户和商机。

    请求体可选字段（除了原有的 actual_date/actual_time/result，新增跟进信息）：
      - follow_content: 跟进详细内容
      - follow_type: 跟进方式（面谈/电话/邮件等）
      - next_action: 下一步行动
      - next_date: 下一步日期
      - business_ids: 关联的商机ID列表
    """
    payload = request.current_user
    data = request.get_json(silent=True) or {}
    username = payload['username']

    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT * FROM visits WHERE id=?", (visit_id,))
    visit = cursor.fetchone()
    if not visit:
        return jsonify({'code': 404, 'message': '拜访记录不存在', 'data': None})

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    today = datetime.now().strftime('%Y-%m-%d')
    current_time = datetime.now().strftime('%H:%M')

    try:
        cursor.execute("""
            UPDATE visits 
            SET status = 'completed',
                actual_date = ?,
                actual_time = ?,
                result = ?,
                updated_at = ?
            WHERE id = ?
        """, (
            data.get('actual_date', today),
            data.get('actual_time', current_time),
            data.get('result'),
            now,
            visit_id
        ))

        # 仅对"客户拜访"类型创建跟进记录
        visit_dict = dict(visit)
        if visit_dict.get('work_type') == 'visit' and visit_dict.get('cust_id'):
            follow_content = data.get('follow_content') or data.get('result', '')
            follow_type = data.get('follow_type', '拜访')
            next_action = data.get('next_action', '')
            next_date = data.get('next_date', '')

            # 跟进内容格式：[跟进方式] + 跟进详情 + 拜访目的
            full_content = follow_content
            if follow_type:
                full_content = f"[{follow_type}] {follow_content}" if follow_content else f"[{follow_type}]"
            if visit_dict.get('purpose'):
                full_content += f"\n拜访目的：{visit_dict['purpose']}"

            # 1. 创建客户跟进记录
            cursor.execute("""
                INSERT INTO follow_logs (ref_type, ref_id, user_id, content, next_plan, log_time, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                'customer',
                visit_dict['cust_id'],
                username,
                full_content,
                next_action,
                now,
                now
            ))
            follow_id = cursor.lastrowid

            # 更新客户的最后跟进时间
            cursor.execute("UPDATE customers SET last_follow = ? WHERE id = ?", (now, visit_dict['cust_id']))

            # 2. 创建商机跟进记录（关联商机或自动关联客户活跃商机）
            business_ids = data.get('business_ids', [])
            if not business_ids:
                # 自动关联该客户的活跃商机
                cursor.execute("""
                    SELECT id FROM business
                    WHERE cust_id = ? AND status IN ('active', '跟进中')
                    ORDER BY id DESC LIMIT 5
                """, (visit_dict['cust_id'],))
                business_ids = [row['id'] for row in cursor.fetchall()]
            else:
                # 验证商机归属
                valid_ids = []
                for bid in business_ids:
                    cursor.execute("SELECT id FROM business WHERE id=? AND cust_id=?", (bid, visit_dict['cust_id']))
                    if cursor.fetchone():
                        valid_ids.append(bid)
                business_ids = valid_ids

            biz_follow_count = 0
            for bid in business_ids:
                cursor.execute("""
                    INSERT INTO follow_logs (ref_type, ref_id, user_id, content, next_plan, log_time, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    'business',
                    bid,
                    username,
                    f"[{follow_type or '拜访'}] {follow_content}",
                    next_action,
                    now,
                    now
                ))
                # 更新商机的最后跟进时间
                cursor.execute("UPDATE business SET last_follow = ? WHERE id = ?", (now, bid))
                biz_follow_count += 1

            record_operation_log(username, '完成', '拜访排班',
                f'完成拜访(ID:{visit_id})，创建客户跟进(ID:{follow_id})，关联{len(business_ids)}个商机跟进')

        db.commit()

        record_operation_log(username, '完成', '拜访排班', f'完成拜访计划 ID:{visit_id}')

        return jsonify({
            'code': 200,
            'message': '拜访已完成，跟进记录已同步',
            'data': {
                'visit_id': visit_id,
                'customer_follow_created': visit_dict.get('work_type') == 'visit' and visit_dict.get('cust_id') is not None,
                'business_follow_count': biz_follow_count if visit_dict.get('work_type') == 'visit' else 0
            }
        })
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@visits_bp.route('/api/visits/<int:visit_id>/update-complete', methods=['POST'])
@token_required
def update_complete_visit(visit_id):
    """修改已完成拜访的完成内容（不重复创建跟进记录）。

    请求体可选字段：
      - actual_date: 实际日期
      - actual_time: 实际时间
      - result: 完成结果/跟进内容
    """
    payload = request.current_user
    data = request.get_json(silent=True) or {}
    username = payload['username']
    role = payload['role']

    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT visitor_id, status, work_type, cust_id FROM visits WHERE id=?", (visit_id,))
    visit = cursor.fetchone()
    if not visit:
        return jsonify({'code': 404, 'message': '拜访记录不存在', 'data': None})

    if role not in ('主任', '院长') and visit['visitor_id'] != username:
        return jsonify({'code': 403, 'message': '权限不足', 'data': None})

    if visit['status'] != 'completed':
        return jsonify({'code': 400, 'message': '只能修改已完成的拜访', 'data': None})

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    try:
        cursor.execute("""
            UPDATE visits
            SET actual_date = ?,
                actual_time = ?,
                result = ?,
                updated_at = ?
            WHERE id = ?
        """, (
            data.get('actual_date'),
            data.get('actual_time'),
            data.get('result'),
            now,
            visit_id
        ))
        db.commit()

        record_operation_log(username, '修改', '拜访排班',
                             f'修改已完成拜访(ID:{visit_id})的完成内容')

        return jsonify({
            'code': 200,
            'message': '完成记录已更新',
            'data': {'visit_id': visit_id}
        })
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@visits_bp.route('/api/visits/<int:visit_id>/cancel', methods=['POST'])
@token_required
def cancel_visit(visit_id):
    payload = request.current_user
    username = payload['username']
    role = payload['role']

    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT visitor_id FROM visits WHERE id=?", (visit_id,))
    visit = cursor.fetchone()
    if not visit:
        return jsonify({'code': 404, 'message': '拜访记录不存在', 'data': None})

    if role not in ('主任', '院长') and visit['visitor_id'] != username:
        return jsonify({'code': 403, 'message': '权限不足', 'data': None})

    try:
        cursor.execute("""
            UPDATE visits 
            SET status = 'cancelled',
                updated_at = ?
            WHERE id = ?
        """, (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), visit_id))
        db.commit()

        record_operation_log(username, '取消', '拜访排班', f'取消拜访计划 ID:{visit_id}')

        return jsonify({'code': 200, 'message': '拜访已取消', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@visits_bp.route('/api/visits/stats/summary', methods=['GET'])
@token_required
def get_visit_stats():
    payload = request.current_user
    username = payload['username']
    role = payload['role']
    
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))

    db = get_db()
    cursor = db.cursor()

    if role in ('主任', '院长'):
        base_where = "1=1"
        params = []
    else:
        base_where = "visitor_id = ?"
        params = [username]

    cursor.execute(f"""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN status = 'planned' THEN 1 ELSE 0 END) as planned,
            SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) as completed,
            SUM(CASE WHEN status = 'cancelled' THEN 1 ELSE 0 END) as cancelled
        FROM visits
        WHERE {base_where} AND plan_date LIKE ?
    """, params + [f'{month}%'])

    row = cursor.fetchone()
    stats = dict(row)

    return jsonify({'code': 200, 'message': 'success', 'data': stats})


@visits_bp.route('/api/visits/customer-businesses/<int:cust_id>', methods=['GET'])
@token_required
def get_customer_businesses(cust_id):
    """获取客户的活跃商机列表，用于拜访完成时选择关联商机。"""
    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT id, title as name, stage, probability, status
        FROM business
        WHERE cust_id = ? AND status IN ('active', '跟进中', '新建')
        ORDER BY CASE status 
            WHEN 'active' THEN 1 
            WHEN '跟进中' THEN 2 
            WHEN '新建' THEN 3 
        END, id DESC
    """, (cust_id,))

    rows = cursor.fetchall()
    businesses = [dict(row) for row in rows]

    return jsonify({'code': 200, 'message': 'success', 'data': businesses})


@visits_bp.route('/api/visits/export-weekly-report', methods=['GET'])
@token_required
def export_weekly_report():
    """导出应用中心工作周报 Excel：每人本周工作 + 下周工作安排。

    - 本周工作：plan_date 落在本周一~本周日，排除已取消（反映本周实际做了/要做的事）
    - 下周安排：plan_date 落在下周一~下周日，且状态为 planned
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import timedelta
    from flask import send_file
    import io

    payload = request.current_user

    db = get_db()
    cursor = db.cursor()

    # 应用中心部门在职用户，主任/院长排在前面
    cursor.execute("""
        SELECT username, name, role, department
        FROM users
        WHERE department = '应用中心' AND (status = '在职' OR status IS NULL OR status = '')
        ORDER BY
            CASE role WHEN '主任' THEN 0 WHEN '院长' THEN 1 ELSE 2 END,
            name
    """)
    users = [dict(r) for r in cursor.fetchall()]

    # 本周/下周日期范围（周一到周日）
    today = datetime.now().date()
    this_monday = today - timedelta(days=today.weekday())
    this_sunday = this_monday + timedelta(days=6)
    next_monday = this_monday + timedelta(days=7)
    next_sunday = next_monday + timedelta(days=6)

    this_monday_str = this_monday.strftime('%Y-%m-%d')
    this_sunday_str = this_sunday.strftime('%Y-%m-%d')
    next_monday_str = next_monday.strftime('%Y-%m-%d')
    next_sunday_str = next_sunday.strftime('%Y-%m-%d')

    def fetch_visits_for(username, start, end, planned_only):
        if planned_only:
            sql = """
                SELECT v.*, c.name as customer_name, c.company as customer_company
                FROM visits v
                LEFT JOIN customers c ON v.cust_id = c.id
                WHERE v.visitor_id = ? AND v.plan_date >= ? AND v.plan_date <= ?
                  AND v.status = 'planned'
                ORDER BY v.plan_date, v.plan_time
            """
        else:
            sql = """
                SELECT v.*, c.name as customer_name, c.company as customer_company
                FROM visits v
                LEFT JOIN customers c ON v.cust_id = c.id
                WHERE v.visitor_id = ? AND v.plan_date >= ? AND v.plan_date <= ?
                  AND v.status != 'cancelled'
                ORDER BY v.plan_date, v.plan_time
            """
        cursor.execute(sql, (username, start, end))
        return [dict(r) for r in cursor.fetchall()]

    report_data = []
    for u in users:
        report_data.append({
            'user': u,
            'this_week': fetch_visits_for(u['username'], this_monday_str, this_sunday_str, False),
            'next_week': fetch_visits_for(u['username'], next_monday_str, next_sunday_str, True),
        })

    # ---------- 生成 Excel ----------
    wb = Workbook()
    ws = wb.active
    ws.title = '工作周报'

    title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
    title_fill = PatternFill('solid', fgColor='4472C4')
    subtitle_font = Font(name='微软雅黑', size=11, color='555555')
    user_font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')
    user_fill = PatternFill('solid', fgColor='D6E4F0')
    section_font = Font(name='微软雅黑', size=10, bold=True, color='2E75B6')
    section_fill = PatternFill('solid', fgColor='EAF1FB')
    header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='8EAADB')
    cell_font = Font(name='微软雅黑', size=10)
    empty_font = Font(name='微软雅黑', size=10, italic=True, color='999999')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = [14, 10, 12, 36, 24, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    last_col = chr(64 + len(widths))
    row = 1

    ws.merge_cells(f'A{row}:{last_col}{row}')
    c = ws.cell(row=row, column=1, value='应用中心工作周报')
    c.font = title_font
    c.fill = title_fill
    c.alignment = center
    ws.row_dimensions[row].height = 32
    row += 1

    ws.merge_cells(f'A{row}:{last_col}{row}')
    c = ws.cell(row=row, column=1,
                value=f'本周：{this_monday_str} ~ {this_sunday_str}    下周安排：{next_monday_str} ~ {next_sunday_str}')
    c.font = subtitle_font
    c.alignment = center
    row += 1
    row += 1  # 空行

    headers = ['日期', '时间', '类型', '工作内容', '客户/地点', '状态/结果']

    def write_visit_row(r, visit, period):
        if period == 'this':
            status_map = {'completed': '已完成', 'planned': '待完成', 'cancelled': '已取消'}
            status_text = status_map.get(visit.get('status'), visit.get('status', ''))
            result_val = visit.get('result') or ''
            if result_val:
                status_text += f'\n结果：{result_val}'
        else:
            status_text = '待完成'

        ws.cell(row=r, column=1, value=visit.get('plan_date', '')).alignment = center
        ws.cell(row=r, column=2, value=visit.get('plan_time') or '').alignment = center
        wtype = visit.get('work_type', 'visit')
        ws.cell(row=r, column=3, value='客户拜访' if wtype == 'visit' else '其它工作').alignment = center
        content = visit.get('work_content') if wtype == 'other' else (visit.get('purpose') or '')
        ws.cell(row=r, column=4, value=content or '').alignment = left
        if wtype == 'visit':
            parts = []
            if visit.get('customer_company'):
                parts.append(visit['customer_company'])
            if visit.get('customer_name'):
                parts.append(visit['customer_name'])
            loc = ' / '.join(parts) if parts else (visit.get('location') or '')
        else:
            loc = visit.get('location') or ''
        ws.cell(row=r, column=5, value=loc).alignment = left
        ws.cell(row=r, column=6, value=status_text).alignment = left
        for col in range(1, len(widths) + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = cell_font
            cell.border = border
        ws.row_dimensions[r].height = 30

    def write_header_row(r):
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[r].height = 22

    for item in report_data:
        u = item['user']
        ws.merge_cells(f'A{row}:{last_col}{row}')
        c = ws.cell(row=row, column=1, value=f'■ {u["name"]}（{u["role"]}）')
        c.font = user_font
        c.fill = user_fill
        c.alignment = left
        ws.row_dimensions[row].height = 26
        row += 1

        # 本周工作
        ws.merge_cells(f'A{row}:{last_col}{row}')
        c = ws.cell(row=row, column=1, value=f'【本周工作】（共 {len(item["this_week"])} 项）')
        c.font = section_font
        c.fill = section_fill
        c.alignment = left
        row += 1
        write_header_row(row)
        row += 1
        if item['this_week']:
            for v in item['this_week']:
                write_visit_row(row, v, 'this')
                row += 1
        else:
            ws.merge_cells(f'A{row}:{last_col}{row}')
            c = ws.cell(row=row, column=1, value='（本周暂无工作记录）')
            c.font = empty_font
            c.alignment = center
            row += 1

        # 下周安排
        ws.merge_cells(f'A{row}:{last_col}{row}')
        c = ws.cell(row=row, column=1, value=f'【下周安排】（共 {len(item["next_week"])} 项）')
        c.font = section_font
        c.fill = section_fill
        c.alignment = left
        row += 1
        write_header_row(row)
        row += 1
        if item['next_week']:
            for v in item['next_week']:
                write_visit_row(row, v, 'next')
                row += 1
        else:
            ws.merge_cells(f'A{row}:{last_col}{row}')
            c = ws.cell(row=row, column=1, value='（下周暂无安排）')
            c.font = empty_font
            c.alignment = center
            row += 1

        row += 1  # 人员之间空一行

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'应用中心工作周报_{this_monday_str}_{this_sunday_str}.xlsx'
    try:
        record_operation_log(payload['username'], '导出', '工作周报',
                             f'导出应用中心工作周报 {this_monday_str}~{this_sunday_str}')
    except Exception:
        pass

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def register_routes(app):
    app.register_blueprint(visits_bp)
