import os
import uuid
from io import BytesIO
from datetime import datetime
from flask import request, jsonify, send_from_directory
from extensions import (
    get_db, verify_token, record_operation_log,
    token_required, admin_required, user_can, UPLOAD_DIR,
)

from . import contracts_bp

# 合同附件允许的扩展名白名单
CONTRACT_FILE_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'jpg', 'jpeg', 'png', 'zip', 'rar'}


@contracts_bp.route('/api/contracts', methods=['GET'])
@token_required
def get_contracts():
    payload = request.current_user
    username = payload['username']
    role = payload['role']
    sort_field = request.args.get('sort_field', 'sign_date')
    sort_order = request.args.get('sort_order', 'desc')

    db = get_db()
    cursor = db.cursor()

    valid_fields = ['sign_date', 'total_amt', 'paid_amt', 'contract_name', 'contract_no', 'pending_amt']
    if sort_field not in valid_fields:
        sort_field = 'sign_date'

    sort_direction = 'DESC' if sort_order.lower() in ['desc', 'descending'] else 'ASC'
    order_by_clause = sort_field + " " + sort_direction

    # 关联名：customer_name（客户公司名）、business_title（关联商机标题）
    # 权限：持有 data.view_all（主任/院长）或 contracts.view_all（默认含应用中心）→ 查全部；其他 → 仅自己
    can_view_all = user_can(username, 'data.view_all') or user_can(username, 'contracts.view_all')

    if can_view_all:
        if sort_field == 'pending_amt':
            cursor.execute(
                "SELECT c.*, u.name as owner_name, cu.company as customer_name, b.title as business_title, "
                "(COALESCE(c.total_amt, 0) - COALESCE(c.paid_amt, 0)) as pending_amt "
                "FROM contracts c "
                "LEFT JOIN users u ON c.owner_id = u.username "
                "LEFT JOIN customers cu ON c.cust_id = cu.id "
                "LEFT JOIN business b ON c.b_id = b.id "
                "ORDER BY pending_amt " + sort_direction
            )
        else:
            cursor.execute(
                "SELECT c.*, u.name as owner_name, cu.company as customer_name, b.title as business_title "
                "FROM contracts c "
                "LEFT JOIN users u ON c.owner_id = u.username "
                "LEFT JOIN customers cu ON c.cust_id = cu.id "
                "LEFT JOIN business b ON c.b_id = b.id "
                "ORDER BY c." + order_by_clause
            )
    else:
        if sort_field == 'pending_amt':
            cursor.execute(
                "SELECT c.*, u.name as owner_name, cu.company as customer_name, b.title as business_title, "
                "(COALESCE(c.total_amt, 0) - COALESCE(c.paid_amt, 0)) as pending_amt "
                "FROM contracts c "
                "LEFT JOIN users u ON c.owner_id = u.username "
                "LEFT JOIN customers cu ON c.cust_id = cu.id "
                "LEFT JOIN business b ON c.b_id = b.id "
                "WHERE c.owner_id = ? "
                "ORDER BY pending_amt " + sort_direction,
                (username,)
            )
        else:
            cursor.execute(
                "SELECT c.*, u.name as owner_name, cu.company as customer_name, b.title as business_title "
                "FROM contracts c "
                "LEFT JOIN users u ON c.owner_id = u.username "
                "LEFT JOIN customers cu ON c.cust_id = cu.id "
                "LEFT JOIN business b ON c.b_id = b.id "
                "WHERE c.owner_id = ? "
                "ORDER BY c." + order_by_clause,
                (username,)
            )

    rows = cursor.fetchall()
    contracts = []
    for row in rows:
        contracts.append(dict(row))

    # —— 关键修复：income/待验收额/验收日期 与验收管理数据源对齐 ——
    # 不再信任 contracts.income 列（历史脏数据漏回填为 0），改为实时从
    # contract_acceptances 表汇总（与验收管理 / 考核完全同源）
    if contracts:
        ids = [c['id'] for c in contracts]
        placeholders = ','.join('?' * len(ids))
        cursor.execute(
            "SELECT contract_id, COALESCE(SUM(acceptance_amount), 0) as acc_sum, "
            "MAX(acceptance_date) as acc_date "
            f"FROM contract_acceptances WHERE contract_id IN ({placeholders}) "
            "GROUP BY contract_id",
            ids
        )
        acc_map = {r['contract_id']: (float(r['acc_sum']), r['acc_date']) for r in cursor.fetchall()}
        for c in contracts:
            if c['id'] in acc_map:
                acc_sum, acc_date = acc_map[c['id']]
                c['income'] = acc_sum
                c['pending_acceptance_amount'] = float(c.get('total_amt') or 0) - acc_sum
                if acc_date:
                    c['acceptance_date'] = acc_date

    return jsonify({'code': 200, 'message': 'success', 'data': contracts})


@contracts_bp.route('/api/contracts/check-no', methods=['GET'])
@token_required
def check_contract_no():
    contract_no = request.args.get('contract_no', '')
    exclude_id = request.args.get('exclude_id', type=int, default=None)

    db = get_db()
    cursor = db.cursor()

    if exclude_id:
        cursor.execute("SELECT COUNT(*) FROM contracts WHERE contract_no = ? AND id != ?", (contract_no, exclude_id))
    else:
        cursor.execute("SELECT COUNT(*) FROM contracts WHERE contract_no = ?", (contract_no,))

    count = cursor.fetchone()[0]

    return jsonify({'code': 200, 'message': 'success', 'data': {'exists': count > 0}})


@contracts_bp.route('/api/contracts/<int:contract_id>', methods=['GET'])
@token_required
def get_contract(contract_id):
    db = get_db()
    cursor = db.cursor()
    cursor.execute(
        "SELECT ct.*, u.name as owner_name, cu.company as customer_name, b.title as business_title "
        "FROM contracts ct "
        "LEFT JOIN users u ON ct.owner_id = u.username "
        "LEFT JOIN customers cu ON ct.cust_id = cu.id "
        "LEFT JOIN business b ON ct.b_id = b.id "
        "WHERE ct.id = ?",
        (contract_id,)
    )
    row = cursor.fetchone()

    if row:
        return jsonify({'code': 200, 'message': 'success', 'data': dict(row)})
    return jsonify({'code': 404, 'message': '合同不存在', 'data': None})


@contracts_bp.route('/api/contracts', methods=['POST'])
@token_required
def create_contract():
    payload = request.current_user
    data = request.get_json(silent=True) or {}
    username = payload['username']

    db = get_db()
    cursor = db.cursor()

    try:
        contract_no = data.get('contract_no')
        if not contract_no or contract_no.strip() == '':
            cursor.execute("SELECT MAX(id) FROM contracts")
            max_id = cursor.fetchone()[0] or 0
            contract_no = f"HT{datetime.now().strftime('%Y%m%d%H%M%S')}{str(max_id + 1).zfill(3)}"

        cursor.execute("SELECT COUNT(*) FROM contracts WHERE contract_no = ?", (contract_no,))
        if cursor.fetchone()[0] > 0:
            contract_no = f"HT{datetime.now().strftime('%Y%m%d%H%M%S')}{str(max_id + 1).zfill(3)}{str(uuid.uuid4().hex[:3])}"

        # 关联客户/商机，并做一致性兜底：若传了 b_id，以商机的 cust_id 为准，防数据撕裂
        b_id = data.get('b_id')
        b_id = b_id if b_id else None  # 规范化空字符串为 None
        cust_id = data.get('cust_id')
        if b_id:
            cursor.execute("SELECT cust_id FROM business WHERE id = ?", (b_id,))
            brow = cursor.fetchone()
            if brow and brow['cust_id']:
                cust_id = brow['cust_id']

        cursor.execute("""
            INSERT INTO contracts
            (b_id, cust_id, contract_no, party_a, project_order_no, total_amt, paid_amt, sign_date, owner_id, status,
             contract_name, classification, is_audit, pending_acceptance_amount,
             cost, gross_profit, acceptance_date, expected_income_date,
             expected_income_year, business_type, total_cost, acceptance_nodes, payment_nodes, note, is_framework,
             income, tax_amount, business_direction)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?, ?)
        """, (
            b_id, cust_id, contract_no, data.get('party_a'), data.get('project_order_no'),
            data.get('total_amt'), 0, data.get('sign_date'), data.get('owner_id'), '执行中',
            data.get('contract_name'), data.get('classification'), data.get('is_audit'), data.get('pending_acceptance_amount'),
            data.get('cost'), data.get('gross_profit'), data.get('acceptance_date'), data.get('expected_income_date'),
            data.get('expected_income_year'), data.get('business_type'), data.get('acceptance_nodes'), data.get('payment_nodes'),
            data.get('note'), 1 if data.get('is_framework') else 0,
            data.get('income', 0), data.get('tax_amount', 0), data.get('business_direction')
        ))
        db.commit()
        contract_id = cursor.lastrowid

        record_operation_log(username, '创建', '合同', f'创建合同：{data.get("contract_name")}（{contract_no}）')

        return jsonify({'code': 200, 'message': '合同创建成功', 'data': {'id': contract_id, 'contract_no': contract_no}})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@contracts_bp.route('/api/contracts/<int:contract_id>', methods=['PUT'])
@token_required
def update_contract(contract_id):
    """编辑合同：管理员可编辑任何合同；合同负责人可编辑自己负责的合同，但不能改负责人。"""
    payload = request.current_user
    data = request.get_json(silent=True) or {}
    username = payload['username']

    db = get_db()
    cursor = db.cursor()

    # 检查权限：管理员 OR 合同负责人
    cursor.execute("SELECT owner_id FROM contracts WHERE id=?", (contract_id,))
    contract_row = cursor.fetchone()
    if not contract_row:
        return jsonify({'code': 404, 'message': '合同不存在', 'data': None})

    is_admin = user_can(username, 'data.view_all')
    is_owner = (contract_row['owner_id'] == username)

    if not is_admin and not is_owner:
        return jsonify({'code': 403, 'message': '无权编辑非自己负责的合同', 'data': None})

    # 负责人不能通过编辑修改合同负责人（强制保留原值）
    effective_owner_id = data.get('owner_id') if is_admin else contract_row['owner_id']

    try:
        # 关联客户/商机，并做一致性兜底：若传了 b_id，以商机的 cust_id 为准，防数据撕裂
        b_id = data.get('b_id')
        b_id = b_id if b_id else None  # 规范化空字符串为 None
        cust_id = data.get('cust_id')
        if b_id:
            cursor.execute("SELECT cust_id FROM business WHERE id = ?", (b_id,))
            brow = cursor.fetchone()
            if brow and brow['cust_id']:
                cust_id = brow['cust_id']

        cursor.execute("""
            UPDATE contracts SET
                contract_name=?, contract_no=?, party_a=?, project_order_no=?, total_amt=?, sign_date=?,
                classification=?, is_audit=?, pending_acceptance_amount=?,
                cost=?, gross_profit=?, acceptance_date=?, expected_income_date=?,
                expected_income_year=?, business_type=?, status=?, owner_id=?,
                cust_id=?, b_id=?,
                acceptance_nodes=?, payment_nodes=?, note=?, is_framework=?,
                income=?, tax_amount=?, business_direction=?
            WHERE id=?
        """, (
            data.get('contract_name'), data.get('contract_no'), data.get('party_a'), data.get('project_order_no'),
            data.get('total_amt'), data.get('sign_date'), data.get('classification'), data.get('is_audit'),
            data.get('pending_acceptance_amount'), data.get('cost'), data.get('gross_profit'),
            data.get('acceptance_date'), data.get('expected_income_date'), data.get('expected_income_year'),
            data.get('business_type'), data.get('status'), effective_owner_id,
            cust_id, b_id,
            data.get('acceptance_nodes'), data.get('payment_nodes'), data.get('note'),
            1 if data.get('is_framework') else 0,
            data.get('income', 0), data.get('tax_amount', 0), data.get('business_direction'),
            contract_id
        ))
        db.commit()

        record_operation_log(username, '编辑', '合同', f'编辑合同：{data.get("contract_name")}（ID:{contract_id}）')

        return jsonify({'code': 200, 'message': '合同更新成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@contracts_bp.route('/api/contracts/<int:contract_id>/owner', methods=['PATCH', 'POST'])
@token_required
def update_contract_owner(contract_id):
    payload = request.current_user

    if not user_can(payload['username'], 'data.view_all'):
        return jsonify({'code': 403, 'message': '无权修改负责人', 'data': None})

    data = request.get_json(silent=True) or {}
    if not data:
        return jsonify({'code': 400, 'message': '请求数据为空', 'data': None})

    owner_id = data.get('owner_id')

    if not owner_id:
        return jsonify({'code': 400, 'message': '请选择负责人', 'data': None})

    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute("UPDATE contracts SET owner_id = ? WHERE id = ?", (owner_id, contract_id))
        db.commit()

        return jsonify({'code': 200, 'message': '负责人修改成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


# ==================== 合同销售分成 ====================

@contracts_bp.route('/api/contracts/<int:contract_id>/commissions', methods=['GET'])
@token_required
def get_contract_commissions(contract_id):
    """获取合同的销售分成列表。"""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT id, total_amt, contract_name, owner_id FROM contracts WHERE id=?", (contract_id,))
    c = cur.fetchone()
    if not c:
        return jsonify({'code': 404, 'message': '合同不存在', 'data': None})
    cur.execute(
        "SELECT cc.username, cc.ratio, u.name "
        "FROM contract_commissions cc LEFT JOIN users u ON cc.username = u.username "
        "WHERE cc.contract_id=? ORDER BY cc.ratio DESC",
        (contract_id,)
    )
    rows = [{'username': r['username'], 'name': r['name'] or r['username'], 'ratio': float(r['ratio'])}
            for r in cur.fetchall()]
    return jsonify({
        'code': 200, 'message': 'OK',
        'data': {
            'contract_id': contract_id,
            'contract_name': c['contract_name'],
            'total_amt': float(c['total_amt'] or 0),
            'owner_id': c['owner_id'],
            'commissions': rows,
        }
    })


@contracts_bp.route('/api/contracts/<int:contract_id>/commissions', methods=['POST'])
@token_required
def save_contract_commissions(contract_id):
    """保存合同销售分成（整体替换）。仅主任/院长。
    body: {commissions: [{username, ratio}, ...]}  ratio之和必须=100
    """
    payload = request.current_user
    if not user_can(payload['username'], 'data.view_all'):
        return jsonify({'code': 403, 'message': '无权设置分成', 'data': None})
    data = request.get_json(silent=True) or {}
    items = data.get('commissions') or []
    if not isinstance(items, list):
        return jsonify({'code': 400, 'message': 'commissions 必须是数组', 'data': None})
    # 校验
    total_ratio = 0
    seen = set()
    for item in items:
        u = (item.get('username') or '').strip()
        r = float(item.get('ratio') or 0)
        if not u:
            return jsonify({'code': 400, 'message': '分成人员用户名不能为空', 'data': None})
        if r < 0 or r > 100:
            return jsonify({'code': 400, 'message': f'比例 {r}% 超出范围(0-100)', 'data': None})
        if u in seen:
            return jsonify({'code': 400, 'message': f'人员 {u} 重复', 'data': None})
        seen.add(u)
        total_ratio += r
    if items and abs(total_ratio - 100.0) > 0.01:
        return jsonify({'code': 400, 'message': f'分成比例之和为 {total_ratio}%，必须等于 100%', 'data': None})

    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT 1 FROM contracts WHERE id=?", (contract_id,))
    if not cur.fetchone():
        return jsonify({'code': 404, 'message': '合同不存在', 'data': None})

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    operator = payload['username']
    try:
        # 整体替换：先删后插
        cur.execute("DELETE FROM contract_commissions WHERE contract_id=?", (contract_id,))
        for item in items:
            u = (item.get('username') or '').strip()
            r = float(item.get('ratio') or 0)
            cur.execute(
                "INSERT INTO contract_commissions (contract_id, username, ratio, created_by, created_at) "
                "VALUES (?, ?, ?, ?, ?)",
                (contract_id, u, r, operator, now)
            )
        db.commit()
        try:
            record_operation_log(operator, '设置分成', '合同管理',
                                 f'合同ID={contract_id} 设置销售分成：'
                                 + ', '.join(f"{it['username']}={it.get('ratio',0)}%" for it in items)
                                 if items else f'合同ID={contract_id} 清除分成')
        except Exception:
            pass
        return jsonify({'code': 200, 'message': '分成保存成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


# ==================== 框架合同验收记录 ====================

def _sync_contract_acceptance_fields(cur, contract_id):
    """根据验收记录同步合同的 income/待验收额/验收日期，确保合同管理与验收管理数据联动。
    收入(累计验收额) = SUM(acceptance_amount)
    待验收合同额 = 合同额 - 累计验收额
    验收日期 = 最近一次验收日期
    """
    cur.execute(
        "SELECT COALESCE(SUM(acceptance_amount), 0), MAX(acceptance_date) "
        "FROM contract_acceptances WHERE contract_id=?",
        (contract_id,)
    )
    row = cur.fetchone()
    acc_total = float(row[0] or 0)
    latest_date = row[1]
    cur.execute("SELECT COALESCE(total_amt, 0) FROM contracts WHERE id=?", (contract_id,))
    total_amt = float(cur.fetchone()[0] or 0)
    cur.execute(
        "UPDATE contracts SET income=?, pending_acceptance_amount=?, acceptance_date=? WHERE id=?",
        (acc_total, total_amt - acc_total, latest_date, contract_id)
    )


@contracts_bp.route('/api/contracts/<int:contract_id>/acceptances', methods=['GET'])
@token_required
def get_acceptances(contract_id):
    """获取框架合同的验收记录列表（含每次验收的分成分配）。"""
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT id, is_framework, total_amt, contract_name FROM contracts WHERE id=?", (contract_id,))
    c = cur.fetchone()
    if not c:
        return jsonify({'code': 404, 'message': '合同不存在', 'data': None})
    cur.execute(
        "SELECT id, acceptance_date, acceptance_amount, note, created_by, created_at "
        "FROM contract_acceptances WHERE contract_id=? ORDER BY acceptance_date DESC",
        (contract_id,)
    )
    rows = [dict(r) for r in cur.fetchall()]
    # 查每条验收的分成
    for r in rows:
        cur.execute(
            "SELECT username, ratio FROM acceptance_commissions WHERE acceptance_id=? ORDER BY ratio DESC",
            (r['id'],)
        )
        r['commissions'] = [{'username': cr['username'], 'ratio': float(cr['ratio'])} for cr in cur.fetchall()]
    total_accepted = sum(float(r['acceptance_amount'] or 0) for r in rows)
    return jsonify({
        'code': 200, 'message': 'OK',
        'data': {
            'contract_id': contract_id,
            'contract_name': c['contract_name'],
            'is_framework': int(c['is_framework'] or 0),
            'total_amt': float(c['total_amt'] or 0),
            'total_accepted': round(total_accepted, 2),
            'acceptances': rows,
        }
    })


@contracts_bp.route('/api/contracts/<int:contract_id>/acceptances', methods=['POST'])
@token_required
def add_acceptance(contract_id):
    """新增一条验收记录（可同时传入分成）。
    body: {acceptance_date, acceptance_amount, note, commissions: [{username, ratio}]}
    管理员或合同负责人可用。
    """
    payload = request.current_user
    username = payload['username']
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT owner_id FROM contracts WHERE id=?", (contract_id,))
    c = cur.fetchone()
    if not c:
        return jsonify({'code': 404, 'message': '合同不存在', 'data': None})
    if not user_can(username, 'data.view_all') and c['owner_id'] != username:
        return jsonify({'code': 403, 'message': '无权操作非自己负责的合同验收', 'data': None})

    data = request.get_json(silent=True) or {}
    acc_date = (data.get('acceptance_date') or '').strip()
    acc_amt = float(data.get('acceptance_amount') or 0)
    if not acc_date:
        return jsonify({'code': 400, 'message': '验收日期不能为空', 'data': None})
    if acc_amt == 0:
        return jsonify({'code': 400, 'message': '收入不能为0（正数为验收，负数为核减）', 'data': None})
    commissions = data.get('commissions') or []
    # 校验分成
    total_ratio = 0
    seen = set()
    for item in commissions:
        u = (item.get('username') or '').strip()
        r = float(item.get('ratio') or 0)
        if not u:
            return jsonify({'code': 400, 'message': '分成人员用户名不能为空', 'data': None})
        if u in seen:
            return jsonify({'code': 400, 'message': f'人员 {u} 重复', 'data': None})
        seen.add(u)
        total_ratio += r
    if commissions and abs(total_ratio - 100.0) > 0.01:
        return jsonify({'code': 400, 'message': f'分成比例之和为 {total_ratio}%，必须等于 100%', 'data': None})

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    operator = payload['username']
    try:
        cur.execute(
            "INSERT INTO contract_acceptances (contract_id, acceptance_date, acceptance_amount, note, created_by, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (contract_id, acc_date, acc_amt, data.get('note', ''), operator, now)
        )
        acc_id = cur.lastrowid
        # 写入验收级分成
        for item in commissions:
            u = (item.get('username') or '').strip()
            r = float(item.get('ratio') or 0)
            cur.execute(
                "INSERT INTO acceptance_commissions (acceptance_id, username, ratio, created_by, created_at) "
                "VALUES (?, ?, ?, ?, ?)",
                (acc_id, u, r, operator, now)
            )
        # 联动同步：更新合同的收入/待验收额/验收日期
        _sync_contract_acceptance_fields(cur, contract_id)
        db.commit()
        try:
            record_operation_log(operator, '新增验收', '合同管理',
                                 f'合同ID={contract_id} 验收日期={acc_date} 金额={acc_amt}'
                                 + (f' 分成={commissions}' if commissions else ''))
        except Exception:
            pass
        return jsonify({'code': 200, 'message': '验收记录添加成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@contracts_bp.route('/api/contracts/acceptances/<int:acc_id>', methods=['DELETE'])
@token_required
def delete_acceptance(acc_id):
    """删除一条验收记录（同时删除其分成）。"""
    payload = request.current_user
    if not user_can(payload['username'], 'data.view_all'):
        return jsonify({'code': 403, 'message': '无权删除验收记录', 'data': None})
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT contract_id, acceptance_amount FROM contract_acceptances WHERE id=?", (acc_id,))
    row = cur.fetchone()
    if not row:
        return jsonify({'code': 404, 'message': '验收记录不存在', 'data': None})
    try:
        cur.execute("DELETE FROM acceptance_commissions WHERE acceptance_id=?", (acc_id,))
        cur.execute("DELETE FROM contract_acceptances WHERE id=?", (acc_id,))
        # 联动同步：重新计算合同的收入/待验收额/验收日期
        _sync_contract_acceptance_fields(cur, row['contract_id'])
        db.commit()
        try:
            record_operation_log(payload['username'], '删除验收', '合同管理',
                                 f'删除验收记录 ID={acc_id}（合同ID={row["contract_id"]} 金额={row["acceptance_amount"]}）')
        except Exception:
            pass
        return jsonify({'code': 200, 'message': '删除成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@contracts_bp.route('/api/contracts/import-parse', methods=['POST'])
@token_required
def import_parse_contracts():
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请选择文件', 'data': None})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'code': 400, 'message': '请选择文件', 'data': None})

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'code': 400, 'message': '仅支持Excel文件（.xlsx/.xls）', 'data': None})

    try:
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        col_map = {}
        for idx, header in enumerate(headers):
            header = str(header).strip() if header else ''
            if header == '合同编号':
                col_map['contract_no'] = idx
            elif header == '合同名称':
                col_map['contract_name'] = idx
            elif header == '甲方':
                col_map['party_a'] = idx
            elif header == '项目令号':
                col_map['project_order_no'] = idx
            elif header == '合同总额(万)':
                col_map['total_amt'] = idx
            elif header == '签约日期':
                col_map['sign_date'] = idx
            elif header == '业态':
                col_map['business_type'] = idx
            elif header == '密级':
                col_map['classification'] = idx
            elif header == '负责人':
                col_map['owner_name'] = idx
            elif header == '验收节点':
                col_map['acceptance_nodes'] = idx
            elif header == '回款节点':
                col_map['payment_nodes'] = idx

        required_cols = ['contract_no', 'contract_name', 'total_amt']
        for col in required_cols:
            if col not in col_map:
                return jsonify({'code': 400, 'message': f'缺少必要列：{col}', 'data': None})

        db = get_db()
        cursor = db.cursor()
        cursor.execute("SELECT contract_no FROM contracts")
        existing_nos = set(row[0] for row in cursor.fetchall())

        rows = []
        batch_nos = set()
        valid_count = 0

        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            errors = []

            for key, idx in col_map.items():
                cell = ws.cell(row=row_idx, column=idx + 1)
                value = cell.value

                if key == 'total_amt':
                    if value is None:
                        errors.append('合同总额不能为空')
                    else:
                        try:
                            value = float(value) * 10000
                        except:
                            errors.append('合同总额格式错误')
                elif key == 'sign_date':
                    if value:
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d')
                        else:
                            value = str(value)[:10]

                row_data[key] = value

            contract_no = row_data.get('contract_no')
            if contract_no:
                contract_no = str(contract_no).strip()
                row_data['contract_no'] = contract_no

                if contract_no in existing_nos:
                    errors.append('合同编号已存在')
                if contract_no in batch_nos:
                    errors.append('批内合同编号重复')
                batch_nos.add(contract_no)

            contract_name = row_data.get('contract_name')
            if not contract_name:
                errors.append('合同名称不能为空')

            valid = len(errors) == 0
            if valid:
                valid_count += 1

            rows.append({
                'row_index': row_idx,
                'data': row_data,
                'valid': valid,
                'errors': errors
            })

        return jsonify({
            'code': 200,
            'message': '解析成功',
            'data': {
                'total': len(rows),
                'valid_count': valid_count,
                'invalid_count': len(rows) - valid_count,
                'rows': rows
            }
        })

    except Exception as e:
        return jsonify({'code': 500, 'message': f'解析失败：{str(e)}', 'data': None})


@contracts_bp.route('/api/contracts/import-execute', methods=['POST'])
@token_required
def import_execute_contracts():
    payload = request.current_user
    data = request.get_json(silent=True) or {}
    if not data or not isinstance(data, list):
        return jsonify({'code': 400, 'message': '数据格式错误', 'data': None})

    db = get_db()
    cursor = db.cursor()

    success_count = 0
    fail_count = 0
    results = []

    cursor.execute("SELECT contract_no FROM contracts")
    existing_nos = set(row[0] for row in cursor.fetchall())

    for item in data:
        row_data = item.get('data', {})
        row_index = item.get('row_index', 0)

        contract_no = row_data.get('contract_no')
        if not contract_no or contract_no in existing_nos:
            results.append({
                'row_index': row_index,
                'success': False,
                'message': '合同编号已存在或为空'
            })
            fail_count += 1
            continue

        try:
            cursor.execute("""
                INSERT INTO contracts
                (contract_no, party_a, project_order_no, total_amt, paid_amt, sign_date, owner_id, status,
                 contract_name, classification, is_audit, pending_acceptance_amount,
                 cost, gross_profit, acceptance_date, expected_income_date,
                 expected_income_year, business_type, total_cost, acceptance_nodes, payment_nodes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
            """, (
                contract_no,
                row_data.get('party_a'),
                row_data.get('project_order_no'),
                row_data.get('total_amt', 0),
                0,
                row_data.get('sign_date'),
                payload['username'],
                '执行中',
                row_data.get('contract_name'),
                row_data.get('classification'),
                0,
                0,
                0,
                0,
                '',
                '',
                '',
                row_data.get('business_type'),
                row_data.get('acceptance_nodes'),
                row_data.get('payment_nodes')
            ))

            existing_nos.add(contract_no)
            success_count += 1
            results.append({
                'row_index': row_index,
                'success': True,
                'message': '导入成功'
            })
        except Exception as e:
            fail_count += 1
            results.append({
                'row_index': row_index,
                'success': False,
                'message': str(e)
            })

    db.commit()

    return jsonify({
        'code': 200,
        'message': '导入完成',
        'data': {
            'total': len(data),
            'success_count': success_count,
            'fail_count': fail_count,
            'results': results
        }
    })


@contracts_bp.route('/api/contracts/<int:contract_id>', methods=['DELETE'])
@token_required
def delete_contract(contract_id):
    payload = request.current_user
    if not user_can(payload['username'], 'data.view_all'):
        return jsonify({'code': 403, 'message': '权限不足，仅主任和院长可删除合同', 'data': None})

    db = get_db()
    cursor = db.cursor()

    try:
        cursor.execute("SELECT contract_name, contract_no FROM contracts WHERE id=?", (contract_id,))
        row = cursor.fetchone()
        contract_info = f"{row['contract_name']}（{row['contract_no']}）" if row else f"ID:{contract_id}"

        cursor.execute("DELETE FROM payment_records WHERE contract_id = ?", (contract_id,))
        cursor.execute("DELETE FROM contracts WHERE id=?", (contract_id,))
        db.commit()

        record_operation_log(payload['username'], '删除', '合同', f'删除合同：{contract_info}')

        return jsonify({'code': 200, 'message': '合同删除成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@contracts_bp.route('/api/contracts/upload', methods=['POST'])
@token_required
def upload_contract_file():
    contract_id = request.form.get('contract_id', type=int)
    file_type = request.form.get('file_type')

    if file_type not in ('contract', 'tech'):
        return jsonify({'code': 400, 'message': '无效的文件类型', 'data': None})

    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请选择文件', 'data': None})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'code': 400, 'message': '请选择文件', 'data': None})

    # 扩展名白名单校验，防止上传可执行/脚本文件
    ext = os.path.splitext(file.filename)[1].lower().lstrip('.')
    if ext not in CONTRACT_FILE_EXTENSIONS:
        return jsonify({'code': 400, 'message': f'不支持的文件格式（允许：{"、".join(sorted(CONTRACT_FILE_EXTENSIONS))}）', 'data': None})

    os.makedirs(UPLOAD_DIR, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # contract_id 为 int、file_type 已白名单校验、ext 已白名单校验，文件名无注入风险
    filename = f"{contract_id}_{file_type}_{timestamp}.{ext}"
    file_path = os.path.join(UPLOAD_DIR, filename)

    file.save(file_path)

    db = get_db()
    cursor = db.cursor()

    if file_type == 'contract':
        cursor.execute("UPDATE contracts SET contract_file_path = ? WHERE id = ?", (f"uploads/contracts/{filename}", contract_id))
    elif file_type == 'tech':
        cursor.execute("UPDATE contracts SET tech_agreement_file_path = ? WHERE id = ?", (f"uploads/contracts/{filename}", contract_id))

    db.commit()

    return jsonify({'code': 200, 'message': '文件上传成功', 'data': {'file_path': f"uploads/contracts/{filename}"}})


@contracts_bp.route('/api/contracts/test-download', methods=['GET'])
def test_download_route():
    return jsonify({'code': 200, 'message': '测试路由工作正常', 'data': None})


@contracts_bp.route('/api/download-contract', methods=['GET'])
def download_contract_file():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    if not token:
        token = request.args.get('token', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})

    contract_id = request.args.get('id', type=int)
    file_type = request.args.get('type')

    if not contract_id or not file_type:
        return jsonify({'code': 400, 'message': '参数错误', 'data': None})

    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM contracts WHERE id = ?", (contract_id,))
    row = cursor.fetchone()

    if not row:
        return jsonify({'code': 404, 'message': '合同不存在', 'data': None})

    row_dict = dict(row)

    if file_type == 'contract':
        file_path = row_dict.get('contract_file_path')
    elif file_type == 'tech':
        file_path = row_dict.get('tech_agreement_file_path')
    else:
        return jsonify({'code': 400, 'message': '无效的文件类型', 'data': None})

    if not file_path:
        return jsonify({'code': 404, 'message': '文件不存在', 'data': None})

    filename = os.path.basename(file_path)
    full_path = os.path.join(UPLOAD_DIR, filename)

    if not os.path.exists(full_path):
        return jsonify({'code': 404, 'message': '文件不存在', 'data': None})

    return send_from_directory(UPLOAD_DIR, filename, as_attachment=False)


def register_routes(app):
    app.register_blueprint(contracts_bp)


# ==================== 验收管理 ====================

@contracts_bp.route('/api/acceptances', methods=['GET'])
@token_required
def get_acceptance_list():
    """验收列表：每笔验收记录一行，附带合同上下文与累计已验收额。"""
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT ca.id AS acceptance_id, ca.acceptance_date, ca.acceptance_amount, ca.note,
               ca.created_by, ca.created_at,
               c.id AS contract_id, c.contract_name, c.contract_no, c.party_a,
               c.total_amt, c.income, c.tax_amount, c.business_direction, c.status,
               (SELECT COALESCE(SUM(ca2.acceptance_amount), 0)
                FROM contract_acceptances ca2
                WHERE ca2.contract_id = c.id) AS cumulative_accepted
        FROM contract_acceptances ca
        JOIN contracts c ON ca.contract_id = c.id
        ORDER BY ca.acceptance_date DESC, c.sign_date DESC
    """)
    rows = [dict(r) for r in cursor.fetchall()]
    for r in rows:
        total = float(r.get('total_amt') or 0)
        cum = float(r.get('cumulative_accepted') or 0)
        r['pending_acceptance_amount'] = round(total - cum, 2)
    return jsonify({'code': 200, 'data': rows})


@contracts_bp.route('/api/acceptances/import', methods=['POST'])
@token_required
def import_acceptances():
    """导入验收数据：每行创建一笔验收记录，并可选更新合同级字段(收入/税额/业务方向)。
    必填：合同编号、验收日期、验收金额(万)。可选：验收情况、收入(万)、税额(万)、业务方向。
    """
    payload = request.current_user
    username = payload['username']

    data = request.get_json(silent=True)
    # 兼容前端发送裸数组或 {rows: [...]} 两种形式
    if isinstance(data, list):
        rows = data
    elif isinstance(data, dict):
        rows = data.get('rows') or data.get('data') or []
    else:
        rows = []
    if not rows:
        return jsonify({'code': 400, 'message': '无有效数据', 'data': None})

    db = get_db()
    cursor = db.cursor()

    # 预加载所有合同编号 → id 映射
    cursor.execute("SELECT id, contract_no FROM contracts WHERE contract_no IS NOT NULL")
    no_map = {r['contract_no']: r['id'] for r in cursor.fetchall()}

    # 防重复：同一合同 + 验收日期 已存在则跳过
    cursor.execute("SELECT contract_id, acceptance_date FROM contract_acceptances")
    existing = {(r['contract_id'], r['acceptance_date']) for r in cursor.fetchall()}

    def _parse_amt(val):
        if val is None or val == '':
            return None
        try:
            return float(val)  # 元，无需单位转换
        except Exception:
            return None

    def _parse_date(val):
        if not val:
            return None
        if isinstance(val, datetime):
            return val.strftime('%Y-%m-%d')
        return str(val)[:10]

    success_count = 0
    fail_count = 0
    results = []
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    affected_contracts = set()  # 记录受影响的合同，用于联动同步

    for item in rows:
        row_data = item.get('data', item) if isinstance(item, dict) else {}
        row_index = item.get('row_index', 0) if isinstance(item, dict) else 0

        contract_no = str(row_data.get('contract_no') or '').strip()
        if not contract_no:
            fail_count += 1
            results.append({'row_index': row_index, 'success': False, 'message': '合同编号为空'})
            continue

        contract_id = no_map.get(contract_no)
        if not contract_id:
            fail_count += 1
            results.append({'row_index': row_index, 'success': False, 'message': f'合同编号不存在：{contract_no}'})
            continue

        acc_date = _parse_date(row_data.get('acceptance_date'))
        acc_amt_yuan = _parse_amt(row_data.get('acceptance_amount'))
        if not acc_date:
            fail_count += 1
            results.append({'row_index': row_index, 'success': False, 'message': '验收日期为空或格式错误'})
            continue
        if acc_amt_yuan is None or acc_amt_yuan == 0:
            fail_count += 1
            results.append({'row_index': row_index, 'success': False, 'message': '收入为空或为0（正数验收，负数核减）'})
            continue

        if (contract_id, acc_date) in existing:
            fail_count += 1
            results.append({'row_index': row_index, 'success': False, 'message': f'该合同已有 {acc_date} 的验收记录，跳过'})
            continue

        note = str(row_data.get('note') or row_data.get('acceptance_note') or '').strip()

        try:
            # 1) 创建验收记录
            cursor.execute(
                "INSERT INTO contract_acceptances (contract_id, acceptance_date, acceptance_amount, note, created_by, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (contract_id, acc_date, acc_amt_yuan, note, username, now)
            )
            existing.add((contract_id, acc_date))
            affected_contracts.add(contract_id)

            # 2) 可选：更新合同级字段（税额/业务方向）。
            # 收入/待验收额/验收日期由下方 _sync_contract_acceptance_fields 自动联动
            contract_updates = {}
            for field, parser in [
                ('tax_amount', _parse_amt),
                ('business_direction', lambda v: str(v).strip() if v else None),
            ]:
                if field in row_data:
                    parsed = parser(row_data[field])
                    if parsed is not None:
                        contract_updates[field] = parsed
            if contract_updates:
                set_clauses = ', '.join(f'{k}=?' for k in contract_updates)
                params = list(contract_updates.values()) + [contract_id]
                cursor.execute(f"UPDATE contracts SET {set_clauses} WHERE id=?", params)

            success_count += 1
            results.append({'row_index': row_index, 'success': True, 'message': '验收记录已创建'})
        except Exception as e:
            fail_count += 1
            results.append({'row_index': row_index, 'success': False, 'message': str(e)})

    # 联动同步：更新所有受影响合同的收入/待验收额/验收日期
    for cid in affected_contracts:
        try:
            _sync_contract_acceptance_fields(cursor, cid)
        except Exception:
            pass

    db.commit()
    if success_count > 0:
        record_operation_log(username, '导入', '验收管理', f'导入验收记录：成功{success_count}条，失败{fail_count}条')

    return jsonify({
        'code': 200,
        'message': '导入完成',
        'data': {
            'total': len(rows),
            'success_count': success_count,
            'fail_count': fail_count,
            'results': results
        }
    })


@contracts_bp.route('/api/acceptances/import-parse', methods=['POST'])
@token_required
def parse_acceptance_excel():
    """解析验收导入Excel文件，返回预览数据。"""
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请选择文件', 'data': None})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'code': 400, 'message': '请选择文件', 'data': None})

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'code': 400, 'message': '仅支持Excel文件（.xlsx/.xls）', 'data': None})

    try:
        from openpyxl import load_workbook
        import re

        wb = load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active

        def _norm(h):
            """规范化表头：去除所有空白字符（含全角空格、换行、制表符）。"""
            if h is None:
                return ''
            s = str(h)
            # 去除全角空格、普通空格、换行、回车、制表符及不可见字符
            s = re.sub(r'[\s\u3000\xa0\u200b\ufeff]+', '', s)
            return s

        # 字段匹配规则（精确匹配优先，再按包含关键词回退）
        # 收入即验收金额：收入列的值直接作为本次验收金额
        field_rules = [
            ('contract_no', ['合同编号', '合同号', '编号', '合同Code']),
            ('contract_name', ['合同名称', '合同名']),
            ('party_a', ['甲方']),
            ('total_amt', ['合同额(万)', '合同总额(万)', '合同额', '合同金额(万)', '合同金额', '合同总价(万)']),
            ('tax_amount', ['税额(万)', '税额', '税金']),
            ('acceptance_date', ['验收日期', '验收时间']),
            ('acceptance_amount', ['收入(万)', '收入', '验收金额(万)', '验收金额', '验收额(万)', '验收额', '本次验收金额']),
            ('note', ['验收情况', '验收备注', '备注', '说明']),
            ('pending_acceptance_amount', ['待验收合同额(万)', '待验收合同额', '待验收金额(万)', '待验收']),
            ('business_direction', ['业务方向']),
        ]

        def _match_field(norm_header):
            """返回该规范化表头对应的字段 key，无匹配返回 None。"""
            # 先精确匹配
            for key, aliases in field_rules:
                if norm_header in aliases:
                    return key
            # 再包含匹配（如表头为"合同编号 "已去除空白，或"甲方单位"含"甲方"）
            for key, aliases in field_rules:
                for alias in aliases:
                    if alias and alias in norm_header:
                        return key
            return None

        # 自动定位表头行：扫描前 5 行，找到含"合同编号"特征的那一行
        header_row_idx = None
        max_scan = min(5, ws.max_row or 1)
        for r in range(1, max_scan + 1):
            row_vals = [(_norm(c.value)) for c in ws[r]]
            # 检测是否有合同编号特征列
            matched = False
            for v in row_vals:
                if v and ('合同编号' in v or '合同号' in v or v == '编号'):
                    matched = True
                    break
            if matched:
                header_row_idx = r
                break

        if header_row_idx is None:
            # 回退：默认第1行为表头
            header_row_idx = 1

        raw_headers = [cell.value for cell in ws[header_row_idx]]
        norm_headers = [_norm(h) for h in raw_headers]
        col_map = {}
        for idx, nh in enumerate(norm_headers):
            if not nh:
                continue
            key = _match_field(nh)
            if key and key not in col_map:
                col_map[key] = idx

        if 'contract_no' not in col_map:
            detected = [h for h in norm_headers if h]
            return jsonify({
                'code': 400,
                'message': f'缺少必要列：合同编号。检测到的表头(第{header_row_idx}行)：{detected}',
                'data': None
            })
        if 'acceptance_date' not in col_map:
            return jsonify({'code': 400, 'message': '缺少必要列：验收日期', 'data': None})
        if 'acceptance_amount' not in col_map:
            return jsonify({'code': 400, 'message': '缺少必要列：收入', 'data': None})

        # 查询数据库已存在的验收记录（合同编号+验收日期），用于重复检测
        try:
            db = get_db()
            cur = db.cursor()
            cur.execute(
                "SELECT c.contract_no, ca.acceptance_date FROM contract_acceptances ca "
                "JOIN contracts c ON ca.contract_id = c.id WHERE c.contract_no IS NOT NULL"
            )
            db_existing = {(r['contract_no'], r['acceptance_date']) for r in cur.fetchall()}
        except Exception:
            db_existing = set()

        rows = []
        valid_count = 0
        # 文件内重复检测：记录本文件中已出现过的 (合同编号, 验收日期)
        file_seen = set()

        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row_data = {}
            errors = []

            for key, idx in col_map.items():
                value = ws.cell(row=row_idx, column=idx + 1).value
                if key == 'acceptance_date' and value:
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d')
                    else:
                        value = str(value)[:10]
                row_data[key] = value

            contract_no = str(row_data.get('contract_no') or '').strip()
            if not contract_no:
                errors.append('合同编号不能为空')

            acc_date = row_data.get('acceptance_date')
            if not acc_date:
                errors.append('验收日期不能为空')

            acc_amt = row_data.get('acceptance_amount')
            if acc_amt is None or acc_amt == '':
                errors.append('收入不能为空')
            else:
                try:
                    if float(acc_amt) == 0:
                        errors.append('收入不能为0（正数验收，负数核减）')
                except Exception:
                    errors.append('收入格式错误')

            # 重复检测：合同编号+验收日期 在数据库或本文件中已存在
            dup_key = (contract_no, str(acc_date).strip() if acc_date else '')
            if contract_no and acc_date:
                if dup_key in db_existing:
                    errors.append('系统中已存在该合同此日期的验收记录（重复导入）')
                elif dup_key in file_seen:
                    errors.append('文件内重复：该合同此日期已在本文件中出现')

            valid = len(errors) == 0
            if valid:
                valid_count += 1
                file_seen.add(dup_key)

            rows.append({
                'row_index': row_idx,
                'data': row_data,
                'valid': valid,
                'errors': errors
            })

        return jsonify({
            'code': 200,
            'message': '解析成功',
            'data': {
                'total': len(rows),
                'valid_count': valid_count,
                'invalid_count': len(rows) - valid_count,
                'rows': rows
            }
        })

    except Exception as e:
        return jsonify({'code': 500, 'message': f'解析失败：{str(e)}', 'data': None})