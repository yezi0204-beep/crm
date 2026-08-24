# -*- coding: utf-8 -*-
"""应用中心人员月度考核路由。
蓝图: appraisal_bp，url_prefix=/api/appraisal
接口:
  GET  /monthly                             月度考核总览（主任/院长）
  GET  /mine                                个人考核详情（所有登录用户）
  GET  /export                              导出月度Excel（主任/院长）
  GET  /config/<username>                   读取用户配置（主任/院长）
  POST /config                              保存用户配置（主任/院长，写操作日志）
"""
import io
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from extensions import (
    get_db, token_required, admin_required, appraisal_viewer_required, record_operation_log
)

appraisal_bp = Blueprint('appraisal', __name__)

SALES_ROLES = ('销售',)
APPRAISAL_DEPT = '应用中心'
MAX_RATE_PCT = 150.00
MIN_RATE_PCT = 0.00


def _is_dept_director(row):
    """部门主任：role='主任' 且设置了年度指标（承担部门指标）。
    部门月度指标=主任月度指标，部门年度指标=主任年度指标。
    不再要求 is_sales_override=1。"""
    if (row.get('role') or '') != '主任':
        return False
    annual = float(row.get('annual_target_amount') or 0)
    return annual > 0


# ========== 核心算法 ==========

def _is_sales_user(row):
    """身份判定：is_sales_override=1 强制销售；否则 role in SALES_ROLES → 销售。其他非销售。"""
    if row.get('is_sales_override'):
        return True
    return (row.get('role') or '') in SALES_ROLES


def _get_user_monthly_target(user, year, month, overrides):
    """返回指定用户某年月的目标金额（元）。优先覆盖值，否则 annual/12。"""
    if month in overrides:
        return float(overrides[month])
    annual = float(user.get('annual_target_amount') or 0)
    return round(annual / 12.0, 2) if annual > 0 else 0.0


def _get_user_cumulative_target(user, year, month, overrides):
    """1..month 累计目标（元）：用户设置的月度指标即为累计目标，不再逐月累加。"""
    return _get_user_monthly_target(user, year, month, overrides)


def _get_user_cumulative_actual(cur, username, year, month):
    """1..month 用户累计新签合同额。
    - 普通合同：按 total_amt 计（有合同级分成按 ratio%，无分成 owner 独享 100%）
    - 框架合同：按 acceptance_amount 计
      ① 有验收级分成：acceptance_amount × 验收级 ratio%
      ② 无验收级分成但有合同级分成：acceptance_amount × 合同级 ratio%
      ③ 无任何分成：owner 独享 acceptance_amount
    注意：sign_date/acceptance_date 可能是 '2026-01-05' 或 '2026-01-05T16:00:00.000Z'，
    用 substr(...,1,10) 统一取 YYYY-MM-DD 做比较。
    一月份新签已在去年结算，累计实际从2月1日开始计算。
    """
    start = f'{int(year):04d}-02-01'
    end = f'{int(year):04d}-{int(month):02d}-31'

    # ① 普通合同 + 合同级分成：total_amt × ratio%
    cur.execute(
        "SELECT COALESCE(SUM(c.total_amt * cc.ratio / 100.0), 0) s "
        "FROM contracts c JOIN contract_commissions cc ON c.id = cc.contract_id "
        "WHERE cc.username = ? AND substr(c.sign_date,1,10) >= ? AND substr(c.sign_date,1,10) <= ? "
        "AND c.sign_date IS NOT NULL "
        "AND COALESCE(c.is_framework, 0) = 0",
        (username, start, end)
    )
    normal_share = float(cur.fetchone()['s'] or 0)

    # ② 普通合同 + 无分成：owner 独享 total_amt
    cur.execute(
        "SELECT COALESCE(SUM(c.total_amt), 0) s FROM contracts c "
        "WHERE c.owner_id = ? AND substr(c.sign_date,1,10) >= ? AND substr(c.sign_date,1,10) <= ? "
        "AND c.sign_date IS NOT NULL "
        "AND c.id NOT IN (SELECT DISTINCT contract_id FROM contract_commissions) "
        "AND COALESCE(c.is_framework, 0) = 0",
        (username, start, end)
    )
    normal_solo = float(cur.fetchone()['s'] or 0)

    # ③ 框架合同 + 验收级分成：acceptance_amount × 验收级 ratio%
    cur.execute(
        "SELECT COALESCE(SUM(ca.acceptance_amount * ac.ratio / 100.0), 0) s "
        "FROM contract_acceptances ca "
        "JOIN contracts c ON ca.contract_id = c.id "
        "JOIN acceptance_commissions ac ON ca.id = ac.acceptance_id "
        "WHERE ac.username = ? AND substr(ca.acceptance_date,1,10) >= ? AND substr(ca.acceptance_date,1,10) <= ? "
        "AND COALESCE(c.is_framework, 0) = 1",
        (username, start, end)
    )
    fw_acc_share = float(cur.fetchone()['s'] or 0)

    # ④ 框架合同 + 合同级分成但无验收级分成：acceptance_amount × 合同级 ratio%
    cur.execute(
        "SELECT COALESCE(SUM(ca.acceptance_amount * cc.ratio / 100.0), 0) s "
        "FROM contract_acceptances ca "
        "JOIN contracts c ON ca.contract_id = c.id "
        "JOIN contract_commissions cc ON c.id = cc.contract_id "
        "WHERE cc.username = ? AND substr(ca.acceptance_date,1,10) >= ? AND substr(ca.acceptance_date,1,10) <= ? "
        "AND COALESCE(c.is_framework, 0) = 1 "
        "AND ca.id NOT IN (SELECT DISTINCT acceptance_id FROM acceptance_commissions)",
        (username, start, end)
    )
    fw_contract_share = float(cur.fetchone()['s'] or 0)

    # ⑤ 框架合同 + 无任何分成：owner 独享 acceptance_amount
    cur.execute(
        "SELECT COALESCE(SUM(ca.acceptance_amount), 0) s "
        "FROM contract_acceptances ca "
        "JOIN contracts c ON ca.contract_id = c.id "
        "WHERE c.owner_id = ? AND substr(ca.acceptance_date,1,10) >= ? AND substr(ca.acceptance_date,1,10) <= ? "
        "AND c.id NOT IN (SELECT DISTINCT contract_id FROM contract_commissions) "
        "AND ca.id NOT IN (SELECT DISTINCT acceptance_id FROM acceptance_commissions) "
        "AND COALESCE(c.is_framework, 0) = 1",
        (username, start, end)
    )
    fw_solo = float(cur.fetchone()['s'] or 0)

    return round(normal_share + normal_solo + fw_acc_share + fw_contract_share + fw_solo, 2)


def _get_user_cumulative_details(cur, username, year, month):
    """返回用户累计实际明细列表，每项包含合同/验收信息、分成信息、计算公式。
    返回结构:
      {
        "total": 123456.78,
        "items": [
          {
            "type": "normal_share",       // 分类: normal_share/normal_solo/fw_acc_share/fw_contract_share/fw_solo
            "type_name": "普通合同-分成",
            "contract_id": 10,
            "contract_no": "HT2024-001",
            "contract_name": "XX项目",
            "is_framework": 0,
            "sign_date": "2026-02-15",
            "base_amount": 1000000.00,    // 合同总额或验收额
            "base_label": "合同额",
            "commission_type": "contract", // contract/acceptance/none
            "commission_members": [        // 分成人员列表
              {"username":"sales_a","name":"销售A","ratio":60.0,"amount":600000.00},
            ],
            "my_ratio": 60.0,             // 当前用户分成的比例
            "my_amount": 600000.00,       // 当前用户分到的金额
            "formula": "1000000 × 60% = 600000"  // 计算公式
          },
          ...
        ]
      }
    """
    start = f'{int(year):04d}-02-01'
    end = f'{int(year):04d}-{int(month):02d}-31'
    items = []
    total = 0.0

    # ① 普通合同 + 合同级分成
    cur.execute(
        "SELECT c.id, c.contract_no, c.contract_name, c.total_amt, c.sign_date, "
        "c.is_framework, cc.ratio, cc.username as comm_user, u.name as comm_name "
        "FROM contracts c JOIN contract_commissions cc ON c.id = cc.contract_id "
        "LEFT JOIN users u ON cc.username = u.username "
        "WHERE cc.username = ? AND substr(c.sign_date,1,10) >= ? AND substr(c.sign_date,1,10) <= ? "
        "AND c.sign_date IS NOT NULL "
        "AND COALESCE(c.is_framework, 0) = 0 "
        "ORDER BY c.sign_date",
        (username, start, end)
    )
    for r in cur.fetchall():
        base = float(r['total_amt'] or 0)
        ratio = float(r['ratio'] or 0)
        amt = round(base * ratio / 100.0, 2)
        total += amt
        items.append({
            'type': 'normal_share',
            'type_name': '普通合同-分成',
            'contract_id': r['id'],
            'contract_no': r['contract_no'] or '',
            'contract_name': r['contract_name'] or '',
            'is_framework': 0,
            'sign_date': r['sign_date'] or '',
            'base_amount': round(base, 2),
            'base_label': '合同额',
            'commission_type': 'contract',
            'commission_members': [{'username': username, 'name': r['comm_name'] or username, 'ratio': ratio, 'amount': amt}],
            'my_ratio': ratio,
            'my_amount': amt,
            'formula': f'{base:.2f} × {ratio}% = {amt:.2f}',
        })

    # ② 普通合同 + 无分成（owner独享）
    cur.execute(
        "SELECT c.id, c.contract_no, c.contract_name, c.total_amt, c.sign_date, c.is_framework "
        "FROM contracts c "
        "WHERE c.owner_id = ? AND substr(c.sign_date,1,10) >= ? AND substr(c.sign_date,1,10) <= ? "
        "AND c.sign_date IS NOT NULL "
        "AND c.id NOT IN (SELECT DISTINCT contract_id FROM contract_commissions) "
        "AND COALESCE(c.is_framework, 0) = 0 "
        "ORDER BY c.sign_date",
        (username, start, end)
    )
    for r in cur.fetchall():
        base = float(r['total_amt'] or 0)
        amt = round(base, 2)
        total += amt
        items.append({
            'type': 'normal_solo',
            'type_name': '普通合同-独享',
            'contract_id': r['id'],
            'contract_no': r['contract_no'] or '',
            'contract_name': r['contract_name'] or '',
            'is_framework': 0,
            'sign_date': r['sign_date'] or '',
            'base_amount': amt,
            'base_label': '合同额',
            'commission_type': 'none',
            'commission_members': [],
            'my_ratio': 100.0,
            'my_amount': amt,
            'formula': f'{amt:.2f} × 100% = {amt:.2f}',
        })

    # ③ 框架合同 + 验收级分成
    cur.execute(
        "SELECT ca.id as acc_id, ca.acceptance_amount, ca.acceptance_date, ca.note as acc_note, "
        "c.id, c.contract_no, c.contract_name, c.is_framework, "
        "ac.ratio, ac.username as comm_user, u.name as comm_name "
        "FROM contract_acceptances ca "
        "JOIN contracts c ON ca.contract_id = c.id "
        "JOIN acceptance_commissions ac ON ca.id = ac.acceptance_id "
        "LEFT JOIN users u ON ac.username = u.username "
        "WHERE ac.username = ? AND substr(ca.acceptance_date,1,10) >= ? AND substr(ca.acceptance_date,1,10) <= ? "
        "AND COALESCE(c.is_framework, 0) = 1 "
        "ORDER BY ca.acceptance_date",
        (username, start, end)
    )
    for r in cur.fetchall():
        base = float(r['acceptance_amount'] or 0)
        ratio = float(r['ratio'] or 0)
        amt = round(base * ratio / 100.0, 2)
        total += amt
        items.append({
            'type': 'fw_acc_share',
            'type_name': '框架合同-验收分成',
            'contract_id': r['id'],
            'contract_no': r['contract_no'] or '',
            'contract_name': r['contract_name'] or '',
            'is_framework': 1,
            'sign_date': r['acceptance_date'] or '',
            'base_amount': round(base, 2),
            'base_label': '验收额',
            'commission_type': 'acceptance',
            'commission_members': [{'username': username, 'name': r['comm_name'] or username, 'ratio': ratio, 'amount': amt}],
            'my_ratio': ratio,
            'my_amount': amt,
            'formula': f'{base:.2f} × {ratio}% = {amt:.2f}',
        })

    # ④ 框架合同 + 合同级分成但无验收级分成
    cur.execute(
        "SELECT ca.id as acc_id, ca.acceptance_amount, ca.acceptance_date, "
        "c.id, c.contract_no, c.contract_name, c.is_framework, "
        "cc.ratio, cc.username as comm_user, u.name as comm_name "
        "FROM contract_acceptances ca "
        "JOIN contracts c ON ca.contract_id = c.id "
        "JOIN contract_commissions cc ON c.id = cc.contract_id "
        "LEFT JOIN users u ON cc.username = u.username "
        "WHERE cc.username = ? AND substr(ca.acceptance_date,1,10) >= ? AND substr(ca.acceptance_date,1,10) <= ? "
        "AND COALESCE(c.is_framework, 0) = 1 "
        "AND ca.id NOT IN (SELECT DISTINCT acceptance_id FROM acceptance_commissions) "
        "ORDER BY ca.acceptance_date",
        (username, start, end)
    )
    for r in cur.fetchall():
        base = float(r['acceptance_amount'] or 0)
        ratio = float(r['ratio'] or 0)
        amt = round(base * ratio / 100.0, 2)
        total += amt
        items.append({
            'type': 'fw_contract_share',
            'type_name': '框架合同-合同分成',
            'contract_id': r['id'],
            'contract_no': r['contract_no'] or '',
            'contract_name': r['contract_name'] or '',
            'is_framework': 1,
            'sign_date': r['acceptance_date'] or '',
            'base_amount': round(base, 2),
            'base_label': '验收额',
            'commission_type': 'contract',
            'commission_members': [{'username': username, 'name': r['comm_name'] or username, 'ratio': ratio, 'amount': amt}],
            'my_ratio': ratio,
            'my_amount': amt,
            'formula': f'{base:.2f} × {ratio}% = {amt:.2f}',
        })

    # ⑤ 框架合同 + 无任何分成（owner独享验收额）
    cur.execute(
        "SELECT ca.id as acc_id, ca.acceptance_amount, ca.acceptance_date, "
        "c.id, c.contract_no, c.contract_name, c.is_framework "
        "FROM contract_acceptances ca "
        "JOIN contracts c ON ca.contract_id = c.id "
        "WHERE c.owner_id = ? AND substr(ca.acceptance_date,1,10) >= ? AND substr(ca.acceptance_date,1,10) <= ? "
        "AND c.id NOT IN (SELECT DISTINCT contract_id FROM contract_commissions) "
        "AND ca.id NOT IN (SELECT DISTINCT acceptance_id FROM acceptance_commissions) "
        "AND COALESCE(c.is_framework, 0) = 1 "
        "ORDER BY ca.acceptance_date",
        (username, start, end)
    )
    for r in cur.fetchall():
        base = float(r['acceptance_amount'] or 0)
        amt = round(base, 2)
        total += amt
        items.append({
            'type': 'fw_solo',
            'type_name': '框架合同-独享',
            'contract_id': r['id'],
            'contract_no': r['contract_no'] or '',
            'contract_name': r['contract_name'] or '',
            'is_framework': 1,
            'sign_date': r['acceptance_date'] or '',
            'base_amount': amt,
            'base_label': '验收额',
            'commission_type': 'none',
            'commission_members': [],
            'my_ratio': 100.0,
            'my_amount': amt,
            'formula': f'{amt:.2f} × 100% = {amt:.2f}',
        })

    return {'total': round(total, 2), 'items': items}


def _load_monthly_overrides(cur, username, year):
    """返回 {1: amount, 2: amount, ...}，只包含覆盖过的月份。"""
    cur.execute(
        "SELECT month, target_amount FROM monthly_targets "
        "WHERE username = ? AND year = ?",
        (username, int(year))
    )
    return {int(r['month']): float(r['target_amount']) for r in cur.fetchall()}


def _load_all_appraisal_users(cur):
    """读取 users 表（在职 + 离职也带，允许考核历史）。返回列表字典。"""
    cur.execute(
        "SELECT username, name, role, department, status,"
        " basic_salary, base_performance, annual_target_amount, is_sales_override"
        " FROM users ORDER BY department, username"
    )
    return [dict(r) for r in cur.fetchall()]


def _is_appraisal_scope(u):
    """考核范围：应用中心（按部门）。"""
    return (u.get('department') or '') == APPRAISAL_DEPT


def build_monthly_rows(cur, year, month):
    """构造月度考核行 + avg_sales_rate_pct。"""
    users = _load_all_appraisal_users(cur)
    # 只保留应用中心的人（主任/非销售/销售都在内），非应用中心剔除（普通用户自己查 /mine 用）
    rows = []
    sales_rates_with_target = []  # 有指标的销售的rate（不含主任）
    # 先算出每个销售个人的实际额
    for u in users:
        if not _is_appraisal_scope(u):
            continue
        username = u['username']
        overrides = _load_monthly_overrides(cur, username, year)
        cum_target = _get_user_cumulative_target(u, year, month, overrides)
        cum_actual = _get_user_cumulative_actual(cur, username, year, month)
        is_sales = _is_sales_user(u)
        is_director = _is_dept_director(u)
        month_target = _get_user_monthly_target(u, year, month, overrides)

        if is_sales:
            if cum_target > 0:
                rate = cum_actual * 100.0 / cum_target
            else:
                rate = 0.0
            rate = max(MIN_RATE_PCT, min(MAX_RATE_PCT, rate))
        else:
            # 非销售先占位，后面用销售均值填
            rate = None

        basic = float(u.get('basic_salary') or 0)
        base_perf = float(u.get('base_performance') or 0)
        row = {
            'username': username,
            'name': u.get('name') or username,
            'role': u.get('role') or '',
            'department': u.get('department') or '',
            'status': u.get('status') or '',
            'is_sales': is_sales,
            'is_director': is_director,
            'annual_target_amount': float(u.get('annual_target_amount') or 0),
            'monthly_target_amt': round(month_target, 2),
            'cumulative_target_amt': round(cum_target, 2),
            'cumulative_actual_amt': round(cum_actual, 2),
            'actual_amt': round(cum_actual, 2),   # 兼容测试断言键名
            'target_amt': round(cum_target, 2),  # 兼容测试断言键名
            'rate_pct': round(rate, 2) if rate is not None else None,
            'basic_salary': round(basic, 2),
            'base_performance': round(base_perf, 2),
            'perf_pay': 0.0,
            'total_pay': 0.0,
        }
        rows.append(row)
        # 主任不参与销售均值（主任的完成率=部门完成率）
        if is_sales and cum_target > 0 and not is_director:
            sales_rates_with_target.append(rate)

    # 部门累计新签 = 所有应用中心人员的业绩累计之和（主任+销售+非销售）
    dept_actual = sum(r['cumulative_actual_amt'] for r in rows)
    for r in rows:
        if r['is_director']:
            r['cumulative_actual_amt'] = round(dept_actual, 2)
            r['actual_amt'] = round(dept_actual, 2)
            target = r['cumulative_target_amt']
            if target > 0:
                dir_rate = dept_actual * 100.0 / target
            else:
                dir_rate = 0.0
            r['rate_pct'] = round(max(MIN_RATE_PCT, min(MAX_RATE_PCT, dir_rate)), 2)

    # 计算销售均值（不含主任）
    if sales_rates_with_target:
        avg_sales_rate = round(sum(sales_rates_with_target) / len(sales_rates_with_target), 2)
    else:
        avg_sales_rate = 0.0

    # 补齐非销售 rate 和 薪资（主任的 rate 已按部门完成率算好，不被覆盖）
    for r in rows:
        if not r['is_sales'] and not r['is_director']:
            r['rate_pct'] = round(avg_sales_rate, 2)
        rate = r['rate_pct'] or 0
        r['perf_pay'] = round(r['base_performance'] * rate / 100.0, 2)
        r['total_pay'] = round(r['basic_salary'] + r['perf_pay'], 2)

    # 汇总行兼容字段：actual_amt/target_amt 为累计值（已赋值）
    # 部门完成率 = 主任的完成率（主任承担部门指标）；无主任则为 None
    dept_rate = None
    dept_monthly_target = 0.0
    dept_cumulative_actual = 0.0
    for r in rows:
        if r['is_director']:
            dept_rate = r['rate_pct']
            dept_monthly_target = r['monthly_target_amt']
            dept_cumulative_actual = r['cumulative_actual_amt']
            break
    return rows, round(avg_sales_rate, 2), dept_rate, round(dept_monthly_target, 2), round(dept_cumulative_actual, 2)


# ========== 接口 ==========

@appraisal_bp.route('/monthly', methods=['GET'])
@appraisal_viewer_required
def monthly_overview():
    """GET ?year=2026&month=8  主任/院长/人力：应用中心全体月度考核总览"""
    try:
        year = int(request.args.get('year') or datetime.now().year)
        month = int(request.args.get('month') or datetime.now().month)
    except ValueError:
        return jsonify({'code': 400, 'message': '年月参数非法', 'data': None})
    if month < 1 or month > 12:
        return jsonify({'code': 400, 'message': '月份需在1-12之间', 'data': None})
    db = get_db()
    cur = db.cursor()
    rows, avg_rate, dept_rate, dept_target, dept_actual = build_monthly_rows(cur, year, month)
    return jsonify({
        'code': 200,
        'message': 'OK',
        'data': {
            'year': year, 'month': month,
            'avg_sales_rate_pct': avg_rate,
            'dept_rate_pct': dept_rate,
            'dept_monthly_target': dept_target,
            'dept_cumulative_actual': dept_actual,
            'rows': rows,
        }
    })


@appraisal_bp.route('/mine', methods=['GET'])
@token_required
def mine_appraisal():
    """GET ?year=2026&month=8  登录用户查看自己的考核详情（非应用中心也能查自己）。"""
    try:
        year = int(request.args.get('year') or datetime.now().year)
        month = int(request.args.get('month') or datetime.now().month)
    except ValueError:
        return jsonify({'code': 400, 'message': '年月参数非法', 'data': None})
    if month < 1 or month > 12:
        return jsonify({'code': 400, 'message': '月份需在1-12之间', 'data': None})
    me = request.current_user
    username = me['username']
    db = get_db()
    cur = db.cursor()

    # 先拿所有应用中心的销售均值，再构造自己的行（自己非应用中心销售角色也算销售）
    all_rows, avg_sales_rate, dept_rate, _dt, _da = build_monthly_rows(cur, year, month)
    # 看是否在 rows 里（应用中心）
    my_row = next((r for r in all_rows if r['username'] == username), None)
    if not my_row:
        # 非应用中心：单独构造（销售按自己指标算，非销售按平均）
        cur.execute(
            "SELECT username, name, role, department, status,"
            " basic_salary, base_performance, annual_target_amount, is_sales_override"
            " FROM users WHERE username=?", (username,)
        )
        u = cur.fetchone()
        if not u:
            return jsonify({'code': 404, 'message': '用户不存在', 'data': None})
        u = dict(u)
        overrides = _load_monthly_overrides(cur, username, year)
        cum_target = _get_user_cumulative_target(u, year, month, overrides)
        cum_actual = _get_user_cumulative_actual(cur, username, year, month)
        month_target = _get_user_monthly_target(u, year, month, overrides)
        is_sales = _is_sales_user(u)
        if is_sales:
            rate = (cum_actual / cum_target * 100.0) if cum_target > 0 else 0.0
            rate = max(MIN_RATE_PCT, min(MAX_RATE_PCT, rate))
        else:
            rate = avg_sales_rate
        basic = float(u.get('basic_salary') or 0)
        base_perf = float(u.get('base_performance') or 0)
        perf = round(base_perf * rate / 100.0, 2)
        my_row = {
            'username': u['username'], 'name': u.get('name') or username,
            'role': u.get('role') or '', 'department': u.get('department') or '',
            'is_sales': is_sales,
            'annual_target_amount': float(u.get('annual_target_amount') or 0),
            'monthly_target_amt': round(month_target, 2),
            'cumulative_target_amt': round(cum_target, 2),
            'cumulative_actual_amt': round(cum_actual, 2),
            'actual_amt': round(cum_actual, 2),
            'target_amt': round(cum_target, 2),
            'rate_pct': round(rate, 2),
            'basic_salary': round(basic, 2),
            'base_performance': round(base_perf, 2),
            'perf_pay': perf,
            'total_pay': round(basic + perf, 2),
        }
    return jsonify({
        'code': 200, 'message': 'OK',
        'data': {
            'year': year, 'month': month,
            'avg_sales_rate_pct': avg_sales_rate,
            'myself': my_row,
        }
    })


@appraisal_bp.route('/details/<username>', methods=['GET'])
@appraisal_viewer_required
def user_details(username):
    """GET ?year=2026&month=8  返回指定用户的累计实际明细（项目明细+分成明细+计算公式）。
    主任/院长/人力可查看任意用户；普通用户只能查自己。
    """
    try:
        year = int(request.args.get('year') or datetime.now().year)
        month = int(request.args.get('month') or datetime.now().month)
    except ValueError:
        return jsonify({'code': 400, 'message': '年月参数非法', 'data': None})
    if month < 1 or month > 12:
        return jsonify({'code': 400, 'message': '月份需在1-12之间', 'data': None})
    me = request.current_user
    # 普通用户（非主任/院长/人力）只能查自己
    if me['username'] != username and me.get('role') not in ('主任', '院长', '人力'):
        return jsonify({'code': 403, 'message': '只能查看自己的明细', 'data': None})

    db = get_db()
    cur = db.cursor()
    # 查用户基本信息
    cur.execute("SELECT username, name, role, department FROM users WHERE username=?", (username,))
    u = cur.fetchone()
    if not u:
        return jsonify({'code': 404, 'message': '用户不存在', 'data': None})
    u = dict(u)
    # 判断身份
    is_sales = _is_sales_user(u)
    is_director = _is_dept_director(u)
    # 获取明细
    details = _get_user_cumulative_details(cur, username, year, month)

    # 如果是主任，附加部门所有销售的明细汇总
    dept_members = []
    if is_director:
        cur.execute(
            "SELECT username, name FROM users WHERE department=? AND status='在职' "
            "ORDER BY username",
            (u.get('department') or '',)
        )
        for r in cur.fetchall():
            member = dict(r)
            member_details = _get_user_cumulative_details(cur, member['username'], year, month)
            dept_members.append({
                'username': member['username'],
                'name': member['name'],
                'total': member_details['total'],
                'items': member_details['items'],
            })

    # 构造公式说明
    formula_explain = []
    if details['items']:
        amounts = ' + '.join([f"{it['my_amount']:.2f}" for it in details['items']])
        formula_explain.append(f"累计实际 = {amounts} = {details['total']:.2f}")
    else:
        formula_explain.append(f"累计实际 = 0.00（{year}年{month}月无符合条件的合同/验收记录）")

    if is_director:
        dept_total = sum(m['total'] for m in dept_members)
        formula_explain.insert(0, f"部门累计实际 = 部门所有销售累计实际之和 = {dept_total:.2f}")

    return jsonify({
        'code': 200, 'message': 'OK',
        'data': {
            'year': year, 'month': month,
            'username': username,
            'name': u.get('name') or username,
            'role': u.get('role') or '',
            'department': u.get('department') or '',
            'is_sales': is_sales,
            'is_director': is_director,
            'total': details['total'],
            'items': details['items'],
            'dept_members': dept_members,
            'formula_explain': formula_explain,
        }
    })


@appraisal_bp.route('/export', methods=['GET'])
@admin_required
def export_monthly():
    """GET ?year=2026&month=8 → xlsx 下载"""
    try:
        year = int(request.args.get('year') or datetime.now().year)
        month = int(request.args.get('month') or datetime.now().month)
    except ValueError:
        return jsonify({'code': 400, 'message': '年月参数非法', 'data': None})
    if month < 1 or month > 12:
        return jsonify({'code': 400, 'message': '月份需在1-12之间', 'data': None})

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except Exception as e:
        return jsonify({'code': 500, 'message': f'缺少 openpyxl：{e}', 'data': None})

    db = get_db()
    cur = db.cursor()
    rows, avg_rate, dept_rate, _dt, _da = build_monthly_rows(cur, year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = f'{year}年{month}月考核'
    headers = ['姓名', '用户名', '部门', '角色', '身份',
               '年度指标', '当月指标', '累计实际', '完成率(%)',
               '基本工资', '基础绩效', '绩效工资', '月应发合计']
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='DDEBF7')
        c.alignment = Alignment(horizontal='center', vertical='center')

    for r in rows:
        ws.append([
            r['name'], r['username'], r['department'], r['role'],
            '销售' if r['is_sales'] else '非销售',
            r['annual_target_amount'], r['monthly_target_amt'],
            r['cumulative_actual_amt'],
            r['rate_pct'],
            r['basic_salary'], r['base_performance'],
            r['perf_pay'], r['total_pay'],
        ])
    # 汇总行
    ws.append(['', '', '', '', '销售平均完成率', '', '', '', avg_rate, '', '', '', ''])

    # 写操作日志
    payload = request.current_user
    try:
        record_operation_log(payload['username'], '导出', '月度考核',
                             f'导出{year}年{month}月应用中心考核表（{len(rows)}人）')
    except Exception:
        pass

    # 列宽
    widths = [12, 14, 10, 8, 8, 12, 12, 12, 12, 10, 10, 10, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname_utf8 = f'应用中心{year}年{month}月考核.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname_utf8,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@appraisal_bp.route('/yearly', methods=['GET'])
@admin_required
def yearly_trend():
    """GET ?year=2026  年度趋势：1-12月每月部门完成率 + 每个销售的月度完成率。
    返回格式：
      {
        "year": 2026,
        "dept_rates": {"1": null, "2": 89.6, ...},  // 部门完成率（无主任则为 null）
        "sales_trend": [
          {"username":"sales_a","name":"销售A","rates":{"1":0,"2":80.0,...}},
          ...
        ]
      }
    """
    try:
        year = int(request.args.get('year') or datetime.now().year)
    except ValueError:
        return jsonify({'code': 400, 'message': '年参数非法', 'data': None})
    db = get_db()
    cur = db.cursor()

    dept_rates = {}   # {1: rate_or_null, ...}
    sales_map = {}    # {username: {"name":..., "rates": {1: rate, ...}}}

    for m in range(1, 13):
        rows, _avg, dept_rate, _dt, _da = build_monthly_rows(cur, year, m)
        dept_rates[m] = dept_rate
        for r in rows:
            if not r['is_sales']:
                continue  # 只展示销售的月度完成率（主任的部门完成率单独在 dept_rates）
            if r['is_director']:
                continue  # 主任不作为个体销售展示
            uname = r['username']
            if uname not in sales_map:
                sales_map[uname] = {
                    'username': uname,
                    'name': r.get('name') or uname,
                    'rates': {},
                }
            sales_map[uname]['rates'][m] = r.get('rate_pct')

    # 转为 JSON 友好的字符串键
    dept_rates_str = {str(k): v for k, v in dept_rates.items()}
    sales_trend = []
    for uname in sorted(sales_map.keys()):
        item = sales_map[uname]
        item['rates'] = {str(k): v for k, v in item['rates'].items()}
        sales_trend.append(item)

    return jsonify({
        'code': 200,
        'message': 'OK',
        'data': {
            'year': year,
            'dept_rates': dept_rates_str,
            'sales_trend': sales_trend,
        }
    })


@appraisal_bp.route('/config/<username>', methods=['GET'])
@admin_required
def get_user_config(username):
    """GET ?year=2026  读取用户基础配置 + 月度覆盖值。"""
    try:
        year = int(request.args.get('year') or datetime.now().year)
    except ValueError:
        return jsonify({'code': 400, 'message': '年参数非法', 'data': None})
    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT username, name, role, department,"
        " basic_salary, base_performance, annual_target_amount, is_sales_override"
        " FROM users WHERE username=?", (username,)
    )
    u = cur.fetchone()
    if not u:
        return jsonify({'code': 404, 'message': '用户不存在', 'data': None})
    u = dict(u)
    overrides = _load_monthly_overrides(cur, username, year)
    # 默认分解
    annual = float(u['annual_target_amount'] or 0)
    default_monthly = {}
    for m in range(1, 13):
        default_monthly[m] = round(annual / 12.0, 2) if annual > 0 else 0.0
    return jsonify({
        'code': 200, 'message': 'OK',
        'data': {
            'username': u['username'], 'name': u.get('name') or username,
            'role': u.get('role') or '', 'department': u.get('department') or '',
            'basic_salary': float(u.get('basic_salary') or 0),
            'base_performance': float(u.get('base_performance') or 0),
            'annual_target_amount': float(u.get('annual_target_amount') or 0),
            'is_sales_override': int(u.get('is_sales_override') or 0),
            'year': year,
            'default_monthly': default_monthly,  # 12个月默认值
            'monthly_overrides': overrides,
        }
    })


@appraisal_bp.route('/config', methods=['POST'])
@admin_required
def save_user_config():
    """POST 保存配置（写 users 4 字段 + 写入/更新 monthly_targets 覆盖）。"""
    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip()
    if not username:
        return jsonify({'code': 400, 'message': 'username 必传', 'data': None})
    try:
        year = int(data.get('year') or datetime.now().year)
    except Exception:
        return jsonify({'code': 400, 'message': 'year 非法', 'data': None})
    basic = float(data.get('basic_salary') or 0)
    base_perf = float(data.get('base_performance') or 0)
    annual = float(data.get('annual_target_amount') or 0)
    override_flag = 1 if int(data.get('is_sales_override') or 0) else 0
    overrides = data.get('monthly_overrides') or {}
    if not isinstance(overrides, dict):
        return jsonify({'code': 400, 'message': 'monthly_overrides 必须是字典', 'data': None})

    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT 1 FROM users WHERE username=?", (username,))
    if not cur.fetchone():
        return jsonify({'code': 404, 'message': '用户不存在', 'data': None})

    # 更新 users 4 字段
    cur.execute(
        "UPDATE users SET basic_salary=?, base_performance=?, annual_target_amount=?, is_sales_override=?"
        " WHERE username=?",
        (basic, base_perf, annual, override_flag, username)
    )

    # 处理覆盖值：UPSERT 到 monthly_targets，传入外字段删除
    # 先查现存覆盖
    cur.execute(
        "SELECT id, month FROM monthly_targets WHERE username=? AND year=?",
        (username, year)
    )
    existing = {int(r['month']): r['id'] for r in cur.fetchall()}
    input_months = set()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    operator = request.current_user.get('username', '')
    for m, amt in overrides.items():
        try:
            mi = int(m)
            av = float(amt)
        except Exception:
            return jsonify({'code': 400, 'message': f'monthly_overrides 非法：m={m} amt={amt}', 'data': None})
        if mi < 1 or mi > 12:
            return jsonify({'code': 400, 'message': f'月份{m}超出范围', 'data': None})
        input_months.add(mi)
        # 与 annual/12 相等则删除覆盖（不需要保留）
        default_amt = round(annual / 12.0, 2) if annual > 0 else 0.0
        if round(av, 2) == round(default_amt, 2):
            if mi in existing:
                cur.execute("DELETE FROM monthly_targets WHERE id=?", (existing[mi],))
            continue
        if mi in existing:
            cur.execute(
                "UPDATE monthly_targets SET target_amount=?, updated_by=?, updated_at=? WHERE id=?",
                (av, operator, now, existing[mi])
            )
        else:
            cur.execute(
                "INSERT INTO monthly_targets (username, year, month, target_amount, updated_by, updated_at)"
                " VALUES (?, ?, ?, ?, ?, ?)",
                (username, year, mi, av, operator, now)
            )
    # 未传入但存在的覆盖 → 删除
    for m_existing, rid in existing.items():
        if m_existing not in input_months:
            cur.execute("DELETE FROM monthly_targets WHERE id=?", (rid,))

    db.commit()

    try:
        record_operation_log(
            operator, '配置指标', '月度考核',
            f'配置用户[{username}] 年度指标={annual} 基本={basic} 绩效={base_perf}'
            f' is_sales_override={override_flag} 覆盖月={sorted(input_months)} (year={year})'
        )
    except Exception:
        pass

    return jsonify({'code': 200, 'message': '配置保存成功', 'data': None})
