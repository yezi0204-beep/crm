from flask import Flask, jsonify, request, send_from_directory, g
import sqlite3
import bcrypt
import os
import hashlib
import uuid
from datetime import datetime, timedelta

app = Flask(__name__)

@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,PATCH,OPTIONS')
    response.headers.add('Access-Control-Allow-Credentials', 'true')
    return response

@app.route('/api/', methods=['OPTIONS'])
def options():
    return jsonify({'code': 200, 'message': 'OK', 'data': None})

SECRET_KEY = "crm_secret_key_2026"
DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "crm_app.db")
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads", "contracts")

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH, check_same_thread=False)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(error):
    if hasattr(g, 'db'):
        g.db.close()


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


def check_password(password: str, hash: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hash.encode('utf-8'))


def create_token(username: str, name: str, role: str) -> str:
    token = str(uuid.uuid4())
    expires = datetime.now() + timedelta(hours=24)
    
    db = get_db()
    cursor = db.cursor()
    cursor.execute('''
        INSERT INTO tokens (token, username, name, role, expires)
        VALUES (?, ?, ?, ?, ?)
    ''', (token, username, name, role, expires.strftime('%Y-%m-%d %H:%M:%S')))
    db.commit()
    
    return token


def verify_token(token: str):
    if not token:
        return None
    
    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT * FROM tokens WHERE token = ?', (token,))
    row = cursor.fetchone()
    
    if not row:
        return None
    
    expires = datetime.strptime(row['expires'], '%Y-%m-%d %H:%M:%S')
    if datetime.now() > expires:
        cursor.execute('DELETE FROM tokens WHERE token = ?', (token,))
        db.commit()
        return None
    
    return {
        'username': row['username'],
        'name': row['name'],
        'role': row['role'],
        'expires': expires
    }


@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT username, password_hash, name, role FROM users WHERE username = ?", (username,))
    row = cursor.fetchone()
    
    if row and check_password(password, row['password_hash']):
        token = create_token(row['username'], row['name'], row['role'])
        return jsonify({
            'code': 200,
            'message': '登录成功',
            'data': {
                'token': token,
                'username': row['username'],
                'name': row['name'],
                'role': row['role']
            }
        })
    return jsonify({'code': 401, 'message': '账号或密码错误', 'data': None})


@app.route('/api/auth/info', methods=['GET'])
def get_user_info():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT role FROM user_roles WHERE username = ?", (payload['username'],))
    rows = cursor.fetchall()
    roles = [r['role'] for r in rows] if rows else [payload['role']]
    
    return jsonify({
        'code': 200,
        'message': 'success',
        'data': {
            'username': payload['username'],
            'name': payload['name'],
            'role': payload['role'],
            'roles': roles
        }
    })


@app.route('/api/auth/logout', methods=['POST'])
def logout():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    if token:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('DELETE FROM tokens WHERE token = ?', (token,))
        db.commit()
    return jsonify({'code': 200, 'message': '退出成功', 'data': None})


@app.route('/api/contracts', methods=['GET'])
def get_contracts():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    username = payload['username']
    role = payload['role']
    sort_field = request.args.get('sort_field', 'sign_date')
    sort_order = request.args.get('sort_order', 'desc')
    
    db = get_db()
    cursor = db.cursor()
    
    valid_fields = ['sign_date', 'total_amt', 'paid_amt', 'contract_name', 'contract_no', 'pending_amt']
    if sort_field not in valid_fields:
        sort_field = 'sign_date'
    
    sort_direction = 'DESC' if sort_order.lower() in ['desc', 'descending'] else 'ASC'
    order_by = f"{sort_field} {sort_direction}"
    
    if role == '主任' or role == '院长':
        if sort_field == 'pending_amt':
            cursor.execute(f"""
                SELECT c.*, u.name as owner_name, 
                       (COALESCE(c.total_amt, 0) - COALESCE(c.paid_amt, 0)) as pending_amt
                FROM contracts c 
                LEFT JOIN users u ON c.owner_id = u.username 
                ORDER BY pending_amt {sort_direction}
            """)
        else:
            cursor.execute(f"""
                SELECT c.*, u.name as owner_name 
                FROM contracts c 
                LEFT JOIN users u ON c.owner_id = u.username 
                ORDER BY c.{order_by}
            """)
    else:
        if sort_field == 'pending_amt':
            cursor.execute(f"""
                SELECT c.*, u.name as owner_name, 
                       (COALESCE(c.total_amt, 0) - COALESCE(c.paid_amt, 0)) as pending_amt
                FROM contracts c 
                LEFT JOIN users u ON c.owner_id = u.username 
                WHERE c.owner_id = ? 
                ORDER BY pending_amt {sort_direction}
            """, (username,))
        else:
            cursor.execute(f"""
                SELECT c.*, u.name as owner_name 
                FROM contracts c 
                LEFT JOIN users u ON c.owner_id = u.username 
                WHERE c.owner_id = ? 
                ORDER BY c.{order_by}
            """, (username,))
    
    rows = cursor.fetchall()
    contracts = []
    for row in rows:
        contracts.append(dict(row))
    
    return jsonify({'code': 200, 'message': 'success', 'data': contracts})


@app.route('/api/contracts/check-no', methods=['GET'])
def check_contract_no():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    contract_no = request.args.get('contract_no', '')
    exclude_id = request.args.get('exclude_id', type=int, default=None)
    
    db = get_db()
    cursor = db.cursor()
    
    if exclude_id:
        cursor.execute("SELECT COUNT(*) FROM contracts WHERE contract_no = ? AND id != ?", (contract_no, exclude_id))
    else:
        cursor.execute("SELECT COUNT(*) FROM contracts WHERE contract_no = ?", (contract_no,))
    
    count = cursor.fetchone()[0]
    
    return jsonify({'code': 200, 'message': 'success', 'data': {'exists': count > 0}})


@app.route('/api/contracts/<int:contract_id>', methods=['GET'])
def get_contract(contract_id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM contracts WHERE id = ?", (contract_id,))
    row = cursor.fetchone()
    
    if row:
        return jsonify({'code': 200, 'message': 'success', 'data': dict(row)})
    return jsonify({'code': 404, 'message': '合同不存在', 'data': None})


@app.route('/api/contracts', methods=['POST'])
def create_contract():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    data = request.json
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        contract_no = data.get('contract_no')
        if not contract_no or contract_no.strip() == '':
            cursor.execute("SELECT MAX(id) FROM contracts")
            max_id = cursor.fetchone()[0] or 0
            contract_no = f"HT{datetime.now().strftime('%Y%m%d%H%M%S')}{str(max_id + 1).zfill(3)}"
        
        cursor.execute("SELECT COUNT(*) FROM contracts WHERE contract_no = ?", (contract_no,))
        if cursor.fetchone()[0] > 0:
            contract_no = f"HT{datetime.now().strftime('%Y%m%d%H%M%S')}{str(max_id + 1).zfill(3)}{str(uuid.uuid4().hex[:3])}"
        
        cursor.execute("""
            INSERT INTO contracts
            (b_id, contract_no, party_a, project_order_no, total_amt, paid_amt, sign_date, owner_id, status,
             contract_name, classification, is_audit, pending_acceptance_amount,
             cost, gross_profit, acceptance_date, expected_income_date,
             expected_income_year, business_type, total_cost, acceptance_nodes, payment_nodes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
        """, (
            data.get('b_id'), contract_no, data.get('party_a'), data.get('project_order_no'),
            data.get('total_amt'), 0, data.get('sign_date'), data.get('owner_id'), '执行中',
            data.get('contract_name'), data.get('classification'), data.get('is_audit'), data.get('pending_acceptance_amount'),
            data.get('cost'), data.get('gross_profit'), data.get('acceptance_date'), data.get('expected_income_date'),
            data.get('expected_income_year'), data.get('business_type'), data.get('acceptance_nodes'), data.get('payment_nodes')
        ))
        db.commit()
        contract_id = cursor.lastrowid
        
        return jsonify({'code': 200, 'message': '合同创建成功', 'data': {'id': contract_id, 'contract_no': contract_no}})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/contracts/<int:contract_id>', methods=['PUT'])
def update_contract(contract_id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    data = request.json
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("""
            UPDATE contracts SET
                contract_name=?, contract_no=?, party_a=?, project_order_no=?, total_amt=?, sign_date=?,
                classification=?, is_audit=?, pending_acceptance_amount=?,
                cost=?, gross_profit=?, acceptance_date=?, expected_income_date=?,
                expected_income_year=?, business_type=?, status=?, owner_id=?,
                acceptance_nodes=?, payment_nodes=?
            WHERE id=?
        """, (
            data.get('contract_name'), data.get('contract_no'), data.get('party_a'), data.get('project_order_no'),
            data.get('total_amt'), data.get('sign_date'), data.get('classification'), data.get('is_audit'),
            data.get('pending_acceptance_amount'), data.get('cost'), data.get('gross_profit'),
            data.get('acceptance_date'), data.get('expected_income_date'), data.get('expected_income_year'),
            data.get('business_type'), data.get('status'), data.get('owner_id'),
            data.get('acceptance_nodes'), data.get('payment_nodes'), contract_id
        ))
        db.commit()
        
        return jsonify({'code': 200, 'message': '合同更新成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/contracts/<int:contract_id>/owner', methods=['PATCH', 'POST'])
def update_contract_owner(contract_id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    if payload['role'] != '主任' and payload['role'] != '院长':
        return jsonify({'code': 403, 'message': '无权修改负责人', 'data': None})
    
    data = request.get_json(silent=True) or request.json
    if not data:
        return jsonify({'code': 400, 'message': '请求数据为空', 'data': None})
    
    owner_id = data.get('owner_id')
    
    if not owner_id:
        return jsonify({'code': 400, 'message': '请选择负责人', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("UPDATE contracts SET owner_id = ? WHERE id = ?", (owner_id, contract_id))
        db.commit()
        
        return jsonify({'code': 200, 'message': '负责人修改成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/contracts/import-parse', methods=['POST'])
def import_parse_contracts():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请选择文件', 'data': None})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'code': 400, 'message': '请选择文件', 'data': None})
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'code': 400, 'message': '仅支持Excel文件（.xlsx/.xls）', 'data': None})
    
    try:
        from io import BytesIO
        from openpyxl import load_workbook
        
        wb = load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        col_map = {}
        for idx, header in enumerate(headers):
            header = str(header).strip() if header else ''
            if header == '合同编号':
                col_map['contract_no'] = idx
            elif header == '合同名称':
                col_map['contract_name'] = idx
            elif header == '甲方':
                col_map['party_a'] = idx
            elif header == '项目令号':
                col_map['project_order_no'] = idx
            elif header == '合同总额(万)':
                col_map['total_amt'] = idx
            elif header == '签约日期':
                col_map['sign_date'] = idx
            elif header == '业态':
                col_map['business_type'] = idx
            elif header == '密级':
                col_map['classification'] = idx
            elif header == '负责人':
                col_map['owner_name'] = idx
            elif header == '验收节点':
                col_map['acceptance_nodes'] = idx
            elif header == '回款节点':
                col_map['payment_nodes'] = idx
        
        required_cols = ['contract_no', 'contract_name', 'total_amt']
        for col in required_cols:
            if col not in col_map:
                return jsonify({'code': 400, 'message': f'缺少必要列：{col}', 'data': None})
        
        db = get_db()
        cursor = db.cursor()
        cursor.execute("SELECT contract_no FROM contracts")
        existing_nos = set(row[0] for row in cursor.fetchall())
        
        rows = []
        batch_nos = set()
        valid_count = 0
        
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            errors = []
            
            for key, idx in col_map.items():
                cell = ws.cell(row=row_idx, column=idx + 1)
                value = cell.value
                
                if key == 'total_amt':
                    if value is None:
                        errors.append('合同总额不能为空')
                    else:
                        try:
                            value = float(value) * 10000
                        except:
                            errors.append('合同总额格式错误')
                elif key == 'sign_date':
                    if value:
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d')
                        else:
                            value = str(value)[:10]
                
                row_data[key] = value
            
            contract_no = row_data.get('contract_no')
            if contract_no:
                contract_no = str(contract_no).strip()
                row_data['contract_no'] = contract_no
                
                if contract_no in existing_nos:
                    errors.append('合同编号已存在')
                if contract_no in batch_nos:
                    errors.append('批内合同编号重复')
                batch_nos.add(contract_no)
            
            contract_name = row_data.get('contract_name')
            if not contract_name:
                errors.append('合同名称不能为空')
            
            valid = len(errors) == 0
            if valid:
                valid_count += 1
            
            rows.append({
                'row_index': row_idx,
                'data': row_data,
                'valid': valid,
                'errors': errors
            })
        
        return jsonify({
            'code': 200,
            'message': '解析成功',
            'data': {
                'total': len(rows),
                'valid_count': valid_count,
                'invalid_count': len(rows) - valid_count,
                'rows': rows
            }
        })
    
    except Exception as e:
        return jsonify({'code': 500, 'message': f'解析失败：{str(e)}', 'data': None})


@app.route('/api/contracts/import-execute', methods=['POST'])
def import_execute_contracts():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    data = request.json
    if not data or not isinstance(data, list):
        return jsonify({'code': 400, 'message': '数据格式错误', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    
    success_count = 0
    fail_count = 0
    results = []
    
    cursor.execute("SELECT contract_no FROM contracts")
    existing_nos = set(row[0] for row in cursor.fetchall())
    
    for item in data:
        row_data = item.get('data', {})
        row_index = item.get('row_index', 0)
        
        contract_no = row_data.get('contract_no')
        if not contract_no or contract_no in existing_nos:
            results.append({
                'row_index': row_index,
                'success': False,
                'message': '合同编号已存在或为空'
            })
            fail_count += 1
            continue
        
        try:
            cursor.execute("""
                INSERT INTO contracts
                (contract_no, party_a, project_order_no, total_amt, paid_amt, sign_date, owner_id, status,
                 contract_name, classification, is_audit, pending_acceptance_amount,
                 cost, gross_profit, acceptance_date, expected_income_date,
                 expected_income_year, business_type, total_cost, acceptance_nodes, payment_nodes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
            """, (
                contract_no, 
                row_data.get('party_a'), 
                row_data.get('project_order_no'),
                row_data.get('total_amt', 0), 
                0, 
                row_data.get('sign_date'), 
                payload['username'], 
                '执行中',
                row_data.get('contract_name'), 
                row_data.get('classification'), 
                0, 
                0,
                0, 
                0, 
                '', 
                '',
                '', 
                row_data.get('business_type'), 
                row_data.get('acceptance_nodes'), 
                row_data.get('payment_nodes')
            ))
            
            existing_nos.add(contract_no)
            success_count += 1
            results.append({
                'row_index': row_index,
                'success': True,
                'message': '导入成功'
            })
        except Exception as e:
            fail_count += 1
            results.append({
                'row_index': row_index,
                'success': False,
                'message': str(e)
            })
    
    db.commit()
    
    return jsonify({
        'code': 200,
        'message': '导入完成',
        'data': {
            'total': len(data),
            'success_count': success_count,
            'fail_count': fail_count,
            'results': results
        }
    })


@app.route('/api/contracts/<int:contract_id>', methods=['DELETE'])
def delete_contract(contract_id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("DELETE FROM contracts WHERE id=?", (contract_id,))
        db.commit()
        
        return jsonify({'code': 200, 'message': '合同删除成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/customers', methods=['GET'])
def get_customers():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    username = payload['username']
    role = payload['role']
    keyword = request.args.get('keyword', '')
    
    db = get_db()
    cursor = db.cursor()
    
    if role == '主任' or role == '院长':
        if keyword:
            cursor.execute("""
                SELECT c.*, u.name as owner_name 
                FROM customers c 
                LEFT JOIN users u ON c.owner_id = u.username 
                WHERE c.company LIKE ? OR c.name LIKE ? OR c.contact_name LIKE ?
                ORDER BY c.created_at DESC
            """, (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%'))
        else:
            cursor.execute("""
                SELECT c.*, u.name as owner_name 
                FROM customers c 
                LEFT JOIN users u ON c.owner_id = u.username 
                ORDER BY c.created_at DESC
            """)
    else:
        if keyword:
            cursor.execute("""
                SELECT c.*, u.name as owner_name 
                FROM customers c 
                LEFT JOIN users u ON c.owner_id = u.username 
                WHERE c.owner_id = ? AND (c.company LIKE ? OR c.name LIKE ? OR c.contact_name LIKE ?)
                ORDER BY c.created_at DESC
            """, (username, f'%{keyword}%', f'%{keyword}%', f'%{keyword}%'))
        else:
            cursor.execute("""
                SELECT c.*, u.name as owner_name 
                FROM customers c 
                LEFT JOIN users u ON c.owner_id = u.username 
                WHERE c.owner_id = ? 
                ORDER BY c.created_at DESC
            """, (username,))
    
    rows = cursor.fetchall()
    customers = []
    for row in rows:
        customers.append(dict(row))
    
    return jsonify({'code': 200, 'message': 'success', 'data': customers})


@app.route('/api/customers', methods=['POST'])
def create_customer():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    data = request.json
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("""
            INSERT INTO customers (name, company, phone, level, source, owner_id, contact_name, email, industry, region, created_at, last_follow)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        """, (
            data.get('name'), data.get('company'), data.get('phone'),
            data.get('level'), data.get('source'), data.get('owner_id'),
            data.get('contact_name'), data.get('email'), data.get('industry'), data.get('region')
        ))
        db.commit()
        
        return jsonify({'code': 200, 'message': '客户创建成功', 'data': {'id': cursor.lastrowid}})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/customers/<int:cust_id>', methods=['DELETE'])
def delete_customer(cust_id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("DELETE FROM customers WHERE id=?", (cust_id,))
        db.commit()
        
        return jsonify({'code': 200, 'message': '客户删除成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/customers/<int:cust_id>', methods=['PUT'])
def update_customer(cust_id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    data = request.json
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("""
            UPDATE customers SET
                name=?, company=?, phone=?, level=?, source=?,
                contact_name=?, email=?, industry=?, region=?
            WHERE id=?
        """, (
            data.get('name'), data.get('company'), data.get('phone'),
            data.get('level'), data.get('source'),
            data.get('contact_name'), data.get('email'),
            data.get('industry'), data.get('region'), cust_id
        ))
        db.commit()
        
        return jsonify({'code': 200, 'message': '客户更新成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/business', methods=['GET'])
def get_business():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    username = payload['username']
    role = payload['role']
    
    db = get_db()
    cursor = db.cursor()
    
    if role == '主任' or role == '院长':
        cursor.execute("""
            SELECT b.*, c.company as customer_name, c.name as customer_contact, u.name as owner_name 
            FROM business b 
            LEFT JOIN customers c ON b.cust_id = c.id 
            LEFT JOIN users u ON b.owner_id = u.username 
            WHERE b.status = 'active' 
            ORDER BY b.created_at DESC
        """)
    else:
        cursor.execute("""
            SELECT b.*, c.company as customer_name, c.name as customer_contact, u.name as owner_name 
            FROM business b 
            LEFT JOIN customers c ON b.cust_id = c.id 
            LEFT JOIN users u ON b.owner_id = u.username 
            WHERE b.owner_id = ? AND b.status = 'active' 
            ORDER BY b.created_at DESC
        """, (username,))
    
    rows = cursor.fetchall()
    business = []
    for row in rows:
        business.append(dict(row))
    
    return jsonify({'code': 200, 'message': 'success', 'data': business})


@app.route('/api/business', methods=['POST'])
def create_business():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    data = request.json
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("""
            INSERT INTO business (title, cust_id, stakeholder, amount, stage, predict_date, source, industry, region, owner_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data.get('title'), data.get('cust_id'), data.get('stakeholder'), data.get('amount'), data.get('stage'),
            data.get('predict_date'), data.get('source'), data.get('industry'), data.get('region'), data.get('owner_id')
        ))
        db.commit()
        
        return jsonify({'code': 200, 'message': '商机创建成功', 'data': {'id': cursor.lastrowid}})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/business/<int:business_id>', methods=['DELETE'])
def delete_business(business_id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("DELETE FROM business WHERE id=?", (business_id,))
        db.commit()
        
        return jsonify({'code': 200, 'message': '商机删除成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/business/<int:business_id>', methods=['PUT'])
def update_business(business_id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    data = request.json
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("""
            UPDATE business SET
                title=?, cust_id=?, stakeholder=?, amount=?, stage=?, predict_date=?,
                source=?, industry=?, region=?
            WHERE id=?
        """, (
            data.get('title'), data.get('cust_id'), data.get('stakeholder'), 
            data.get('amount'), data.get('stage'), data.get('predict_date'),
            data.get('source'), data.get('industry'), data.get('region'), business_id
        ))
        db.commit()
        
        return jsonify({'code': 200, 'message': '商机更新成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/dashboard', methods=['GET'])
def get_dashboard():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    username = payload['username']
    role = payload['role']
    time_range = request.args.get('time_range', 'all')
    
    db = get_db()
    cursor = db.cursor()
    
    result = {}
    
    now = datetime.now()
    if time_range == 'month':
        date_filter = f"{now.year}-{str(now.month).zfill(2)}-01"
        date_condition = "WHERE strftime('%Y-%m', created_at) = strftime('%Y-%m', 'now')"
        contract_date_condition = "WHERE strftime('%Y-%m', sign_date) = strftime('%Y-%m', 'now')"
        payment_date_condition = "WHERE strftime('%Y-%m', payment_date) = strftime('%Y-%m', 'now')"
    elif time_range == 'quarter':
        quarter = (now.month - 1) // 3 + 1
        start_month = (quarter - 1) * 3 + 1
        date_condition = f"WHERE created_at >= '{now.year}-{str(start_month).zfill(2)}-01'"
        contract_date_condition = f"WHERE sign_date >= '{now.year}-{str(start_month).zfill(2)}-01'"
        payment_date_condition = f"WHERE payment_date >= '{now.year}-{str(start_month).zfill(2)}-01'"
    elif time_range == 'year':
        date_condition = f"WHERE strftime('%Y', created_at) = '{now.year}'"
        contract_date_condition = f"WHERE strftime('%Y', sign_date) = '{now.year}'"
        payment_date_condition = f"WHERE strftime('%Y', payment_date) = '{now.year}'"
    else:
        date_condition = ""
        contract_date_condition = ""
        payment_date_condition = ""
    
    if role == '主任' or role == '院长':
        cursor.execute(f"SELECT COUNT(*) as total FROM customers {date_condition}")
        result['total_customers'] = cursor.fetchone()['total']
        
        cursor.execute(f"SELECT COUNT(*) as total FROM business {date_condition} AND status = 'active'" if date_condition else "SELECT COUNT(*) as total FROM business WHERE status = 'active'")
        result['total_business'] = cursor.fetchone()['total']
        
        cursor.execute(f"SELECT COUNT(*) as total FROM contracts {contract_date_condition}")
        result['total_contracts'] = cursor.fetchone()['total']
        
        cursor.execute(f"SELECT SUM(total_amt) as total FROM contracts {contract_date_condition}")
        total = cursor.fetchone()['total'] or 0
        result['contracts_amount'] = total
        
        cursor.execute(f"SELECT SUM(amount) as total FROM payment_records {payment_date_condition}")
        total = cursor.fetchone()['total'] or 0
        result['total_payments'] = total
    else:
        cursor.execute(f"SELECT COUNT(*) as total FROM customers WHERE owner_id = ? {('AND' if date_condition else '')} {date_condition.replace('WHERE', '')}", (username,))
        result['total_customers'] = cursor.fetchone()['total']
        
        if date_condition:
            cursor.execute(f"SELECT COUNT(*) as total FROM business WHERE owner_id = ? AND status = 'active' AND {date_condition.replace('WHERE', '')}", (username,))
        else:
            cursor.execute("SELECT COUNT(*) as total FROM business WHERE owner_id = ? AND status = 'active'", (username,))
        result['total_business'] = cursor.fetchone()['total']
        
        if contract_date_condition:
            cursor.execute(f"SELECT COUNT(*) as total FROM contracts WHERE owner_id = ? AND {contract_date_condition.replace('WHERE', '')}", (username,))
        else:
            cursor.execute("SELECT COUNT(*) as total FROM contracts WHERE owner_id = ?", (username,))
        result['total_contracts'] = cursor.fetchone()['total']
        
        if contract_date_condition:
            cursor.execute(f"SELECT SUM(total_amt) as total FROM contracts WHERE owner_id = ? AND {contract_date_condition.replace('WHERE', '')}", (username,))
        else:
            cursor.execute("SELECT SUM(total_amt) as total FROM contracts WHERE owner_id = ?", (username,))
        total = cursor.fetchone()['total'] or 0
        result['contracts_amount'] = total
        
        if payment_date_condition:
            cursor.execute(f"""
                SELECT SUM(pr.amount) as total 
                FROM payment_records pr 
                JOIN contracts c ON pr.contract_id = c.id 
                WHERE c.owner_id = ? AND {payment_date_condition.replace('WHERE', '')}
            """, (username,))
        else:
            cursor.execute("""
                SELECT SUM(pr.amount) as total 
                FROM payment_records pr 
                JOIN contracts c ON pr.contract_id = c.id 
                WHERE c.owner_id = ?
            """, (username,))
        total = cursor.fetchone()['total'] or 0
        result['total_payments'] = total
    
    if time_range == 'month':
        cursor.execute("""
            SELECT strftime('%d', created_at) as day, COUNT(*) as count 
            FROM customers 
            WHERE strftime('%Y-%m', created_at) = strftime('%Y-%m', 'now')
            GROUP BY strftime('%d', created_at) 
            ORDER BY day
        """)
        customer_monthly = {row['day']: row['count'] for row in cursor.fetchall()}
        
        cursor.execute("""
            SELECT strftime('%d', created_at) as day, COUNT(*) as count 
            FROM business 
            WHERE status = 'active' AND strftime('%Y-%m', created_at) = strftime('%Y-%m', 'now')
            GROUP BY strftime('%d', created_at) 
            ORDER BY day
        """)
        business_monthly = {row['day']: row['count'] for row in cursor.fetchall()}
        
        cursor.execute("""
            SELECT strftime('%d', sign_date) as day, COUNT(*) as count 
            FROM contracts 
            WHERE strftime('%Y-%m', sign_date) = strftime('%Y-%m', 'now')
            GROUP BY strftime('%d', sign_date) 
            ORDER BY day
        """)
        contract_monthly = {row['day']: row['count'] for row in cursor.fetchall()}
        
        days_in_month = (now.replace(month=now.month % 12 + 1, day=1) - timedelta(days=1)).day
        months = [f"{i}日" for i in range(1, days_in_month + 1)]
        customer_data = [customer_monthly.get(str(i), 0) for i in range(1, days_in_month + 1)]
        business_data = [business_monthly.get(str(i), 0) for i in range(1, days_in_month + 1)]
        contract_data = [contract_monthly.get(str(i), 0) for i in range(1, days_in_month + 1)]
    elif time_range == 'quarter':
        quarter = (now.month - 1) // 3 + 1
        start_month = (quarter - 1) * 3 + 1
        cursor.execute(f"""
            SELECT strftime('%Y-%m', created_at) as month, COUNT(*) as count 
            FROM customers 
            WHERE created_at >= '{now.year}-{str(start_month).zfill(2)}-01'
            GROUP BY strftime('%Y-%m', created_at) 
            ORDER BY month
        """)
        customer_monthly = {row['month']: row['count'] for row in cursor.fetchall()}
        
        cursor.execute(f"""
            SELECT strftime('%Y-%m', created_at) as month, COUNT(*) as count 
            FROM business 
            WHERE status = 'active' AND created_at >= '{now.year}-{str(start_month).zfill(2)}-01'
            GROUP BY strftime('%Y-%m', created_at) 
            ORDER BY month
        """)
        business_monthly = {row['month']: row['count'] for row in cursor.fetchall()}
        
        cursor.execute(f"""
            SELECT strftime('%Y-%m', sign_date) as month, COUNT(*) as count 
            FROM contracts 
            WHERE sign_date >= '{now.year}-{str(start_month).zfill(2)}-01'
            GROUP BY strftime('%Y-%m', sign_date) 
            ORDER BY month
        """)
        contract_monthly = {row['month']: row['count'] for row in cursor.fetchall()}
        
        months = []
        customer_data = []
        business_data = []
        contract_data = []
        for m in range(start_month, start_month + 3):
            month_str = f"{now.year}-{str(m).zfill(2)}"
            months.append(f"{m}月")
            customer_data.append(customer_monthly.get(month_str, 0))
            business_data.append(business_monthly.get(month_str, 0))
            contract_data.append(contract_monthly.get(month_str, 0))
    else:
        cursor.execute("""
            SELECT strftime('%Y-%m', created_at) as month, COUNT(*) as count 
            FROM customers 
            WHERE created_at >= DATE('now', '-12 months') 
            GROUP BY strftime('%Y-%m', created_at) 
            ORDER BY month
        """)
        customer_monthly = {row['month']: row['count'] for row in cursor.fetchall()}
        
        cursor.execute("""
            SELECT strftime('%Y-%m', created_at) as month, COUNT(*) as count 
            FROM business 
            WHERE status = 'active' AND created_at >= DATE('now', '-12 months') 
            GROUP BY strftime('%Y-%m', created_at) 
            ORDER BY month
        """)
        business_monthly = {row['month']: row['count'] for row in cursor.fetchall()}
        
        cursor.execute("""
            SELECT strftime('%Y-%m', sign_date) as month, COUNT(*) as count 
            FROM contracts 
            WHERE sign_date >= DATE('now', '-12 months') 
            GROUP BY strftime('%Y-%m', sign_date) 
            ORDER BY month
        """)
        contract_monthly = {row['month']: row['count'] for row in cursor.fetchall()}
        
        months = []
        customer_data = []
        business_data = []
        contract_data = []
        
        for i in range(12):
            date = now - timedelta(days=i*30)
            month_str = date.strftime('%Y-%m')
            months.insert(0, date.strftime('%m月'))
            customer_data.insert(0, customer_monthly.get(month_str, 0))
            business_data.insert(0, business_monthly.get(month_str, 0))
            contract_data.insert(0, contract_monthly.get(month_str, 0))
    
    result['chart_data'] = {
        'months': months,
        'customer_data': customer_data,
        'business_data': business_data,
        'contract_data': contract_data
    }
    
    cursor.execute("""
        SELECT u.name, u.role, COALESCE(SUM(c.total_amt), 0) as total_amount
        FROM users u
        LEFT JOIN contracts c ON u.username = c.owner_id
        GROUP BY u.username, u.name, u.role
        ORDER BY total_amount DESC
        LIMIT 5
    """)
    sales_ranking = []
    for row in cursor.fetchall():
        sales_ranking.append({
            'name': row['name'],
            'role': row['role'],
            'amount': row['total_amount']
        })
    result['sales_ranking'] = sales_ranking
    
    return jsonify({'code': 200, 'message': 'success', 'data': result})


@app.route('/api/users', methods=['GET'])
def get_users():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT username, name, role FROM users ORDER BY name")
    
    rows = cursor.fetchall()
    users = []
    for row in rows:
        users.append(dict(row))
    
    return jsonify({'code': 200, 'message': 'success', 'data': users})


@app.route('/api/users', methods=['POST'])
def create_user():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    if payload['role'] != '主任' and payload['role'] != '院长':
        return jsonify({'code': 403, 'message': '无权访问', 'data': None})
    
    data = request.json
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("SELECT username FROM users WHERE username = ?", (data.get('username'),))
        if cursor.fetchone():
            return jsonify({'code': 400, 'message': '用户名已存在', 'data': None})
        
        hashed_pwd = hash_password(data.get('password'))
        cursor.execute("""
            INSERT INTO users (username, password_hash, name, role)
            VALUES (?, ?, ?, ?)
        """, (data.get('username'), hashed_pwd, data.get('name'), data.get('role')))
        db.commit()
        
        return jsonify({'code': 200, 'message': '用户创建成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/users/<username>', methods=['DELETE'])
def delete_user(username):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    if payload['role'] != '主任' and payload['role'] != '院长':
        return jsonify({'code': 403, 'message': '无权访问', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("DELETE FROM users WHERE username = ?", (username,))
        db.commit()
        
        return jsonify({'code': 200, 'message': '用户删除成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/payment_records', methods=['GET'])
def get_payment_records():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    contract_id = request.args.get('contract_id')
    username = payload['username']
    role = payload['role']
    
    db = get_db()
    cursor = db.cursor()
    
    if contract_id:
        cursor.execute("""
            SELECT pr.*, c.contract_name, c.contract_no, c.party_a, u.name as owner_name 
            FROM payment_records pr 
            LEFT JOIN contracts c ON pr.contract_id = c.id 
            LEFT JOIN users u ON c.owner_id = u.username 
            WHERE pr.contract_id = ? 
            ORDER BY pr.payment_date DESC
        """, (contract_id,))
    else:
        if role == '主任' or role == '院长':
            cursor.execute("""
                SELECT pr.*, c.contract_name, c.contract_no, c.party_a, u.name as owner_name 
                FROM payment_records pr 
                LEFT JOIN contracts c ON pr.contract_id = c.id 
                LEFT JOIN users u ON c.owner_id = u.username 
                ORDER BY pr.payment_date DESC
            """)
        else:
            cursor.execute("""
                SELECT pr.*, c.contract_name, c.contract_no, c.party_a, u.name as owner_name 
                FROM payment_records pr 
                LEFT JOIN contracts c ON pr.contract_id = c.id 
                LEFT JOIN users u ON c.owner_id = u.username 
                WHERE c.owner_id = ? 
                ORDER BY pr.payment_date DESC
            """, (username,))
    
    rows = cursor.fetchall()
    records = []
    for row in rows:
        records.append(dict(row))
    
    return jsonify({'code': 200, 'message': 'success', 'data': records})


@app.route('/api/payment_records', methods=['POST'])
def create_payment_record():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    data = request.json
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("""
            INSERT INTO payment_records (contract_id, payment_date, amount, note)
            VALUES (?, ?, ?, ?)
        """, (data.get('contract_id'), data.get('payment_date'), data.get('amount'), data.get('note')))
        db.commit()
        
        cursor.execute("SELECT SUM(amount) as total FROM payment_records WHERE contract_id = ?", (data.get('contract_id'),))
        total_paid = cursor.fetchone()['total'] or 0
        cursor.execute("UPDATE contracts SET paid_amt = ? WHERE id = ?", (total_paid, data.get('contract_id')))
        db.commit()
        
        return jsonify({'code': 200, 'message': '回款记录创建成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/payment_records/<int:record_id>', methods=['PUT'])
def update_payment_record(record_id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    data = request.json
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("SELECT contract_id FROM payment_records WHERE id = ?", (record_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'code': 404, 'message': '回款记录不存在', 'data': None})
        
        original_contract_id = row['contract_id']
        
        cursor.execute("""
            UPDATE payment_records 
            SET payment_date = ?, amount = ?, note = ?
            WHERE id = ?
        """, (data.get('payment_date'), data.get('amount'), data.get('note'), record_id))
        db.commit()
        
        cursor.execute("SELECT SUM(amount) as total FROM payment_records WHERE contract_id = ?", (original_contract_id,))
        total_paid = cursor.fetchone()['total'] or 0
        cursor.execute("UPDATE contracts SET paid_amt = ? WHERE id = ?", (total_paid, original_contract_id))
        db.commit()
        
        return jsonify({'code': 200, 'message': '回款记录更新成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/payment_records/<int:record_id>', methods=['DELETE'])
def delete_payment_record(record_id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("SELECT contract_id FROM payment_records WHERE id = ?", (record_id,))
        row = cursor.fetchone()
        if row:
            contract_id = row['contract_id']
            
            cursor.execute("DELETE FROM payment_records WHERE id = ?", (record_id,))
            
            cursor.execute("SELECT SUM(amount) as total FROM payment_records WHERE contract_id = ?", (contract_id,))
            total_paid = cursor.fetchone()['total'] or 0
            cursor.execute("UPDATE contracts SET paid_amt = ? WHERE id = ?", (total_paid, contract_id))
        
        db.commit()
        
        return jsonify({'code': 200, 'message': '回款记录删除成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


def detect_file_format(file_bytes):
    if len(file_bytes) < 8:
        return 'unknown'
    if file_bytes[:4] == b'\x50\x4B\x03\x04':
        return 'xlsx'
    if file_bytes[:8] == b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1':
        return 'xls'
    if file_bytes[:3] == b'\xEF\xBB\xBF':
        return 'csv'
    if b'<!DOCTYPE' in file_bytes[:200] or b'<html' in file_bytes[:200]:
        return 'html'
    try:
        import zipfile
        try:
            with zipfile.ZipFile(BytesIO(file_bytes)) as zf:
                if '[Content_Types].xml' in zf.namelist():
                    return 'xlsx'
        except:
            pass
    except:
        pass
    return 'unknown'

@app.route('/api/payments/import-parse', methods=['POST'])
def import_parse_payments():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请选择文件', 'data': None})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'code': 400, 'message': '请选择文件', 'data': None})
    
    filename = file.filename.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
        return jsonify({'code': 400, 'message': '仅支持Excel文件（.xlsx/.xls）', 'data': None})
    
    try:
        from io import BytesIO
        
        file_bytes = file.read()
        file_format = detect_file_format(file_bytes)
        
        if file_format == 'xls':
            try:
                import xlrd
                wb = xlrd.open_workbook(file_contents=file_bytes)
                ws = wb.sheet_by_index(0)
                headers = [ws.cell_value(0, i) for i in range(ws.ncols)]
                use_xlrd = True
            except ImportError:
                return jsonify({'code': 400, 'message': '请安装xlrd库以支持.xls文件格式，或使用.xlsx格式', 'data': None})
            except Exception as e:
                return jsonify({'code': 400, 'message': f'无法读取.xls文件：{str(e)}', 'data': None})
        elif file_format == 'xlsx':
            try:
                from openpyxl import load_workbook
                wb = load_workbook(BytesIO(file_bytes), data_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                use_xlrd = False
            except Exception as e:
                return jsonify({'code': 400, 'message': f'无法读取.xlsx文件：{str(e)}。请确认文件未被加密或损坏，或尝试另存为.xlsx格式', 'data': None})
        else:
            return jsonify({'code': 400, 'message': f'文件格式不匹配：扩展名是{filename.split(".")[-1]}，但实际文件格式无法识别。请确认文件是有效的Excel文件', 'data': None})
        
        def normalize_header(header):
            if header is None:
                return ''
            s = str(header)
            s = s.replace('\u3000', '')
            s = s.replace('　', '')
            s = ''.join(s.split())
            return s
        
        col_map = {}
        actual_headers = []
        
        for idx, header in enumerate(headers):
            normalized = normalize_header(header)
            actual_headers.append(str(header) if header else '')
            
            if '合同编号' in normalized or '合同号' in normalized:
                col_map['contract_no'] = idx
            elif '合同名称' in normalized or '合同名' in normalized:
                col_map['contract_name'] = idx
            elif '回款日期' in normalized or '日期' in normalized:
                col_map['payment_date'] = idx
            elif '金额' in normalized or '款额' in normalized:
                col_map['amount'] = idx
            elif '备注' in normalized or '说明' in normalized:
                col_map['note'] = idx
        
        required_cols = ['contract_no', 'payment_date', 'amount']
        missing_cols = []
        for col in required_cols:
            if col not in col_map:
                missing_cols.append(col)
        
        if missing_cols:
            col_names = {'contract_no': '合同编号', 'payment_date': '回款日期', 'amount': '金额'}
            missing_names = ', '.join([col_names.get(c, c) for c in missing_cols])
            return jsonify({
                'code': 400, 
                'message': f'缺少必要列：{missing_names}。文件中的表头为：{", ".join(actual_headers)}', 
                'data': None
            })
        
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute("SELECT id, contract_no, contract_name FROM contracts")
        contract_map = {}
        for row in cursor.fetchall():
            contract_map[row['contract_no'].strip()] = {'id': row['id'], 'name': row['contract_name']}
        
        cursor.execute("""
            SELECT pr.id, pr.contract_id, pr.payment_date, pr.amount, pr.note, pr.created_at, c.contract_no, c.contract_name
            FROM payment_records pr
            LEFT JOIN contracts c ON pr.contract_id = c.id
        """)
        existing_payments = {}
        for row in cursor.fetchall():
            key = f"{row['contract_no']}_{row['payment_date']}_{row['amount']}"
            existing_payments[key] = {
                'id': row['id'], 
                'contract_id': row['contract_id'],
                'contract_name': row['contract_name'],
                'note': row['note'],
                'created_at': row['created_at']
            }
        
        rows = []
        valid_count = 0
        
        max_row = ws.nrows if use_xlrd else ws.max_row
        
        for row_idx in range(2, max_row + 1):
            row_data = {}
            errors = []
            is_duplicate = False
            
            for key, idx in col_map.items():
                if use_xlrd:
                    value = ws.cell_value(row_idx - 1, idx)
                else:
                    value = ws.cell(row=row_idx, column=idx + 1).value
                
                if key == 'amount':
                    if value is None or value == '':
                        errors.append('金额不能为空')
                    else:
                        try:
                            value = float(value) * 10000
                        except:
                            errors.append('金额格式错误')
                elif key == 'payment_date':
                    if value:
                        if use_xlrd:
                            if isinstance(value, float):
                                try:
                                    date_tuple = xlrd.xldate_as_tuple(value, wb.datemode)
                                    value = f"{date_tuple[0]}-{str(date_tuple[1]).zfill(2)}-{str(date_tuple[2]).zfill(2)}"
                                except:
                                    value = str(value)[:10]
                            else:
                                value = str(value)[:10]
                        elif isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d')
                        else:
                            value = str(value)[:10]
                    else:
                        errors.append('回款日期不能为空')
                elif key == 'contract_no':
                    if value:
                        value = str(value).strip()
                    else:
                        errors.append('合同编号不能为空')
                
                row_data[key] = value
            
            contract_no = row_data.get('contract_no', '').strip()
            if contract_no:
                if contract_no not in contract_map:
                    errors.append('合同编号不存在于系统中')
                else:
                    row_data['contract_id'] = contract_map[contract_no]['id']
            
            payment_date = row_data.get('payment_date', '')
            amount = row_data.get('amount', 0)
            
            if contract_no and payment_date and amount:
                check_key = f"{contract_no}_{payment_date}_{amount}"
                if check_key in existing_payments:
                    is_duplicate = True
                    existing_info = existing_payments[check_key]
                    row_data['duplicate_id'] = existing_info['id']
                    row_data['existing_data'] = {
                        'id': existing_info['id'],
                        'contract_name': existing_info['contract_name'],
                        'payment_date': payment_date,
                        'amount': amount,
                        'note': existing_info['note'],
                        'created_at': existing_info['created_at']
                    }
            
            valid = len(errors) == 0
            if valid:
                valid_count += 1
            
            rows.append({
                'row_index': row_idx,
                'data': row_data,
                'valid': valid,
                'errors': errors,
                'is_duplicate': is_duplicate
            })
        
        return jsonify({
            'code': 200,
            'message': '解析成功',
            'data': {
                'total': len(rows),
                'valid_count': valid_count,
                'invalid_count': len(rows) - valid_count,
                'duplicate_count': sum(1 for r in rows if r['is_duplicate']),
                'rows': rows
            }
        })
    
    except Exception as e:
        return jsonify({'code': 500, 'message': f'解析失败：{str(e)}', 'data': None})


@app.route('/api/payments/import-execute', methods=['POST'])
def import_execute_payments():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    data = request.json
    if not data or not isinstance(data, list):
        return jsonify({'code': 400, 'message': '数据格式错误', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    
    success_count = 0
    fail_count = 0
    results = []
    
    for item in data:
        row_data = item.get('data', {})
        row_index = item.get('row_index', 0)
        duplicate_action = item.get('duplicate_action', 'keep_import')
        
        contract_id = row_data.get('contract_id')
        payment_date = row_data.get('payment_date')
        amount = row_data.get('amount')
        note = row_data.get('note', '')
        duplicate_id = row_data.get('duplicate_id')
        
        if not contract_id or not payment_date or not amount:
            results.append({
                'row_index': row_index,
                'success': False,
                'message': '数据不完整'
            })
            fail_count += 1
            continue
        
        try:
            if duplicate_id and duplicate_action == 'keep_existing':
                results.append({
                    'row_index': row_index,
                    'success': True,
                    'message': '保留系统数据，跳过导入'
                })
                success_count += 1
                continue
            
            if duplicate_id and duplicate_action == 'replace':
                cursor.execute("DELETE FROM payment_records WHERE id = ?", (duplicate_id,))
            
            cursor.execute("""
                INSERT INTO payment_records (contract_id, payment_date, amount, note)
                VALUES (?, ?, ?, ?)
            """, (contract_id, payment_date, amount, note))
            
            cursor.execute("SELECT SUM(amount) as total FROM payment_records WHERE contract_id = ?", (contract_id,))
            total_paid = cursor.fetchone()['total'] or 0
            cursor.execute("UPDATE contracts SET paid_amt = ? WHERE id = ?", (total_paid, contract_id))
            
            success_count += 1
            results.append({
                'row_index': row_index,
                'success': True,
                'message': '导入成功' if not duplicate_id else '替换成功'
            })
        except Exception as e:
            fail_count += 1
            results.append({
                'row_index': row_index,
                'success': False,
                'message': str(e)
            })
    
    db.commit()
    
    return jsonify({
        'code': 200,
        'message': '导入完成',
        'data': {
            'total': len(data),
            'success_count': success_count,
            'fail_count': fail_count,
            'results': results
        }
    })


@app.route('/api/business_stage_logs', methods=['GET'])
def get_stage_logs():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    business_id = request.args.get('business_id')
    
    db = get_db()
    cursor = db.cursor()
    
    if business_id:
        cursor.execute("SELECT * FROM business_stage_logs WHERE business_id = ? ORDER BY changed_at DESC", (business_id,))
    else:
        cursor.execute("SELECT * FROM business_stage_logs ORDER BY changed_at DESC")
    
    rows = cursor.fetchall()
    logs = []
    for row in rows:
        logs.append(dict(row))
    
    return jsonify({'code': 200, 'message': 'success', 'data': logs})


@app.route('/api/follow_logs', methods=['GET'])
def get_follow_logs():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    ref_type = request.args.get('ref_type')
    ref_id = request.args.get('ref_id')
    keyword = request.args.get('keyword')
    
    db = get_db()
    cursor = db.cursor()
    
    conditions = []
    params = []
    
    if ref_type:
        conditions.append("fl.ref_type = ?")
        params.append(ref_type)
    
    if ref_id:
        conditions.append("fl.ref_id = ?")
        params.append(ref_id)
    
    if keyword:
        conditions.append("(fl.content LIKE ? OR fl.subject LIKE ? OR fl.participants LIKE ? OR u.name LIKE ?)")
        keyword_pattern = f'%{keyword}%'
        params.extend([keyword_pattern, keyword_pattern, keyword_pattern, keyword_pattern])
    
    where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""
    
    query = f"""
        SELECT fl.*, u.name as user_name 
        FROM follow_logs fl 
        LEFT JOIN users u ON fl.user_id = u.username 
        {where_clause}
        ORDER BY fl.created_at DESC
    """
    
    cursor.execute(query, params)
    
    rows = cursor.fetchall()
    logs = []
    for row in rows:
        logs.append(dict(row))
    
    return jsonify({'code': 200, 'message': 'success', 'data': logs})


@app.route('/api/follow_logs', methods=['POST'])
def create_follow_log():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    data = request.json
    username = payload['username']
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("""
            INSERT INTO follow_logs 
            (ref_type, ref_id, user_id, content, log_time, subject, participants, location, next_plan)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data.get('ref_type'), data.get('ref_id'), username,
            data.get('content'), data.get('log_time'), data.get('subject'),
            data.get('participants'), data.get('location'), data.get('next_plan')
        ))
        db.commit()
        
        return jsonify({'code': 200, 'message': '跟进记录添加成功', 'data': {'id': cursor.lastrowid}})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/follow_logs/<int:log_id>', methods=['DELETE'])
def delete_follow_log(log_id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("DELETE FROM follow_logs WHERE id = ?", (log_id,))
        db.commit()
        
        return jsonify({'code': 200, 'message': '跟进记录删除成功', 'data': None})
    except Exception as e:
        db.rollback()
        return jsonify({'code': 500, 'message': str(e), 'data': None})


@app.route('/api/pool', methods=['GET'])
def get_pool():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute("SELECT * FROM customers WHERE owner_id IS NULL OR owner_id = '' ORDER BY created_at DESC")
    rows = cursor.fetchall()
    pool_data = []
    for row in rows:
        item = dict(row)
        item['quality_score'] = 65 + (item.get('id', 0) % 35)
        pool_data.append(item)
    
    return jsonify({'code': 200, 'message': 'success', 'data': pool_data})


@app.route('/api/workhours', methods=['GET'])
def get_workhours():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute("""
        SELECT wh.*, b.title as project_name, u.name as user_name
        FROM work_hours wh
        LEFT JOIN business b ON wh.business_id = b.id
        LEFT JOIN users u ON wh.user_id = u.username
        ORDER BY wh.work_date DESC
    """)
    rows = cursor.fetchall()
    workhours = []
    for row in rows:
        item = dict(row)
        item['date'] = item.pop('work_date')
        item['task_name'] = item.get('description', '')[:20]
        workhours.append(item)
    
    return jsonify({'code': 200, 'message': 'success', 'data': workhours})


@app.route('/api/projects', methods=['GET'])
def get_projects():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute("""
        SELECT b.id, b.title as project_name, c.name as customer_name,
               b.created_at as start_date, b.predict_date as end_date,
               b.stage, b.status, b.project_manager as manager, 
               b.owner_id, u.name as owner_name,
               b.amount, b.probability
        FROM business b
        LEFT JOIN customers c ON b.cust_id = c.id
        LEFT JOIN users u ON b.owner_id = u.username
        WHERE b.status = 'active'
        ORDER BY b.created_at DESC
    """)
    rows = cursor.fetchall()
    projects = []
    for row in rows:
        item = dict(row)
        stage_progress = {
            '需求确认': 20, '方案报价': 40, '商务谈判': 60, 
            '合同签署': 80, '项目启动': 90, '实施中': 95, '已完成': 100
        }
        item['progress'] = stage_progress.get(item.get('stage'), 30)
        status_map = {'active': '进行中'}
        item['status'] = status_map.get(item.get('status'), item.get('status', '进行中'))
        
        cursor.execute("""
            SELECT u.name FROM project_assignments pa
            LEFT JOIN users u ON pa.user_id = u.username
            WHERE pa.project_type = 'business' AND pa.project_id = ?
        """, (item['id'],))
        members = cursor.fetchall()
        item['team_members'] = [m[0] for m in members]
        if not item['team_members'] and item['manager']:
            item['team_members'] = [item['manager']]
        
        projects.append(item)
    
    return jsonify({'code': 200, 'message': 'success', 'data': projects})


@app.route('/api/search', methods=['GET'])
def search():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    keyword = request.args.get('keyword', '')
    
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute("SELECT id, name, company, phone, level, source FROM customers WHERE name LIKE ? OR company LIKE ?", (f'%{keyword}%', f'%{keyword}%'))
    customers = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute("SELECT id, title, amount, stage, probability, owner_id FROM business WHERE title LIKE ? OR stage LIKE ?", (f'%{keyword}%', f'%{keyword}%'))
    business = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute("SELECT id, contract_name, contract_no, total_amt, sign_date FROM contracts WHERE contract_name LIKE ? OR contract_no LIKE ?", (f'%{keyword}%', f'%{keyword}%'))
    contracts = [dict(row) for row in cursor.fetchall()]
    
    cursor.execute("""
        SELECT fl.id, fl.ref_type, fl.ref_id, fl.content, fl.subject, fl.created_at, u.name as user_name
        FROM follow_logs fl
        LEFT JOIN users u ON fl.user_id = u.username
        WHERE fl.content LIKE ? OR fl.subject LIKE ? OR fl.participants LIKE ?
    """, (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%'))
    follow_logs = [dict(row) for row in cursor.fetchall()]
    
    return jsonify({'code': 200, 'message': 'success', 'data': {
        'customers': customers,
        'business': business,
        'contracts': contracts,
        'follow_logs': follow_logs
    }})


@app.route('/api/contracts/upload', methods=['POST'])
def upload_contract_file():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    contract_id = request.form.get('contract_id', type=int)
    file_type = request.form.get('file_type')
    
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请选择文件', 'data': None})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'code': 400, 'message': '请选择文件', 'data': None})
    
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    _, ext = os.path.splitext(file.filename)
    filename = f"{contract_id}_{file_type}_{timestamp}{ext}"
    file_path = os.path.join(UPLOAD_DIR, filename)
    
    file.save(file_path)
    
    db = get_db()
    cursor = db.cursor()
    
    if file_type == 'contract':
        cursor.execute("UPDATE contracts SET contract_file_path = ? WHERE id = ?", (f"uploads/contracts/{filename}", contract_id))
    elif file_type == 'tech':
        cursor.execute("UPDATE contracts SET tech_agreement_file_path = ? WHERE id = ?", (f"uploads/contracts/{filename}", contract_id))
    
    db.commit()
    
    return jsonify({'code': 200, 'message': '文件上传成功', 'data': {'file_path': f"uploads/contracts/{filename}"}})


@app.route('/api/contracts/download/<int:contract_id>/<file_type>', methods=['GET'])
def download_contract_file(contract_id, file_type):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    if not token:
        token = request.args.get('token', '')
    payload = verify_token(token)
    if not payload:
        return jsonify({'code': 401, 'message': '登录已过期', 'data': None})
    
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM contracts WHERE id = ?", (contract_id,))
    row = cursor.fetchone()
    
    if not row:
        return jsonify({'code': 404, 'message': '合同不存在', 'data': None})
    
    row_dict = dict(row)
    
    if file_type == 'contract':
        file_path = row_dict.get('contract_file_path')
    elif file_type == 'tech':
        file_path = row_dict.get('tech_agreement_file_path')
    else:
        return jsonify({'code': 400, 'message': '无效的文件类型', 'data': None})
    
    if not file_path:
        return jsonify({'code': 404, 'message': '文件不存在', 'data': None})
    
    full_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), file_path)
    
    if not os.path.exists(full_path):
        return jsonify({'code': 404, 'message': '文件不存在', 'data': None})
    
    return send_from_directory(os.path.dirname(full_path), os.path.basename(full_path), as_attachment=False)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)